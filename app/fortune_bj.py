"""Fortune BJ order-level scheduling MVP.

This module is intentionally separate from the legacy ton/month optimizer.
All user-facing workbook sheets and columns are Chinese for the Fortune BJ
desktop workflow.
"""
from __future__ import annotations

import calendar
import math
import random
import re
import sys
import time
import unicodedata
from dataclasses import dataclass, field, replace
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo

from app.license_validator import LicenseInfo, validate_license_with_fallback
from app.version import APP_VERSION


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def _resolve_deploy_root() -> Path:
    """Locate the user-facing folder that contains 数据导入 and 报告."""
    candidates: list[Path] = []
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        candidates.extend([exe_dir, *exe_dir.parents[:3]])
    candidates.extend([Path.cwd(), PROJECT_ROOT, PROJECT_ROOT.parent])
    for candidate in candidates:
        if (candidate / "数据导入").exists() or (candidate / "报告").exists():
            return candidate
    return PROJECT_ROOT


DEPLOY_ROOT = _resolve_deploy_root()
DATA_DIR = DEPLOY_ROOT / "数据导入"
REPORT_DIR = DEPLOY_ROOT / "报告"

OPS_TEMPLATE_NAME = "生产订单工序_产能分析输入模板.csv"
DEMAND_TEMPLATE_NAME = "订单交期数量_产能分析输入模板.xlsx"
WC_TEMPLATE_NAME = "工作中心_产能分析输入模板.xlsx"
OPTIONAL_OPS_TEMPLATE_NAME = "可选工序_产能分析输入模板.xlsx"
CALENDAR_TEMPLATE_NAME = "工作日历_产能分析输入模板.xlsx"
FORECAST_TEMPLATE_NAME = "需求预测_产能分析输入模板.xlsx"
FORECAST_ROUTE_SHEET_NAME = "Data"
FORECAST_ROUTE_TEMPLATE_SPECS: tuple[tuple[str, str], ...] = (
    ("北京", "工艺路线-北京.xlsx"),
    ("沈阳", "工艺路线-沈阳.xlsx"),
    ("南通", "工艺路线-南通.xlsx"),
)

PLACEHOLDER_DUE_YEAR = 2049
OUTSOURCE_DURATION_HOURS = 7 * 24
UNMAINTAINED_WORKCENTER = "未维护工作中心"
URGENT_TYPE_PRIORITY = {
    "越库": 1,
    "T-14": 2,
    "RTM": 3,
    "临时加急": 4,
}
ALLOWED_URGENT_TYPES = tuple(URGENT_TYPE_PRIORITY)
PRIORITY_OVERDUE = 5
PRIORITY_NORMAL = 6


@dataclass(frozen=True)
class FortuneBjConfig:
    """Runtime options exposed by the Chinese desktop UI."""

    operations_path: Path
    demand_path: Path
    workcenter_path: Path
    optional_operations_path: Path | None = None
    calendar_path: Path | None = None
    forecast_path: Path | None = None
    output_dir: Path = REPORT_DIR
    schedule_mode: str = "ModeA"
    mode_b_optimize_days: int = 1
    mode_b_optimization_granularity: str = "周"
    mode_b_optimization_start_month: datetime | None = None
    mode_b_max_window_tasks: int = 2000
    mode_b_solver_max_seconds: float | None = 60.0
    enable_urgent: bool = True
    enable_forecast: bool = False
    operation_flow_mode: str = "整批流转"
    hot_surface_mode: str = "同机加逻辑"
    objective_profile: str = "默认：产能缺口最小"
    start_time: datetime | None = None
    due_date_policy: str = "同订单多交期取最早"


@dataclass(frozen=True)
class WorkCenterCapacity:
    work_center: str
    resource_group: str
    quantity: int
    calendar_name: str = "默认日历"
    daily_hours: float = 24.0
    capacity_calc_type: str = "普通工时"
    hot_surface_type: str = "普通"
    batch_capacity: float = 0.0
    capacity_unit: str = "件"
    default_unit_capacity: float = 1.0
    batch_cycle_hours: float = 0.0
    setup_hours: float = 0.0
    allow_partial_batch: bool = True
    min_batch_fill_rate: float = 0.0
    line_throughput_rate: float = 0.0
    throughput_unit: str = "件/小时"
    residence_hours: float = 0.0
    changeover_hours: float = 0.0
    process_groups: str = ""


@dataclass(frozen=True)
class WorkCalendar:
    name: str
    daily_hours: float
    weekly_workdays: float = 7.0

    @property
    def average_daily_hours(self) -> float:
        return max(self.daily_hours, 0.0) * max(self.weekly_workdays, 0.0) / 7.0


@dataclass(frozen=True)
class OptionalOperation:
    material: str
    activity: str
    alternative_work_center: str
    alternative_resource_group: str
    unit_hours: float
    unit_hours_source: str = "可选工序表填写"
    is_outsource: bool = False
    priority_rank: int = 999
    capacity_calc_type: str = ""
    process_group: str = ""
    unit_capacity: float = 0.0
    outsource_return_days: float = 7.0


@dataclass
class OperationTask:
    order_id: str
    activity: float
    material: str
    process_text: str
    work_center: str
    resource_group: str
    quantity: float
    unit_hours: float
    duration_hours: float
    due_date: datetime
    urgent: bool = False
    manual_urgent: bool = False
    urgent_type: str = ""
    overdue: bool = False
    adjusted_to_start_period: bool = False
    original_due_date: datetime | None = None
    priority_rank: int = PRIORITY_NORMAL
    priority_type: str = "普通订单"
    priority_reason: str = "普通订单"
    demand_source: str = "真实订单"
    forecast_month: str = ""
    forecast_week_end: str = ""
    forecast_drawing: str = ""
    forecast_route_source: str = ""
    route_source_order: str = ""
    is_outsource: bool = False
    missing_work_center: bool = False
    is_hot_surface: bool = False
    source_row: int | None = None
    capacity_calc_type: str = ""
    hot_surface_type: str = ""
    process_group: str = ""
    unit_capacity: float = 1.0
    allow_batch_mix: bool = True
    must_same_batch: bool = False
    treatment_program: str = ""


@dataclass
class ScheduledOperation:
    task: OperationTask
    start: datetime
    end: datetime
    on_time: bool
    tardy_hours: float
    note: str = ""
    analysis_status: str = "无限产能估算"
    analysis_source: str = "ModeA"
    window_number: int | None = None
    window_type: str = "窗口外"


@dataclass(frozen=True)
class ModeBAllocation:
    period: str
    period_start: datetime
    period_end: datetime
    source_item: ScheduledOperation
    quantity: int
    destination_type: str
    destination_work_center: str
    destination_resource_group: str
    unit_hours: float
    unit_hours_source: str
    load_hours: float
    original_unit_hours: float
    original_released_hours: float
    extra_hours: float
    unmaintained_load_hours: float = 0.0
    is_outsource: bool = False
    is_unmaintained_work_center: bool = False
    capacity_calc_type: str = "普通工时"
    hot_surface_type: str = "普通"
    process_group: str = ""
    unit_capacity: float = 1.0
    batch_capacity: float = 0.0
    batch_cycle_hours: float = 0.0
    batch_count: float = 0.0
    line_throughput_rate: float = 0.0
    residence_hours: float = 0.0
    capacity_load_units: float = 0.0


@dataclass(frozen=True)
class ModeBAllocationPeriodSegment:
    allocation: ModeBAllocation
    period: str
    period_start: datetime
    period_end: datetime
    load_hours: float
    original_load_hours: float
    original_released_hours: float
    extra_hours: float
    unmaintained_load_hours: float
    capacity_load_units: float


@dataclass
class ScheduleResult:
    scheduled: list[ScheduledOperation] = field(default_factory=list)
    outsource: list[ScheduledOperation] = field(default_factory=list)
    missing_mapping: list[dict[str, Any]] = field(default_factory=list)
    placeholder_due_orders: list[dict[str, Any]] = field(default_factory=list)
    bottleneck_report: list[dict[str, Any]] = field(default_factory=list)
    window_report: list[dict[str, Any]] = field(default_factory=list)
    optional_operation_report: list[dict[str, Any]] = field(default_factory=list)
    order_operation_allocation_report: list[dict[str, Any]] = field(default_factory=list)
    capacity_optimization_summary: list[dict[str, Any]] = field(default_factory=list)
    capacity_recommendation_report: list[dict[str, Any]] = field(default_factory=list)
    capacity_optimization_stats: list[dict[str, Any]] = field(default_factory=list)
    monthly_capacity_report: list[dict[str, Any]] = field(default_factory=list)
    hot_surface_capacity_report: list[dict[str, Any]] = field(default_factory=list)
    unmaintained_workcenter_report: list[dict[str, Any]] = field(default_factory=list)
    input_maintenance_report: list[dict[str, Any]] = field(default_factory=list)
    data_issues: list[dict[str, Any]] = field(default_factory=list)
    report_path: Path | None = None


def _validate_analysis_inputs(
    *,
    config: FortuneBjConfig,
    demand_by_order: dict[str, dict[str, Any]],
    capacities: dict[str, WorkCenterCapacity],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Return blocking pre-analysis issues before any capacity calculation starts."""
    operation_df = _read_workbook_first_sheet(config.operations_path)
    required = ["订单", "活动", "物料", "工序短文本", "标准值1", "标准值2", "标准值3"]
    issues = _missing_columns_issue(operation_df, required, config.operations_path)
    if issues:
        summary_rows = [{
            "校验项": "生产订单工序文件必填列",
            "状态": "失败",
            "问题数": len(issues),
            "说明": "生产订单工序文件缺少必填列，正式产能分析已停止。",
        }]
        return summary_rows, [], [], [], issues

    normalized_orders: list[str | None] = [_normalize_order(value) for value in operation_df["订单"]]
    operation_orders = {order for order in normalized_orders if order}
    demand_orders = set(demand_by_order)
    forecast_orders = {
        order_id
        for order_id, demand in demand_by_order.items()
        if str(demand.get("需求来源") or "") == "预测需求"
    }

    missing_order_rows: list[dict[str, Any]] = []
    for order_id in sorted((demand_orders - forecast_orders) - operation_orders):
        demand = demand_by_order.get(order_id, {})
        due_date = demand.get("交期")
        original_due = demand.get("原始最早交期")
        missing_order_rows.append({
            "订单": order_id,
            "订单数量": demand.get("数量", ""),
            "需求日期": due_date.strftime("%Y-%m-%d") if isinstance(due_date, datetime) else "",
            "原始最早交期": original_due.strftime("%Y-%m-%d") if isinstance(original_due, datetime) else "",
            "紧急类型": demand.get("紧急类型", ""),
            "是否紧急": "是" if demand.get("紧急") else "否",
            "问题类型": "订单缺失工序",
            "处理建议": "请在生产订单工序文件中补充该订单的完整工序，或从订单交期数量产能分析输入文件中移除不参与分析的订单。",
        })

    missing_workcenter_rows: list[dict[str, Any]] = []
    capacity_names = set(capacities)
    total_rows = len(operation_df)
    started_at = time.perf_counter()
    for idx, row in operation_df.iterrows():
        order_id = normalized_orders[idx] if idx < len(normalized_orders) else _normalize_order(row.get("订单"))
        if not order_id or order_id not in demand_orders:
            continue
        process_text = _clean_text(row.get("工序短文本"))
        if not process_text or "外协" in process_text or process_text in capacity_names:
            continue
        demand = demand_by_order.get(order_id, {})
        unit_hours = sum(_to_number(row.get(col), default=0.0) for col in ("标准值1", "标准值2", "标准值3"))
        quantity = float(demand.get("数量") or 0.0)
        missing_workcenter_rows.append({
            "订单": order_id,
            "活动": _activity_key(row.get("活动")),
            "物料": _clean_text(row.get("物料")),
            "工序短文本": process_text,
            "订单数量": quantity,
            "标准值1": row.get("标准值1", ""),
            "标准值2": row.get("标准值2", ""),
            "标准值3": row.get("标准值3", ""),
            "单位工时(小时/pcs)": round(unit_hours, 4),
            "工序生产时间(小时)": round(quantity * unit_hours, 4),
            "源文件行号": idx + 2,
            "问题类型": "工序缺失工作中心",
            "处理建议": "请在工作中心文件中补充该工序短文本对应的工作中心、资源组分类、设备数量和日历名称。",
        })
        if (idx + 1) % 5000 == 0:
            _emit_progress(
                None,
                "分析前数据校验",
                idx + 1,
                total_rows,
                started_at,
                f"缺失工作中心 {len(missing_workcenter_rows):,}",
            )

    missing_workcenter_summary = _summarize_missing_workcenters(missing_workcenter_rows)
    summary_rows = [{
        "校验项": "订单交期数量订单是否都有生产订单工序",
        "状态": "通过" if not missing_order_rows else "失败",
        "问题数": len(missing_order_rows),
        "说明": "所有有效需求订单均已在生产订单工序文件中找到。" if not missing_order_rows else "存在订单缺失生产订单工序，正式产能分析已停止。",
    }, {
        "校验项": "生产订单工序是否都有工作中心映射",
        "状态": "通过" if not missing_workcenter_rows else "提醒",
        "问题数": len(missing_workcenter_rows),
        "说明": "所有参与分析的工序短文本均已在工作中心文件中找到。" if not missing_workcenter_rows else "存在工序短文本未定义工作中心；正式分析继续，这些工序作为未维护工作中心负荷单独报告，不进入工作中心产能优化。",
    }]
    return summary_rows, missing_order_rows, missing_workcenter_summary, missing_workcenter_rows, []


def _summarize_missing_workcenters(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        process_text = _clean_text(row.get("工序短文本")) or "空工序短文本"
        item = grouped.setdefault(process_text, {
            "工序短文本": process_text,
            "缺失工序行数": 0,
            "涉及订单数": set(),
            "涉及总工时": 0.0,
            "示例订单": row.get("订单", ""),
            "示例活动": row.get("活动", ""),
            "示例物料": row.get("物料", ""),
            "示例源文件行号": row.get("源文件行号", ""),
            "处理建议": row.get("处理建议", ""),
        })
        item["缺失工序行数"] += 1
        item["涉及订单数"].add(str(row.get("订单", "")))
        item["涉及总工时"] += _to_number(row.get("工序生产时间(小时)"), default=0.0)
    summary: list[dict[str, Any]] = []
    for item in grouped.values():
        summary.append({
            "工序短文本": item["工序短文本"],
            "缺失工序行数": item["缺失工序行数"],
            "涉及订单数": len(item["涉及订单数"]),
            "涉及总工时": round(item["涉及总工时"], 4),
            "示例订单": item["示例订单"],
            "示例活动": item["示例活动"],
            "示例物料": item["示例物料"],
            "示例源文件行号": item["示例源文件行号"],
            "处理建议": item["处理建议"],
        })
    return sorted(summary, key=lambda row: (-int(row["缺失工序行数"]), row["工序短文本"]))


def _write_precheck_report(
    config: FortuneBjConfig,
    summary_rows: list[dict[str, Any]],
    missing_order_rows: list[dict[str, Any]],
    missing_workcenter_summary: list[dict[str, Any]],
    missing_workcenter_rows: list[dict[str, Any]],
    blocking_issue_rows: list[dict[str, Any]] | None = None,
) -> Path:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    path = config.output_dir / f"Fortune_BJ_分析前数据校验报告_{timestamp}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    _write_dict_sheet(wb, "校验摘要", summary_rows)
    _write_dict_sheet(wb, "数据校验问题", blocking_issue_rows or [])
    _write_dict_sheet(wb, "缺失工序订单", missing_order_rows)
    _write_dict_sheet(wb, "缺失工作中心汇总", missing_workcenter_summary)
    _write_dict_sheet(wb, "缺失工作中心明细", missing_workcenter_rows)
    _autosize_workbook(wb)
    wb.save(path)
    return path


def _blocking_demand_issues(issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    blocking_types = {"缺少字段", "紧急类型异常"}
    return [row for row in issues if str(row.get("类型") or "") in blocking_types]


def _blocking_optional_issues(issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    blocking_types = {"缺少字段", "可选工序缺少可继承工时"}
    return [row for row in issues if str(row.get("类型") or "") in blocking_types]


def ensure_runtime_dirs() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    (DEPLOY_ROOT / "licenses" / "active").mkdir(parents=True, exist_ok=True)
    (DEPLOY_ROOT / "licenses" / "requests").mkdir(parents=True, exist_ok=True)


def generate_input_templates(project_stuff_dir: Path | None = None, output_dir: Path = DATA_DIR) -> dict[str, Path]:
    raise RuntimeError(
        "生成/刷新数据导入模板功能已停用。当前 数据导入 文件夹就是正式数据源，工具运行时只读取，不再从 Project_Stuff 抽取或覆盖输入文件。"
    )


ProgressCallback = Callable[[str], None]


def _format_duration(seconds: float | None) -> str:
    if seconds is None or math.isinf(seconds) or math.isnan(seconds):
        return "计算中"
    seconds = max(float(seconds), 0.0)
    if seconds < 60:
        return f"{seconds:.0f}秒"
    minutes, sec = divmod(int(seconds), 60)
    if minutes < 60:
        return f"{minutes}分{sec:02d}秒"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}小时{minutes:02d}分"


def _emit_progress(
    progress: ProgressCallback | None,
    phase: str,
    current: int,
    total: int,
    started_at: float,
    detail: str = "",
) -> None:
    if progress is None:
        return
    current = max(int(current), 0)
    total = max(int(total), 0)
    elapsed = time.perf_counter() - started_at
    if total <= 0:
        progress(f"{phase}: {detail} | 已用 {_format_duration(elapsed)}")
        return
    current = min(current, total)
    percent = current / total * 100
    eta = (elapsed / current * (total - current)) if current > 0 else None
    suffix = f" | {detail}" if detail else ""
    progress(
        f"{phase}: {current:,}/{total:,} ({percent:.1f}%) | "
        f"已用 {_format_duration(elapsed)} | 预计剩余 {_format_duration(eta)}{suffix}"
    )


def run_fortune_bj_schedule(config: FortuneBjConfig, progress: ProgressCallback | None = None) -> ScheduleResult:
    def notify(message: str) -> None:
        if progress is not None:
            progress(message)

    notify("校验离线授权...")
    license_info = validate_license_with_fallback(
        primary_root=str(DEPLOY_ROOT),
        fallback_roots=[str(PROJECT_ROOT)],
    )
    notify(f"授权有效：{license_info.license_id} / {license_info.customer_name} / 到期 {license_info.expiry_date}")

    notify("读取工作日历...")
    calendars, calendar_issues = load_work_calendars(config.calendar_path)
    notify(f"已读取工作日历 {len(calendars)} 个，问题 {len(calendar_issues)} 条。")
    notify("读取工作中心...")
    capacities, capacity_issues = load_workcenter_capacities(config.workcenter_path, calendars)
    notify(f"已读取工作中心 {len(capacities)} 个，问题 {len(capacity_issues)} 条。")
    notify("读取可选工序...")
    operation_unit_hours = _build_operation_unit_hours_lookup(config.operations_path)
    optional_operations, optional_issues = load_optional_operations(
        config.optional_operations_path,
        capacities,
        operation_unit_hours=operation_unit_hours,
    )
    notify(f"已读取可选工序 {len(optional_operations)} 条，问题 {len(optional_issues)} 条。")
    notify("读取订单交期数量...")
    demand_by_order, demand_issues, placeholder_due_orders = load_order_demand(config.demand_path, config)
    notify(f"已读取有效订单需求 {len(demand_by_order)} 个，问题/调整 {len(demand_issues)} 条。")
    forecast_issues: list[dict[str, Any]] = []
    if config.enable_forecast:
        notify("读取需求预测...")
        forecast_demand_by_order, forecast_issues = load_forecast_demand(config)
        demand_by_order.update(forecast_demand_by_order)
        notify(f"已生成预测虚拟订单 {len(forecast_demand_by_order)} 个，问题 {len(forecast_issues)} 条。")
    else:
        notify("需求预测导入未启用，已忽略需求预测表。")
    blocking_demand_issue_rows = _blocking_demand_issues(demand_issues)
    blocking_optional_issue_rows = _blocking_optional_issues(optional_issues)
    blocking_forecast_issue_rows = _blocking_forecast_issues(forecast_issues)
    if placeholder_due_orders:
        placeholder_order_count = len({row["订单"] for row in placeholder_due_orders})
        notify(
            f"识别占位交期订单 {placeholder_order_count} 个、需求行 {len(placeholder_due_orders)} 条，"
            "已排除出产能分析计算。"
        )
    notify("执行分析前数据完整性校验...")
    precheck_summary, missing_order_rows, missing_workcenter_summary, missing_workcenter_rows, blocking_issue_rows = _validate_analysis_inputs(
        config=config,
        demand_by_order=demand_by_order,
        capacities=capacities,
    )
    if blocking_demand_issue_rows:
        precheck_summary.insert(0, {
            "校验项": "订单交期数量产能分析输入文件",
            "状态": "失败",
            "问题数": len(blocking_demand_issue_rows),
            "说明": "订单交期数量产能分析输入文件存在阻断问题，正式产能分析已停止。",
        })
        blocking_issue_rows = [*blocking_demand_issue_rows, *blocking_issue_rows]
    if blocking_optional_issue_rows:
        precheck_summary.insert(0, {
            "校验项": "可选工序输入文件",
            "状态": "失败",
            "问题数": len(blocking_optional_issue_rows),
            "说明": "可选工序存在阻断问题，正式产能分析已停止。",
        })
        blocking_issue_rows = [*blocking_optional_issue_rows, *blocking_issue_rows]
    if blocking_forecast_issue_rows:
        precheck_summary.insert(0, {
            "校验项": "需求预测输入文件",
            "状态": "失败",
            "问题数": len(blocking_forecast_issue_rows),
            "说明": "需求预测输入文件存在阻断问题，正式产能分析已停止。",
        })
        blocking_issue_rows = [*blocking_forecast_issue_rows, *blocking_issue_rows]
    precheck_failed = any(row.get("状态") == "失败" for row in precheck_summary)
    if precheck_failed or missing_order_rows or blocking_issue_rows:
        precheck_report = _write_precheck_report(
            config,
            precheck_summary,
            missing_order_rows,
            missing_workcenter_summary,
            missing_workcenter_rows,
            blocking_issue_rows,
        )
        raise RuntimeError(
            "分析前数据校验未通过，已停止正式产能分析。\n"
            f"数据结构问题：{len(blocking_issue_rows)} 条。\n"
            f"缺失工序订单：{len(missing_order_rows)} 个。\n"
            f"未维护工作中心工序：{len(missing_workcenter_rows)} 行；"
            f"未维护工作中心类型：{len(missing_workcenter_summary)} 个。\n"
            f"校验报告：{precheck_report}"
        )
    if missing_workcenter_rows:
        notify(
            f"分析前数据完整性校验通过：订单工序完整；发现未维护工作中心工序 "
            f"{len(missing_workcenter_rows):,} 行，将单独报告且不进入工作中心产能优化。"
        )
    else:
        notify("分析前数据完整性校验通过：订单工序和工作中心映射完整。")
    notify("读取生产订单工序并匹配资源映射...")
    tasks, task_issues, missing_mapping = load_operation_tasks(
        config.operations_path,
        capacities,
        demand_by_order,
        config=config,
        progress=progress,
    )
    notify(f"可分析工序 {len(tasks)} 条，缺失映射工序 {len(missing_mapping)} 条。")

    result = ScheduleResult()
    result.data_issues.extend(calendar_issues)
    result.data_issues.extend(capacity_issues)
    result.data_issues.extend(optional_issues)
    result.data_issues.extend(demand_issues)
    result.data_issues.extend(forecast_issues)
    result.data_issues.extend(task_issues)
    result.missing_mapping.extend(missing_mapping)
    result.placeholder_due_orders.extend(placeholder_due_orders)

    mode = _normalize_mode(config.schedule_mode)
    notify(f"执行 {mode} 产能分析计算...")
    scheduled_all = _schedule_tasks(
        tasks,
        capacities,
        mode=mode,
        optional_operations=optional_operations,
        result=result,
        config=config,
        progress=progress,
    )
    notify(f"产能分析计算完成，生成工序产能占用结果 {len(scheduled_all)} 条。")
    order_completion: dict[str, datetime] = {}
    for item in scheduled_all:
        order_completion[item.task.order_id] = max(order_completion.get(item.task.order_id, datetime.min), item.end)

    # Recompute order-level tardiness after outsource delays are known.
    for item in scheduled_all:
        final_end = order_completion.get(item.task.order_id, item.end)
        item.on_time = final_end <= item.task.due_date
        item.tardy_hours = max((final_end - item.task.due_date).total_seconds() / 3600.0, 0.0)

    result.scheduled = [item for item in scheduled_all if not item.task.is_outsource]
    result.outsource = [item for item in scheduled_all if item.task.is_outsource]
    result.unmaintained_workcenter_report = _build_unmaintained_workcenter_report(scheduled_all, config=config)
    result.input_maintenance_report = _build_input_maintenance_report(config)
    notify("写入 Excel 报告...")
    result.report_path = write_report(result, config, order_completion, license_info=license_info, progress=progress)
    notify(f"Excel 报告写入完成：{result.report_path}")
    return result


def load_work_calendars(path: Path | None) -> tuple[dict[str, WorkCalendar], list[dict[str, Any]]]:
    default = {"默认日历": WorkCalendar("默认日历", 24.0, 7.0)}
    if path is None or not Path(path).exists():
        return default, []
    df = _read_workbook_first_sheet(Path(path))
    required = ["日历名称", "每日工作小时", "每周工作天数"]
    issues = _missing_columns_issue(df, required, Path(path))
    if issues:
        return default, issues
    calendars: dict[str, WorkCalendar] = {}
    for idx, row in df.iterrows():
        name = _clean_text(row.get("日历名称"))
        daily_hours = _to_number(row.get("每日工作小时"), default=0)
        weekly_workdays = _to_number(row.get("每周工作天数"), default=7)
        if not name or daily_hours <= 0 or weekly_workdays <= 0:
            issues.append({"类型": "工作日历异常", "文件": str(path), "行号": idx + 2, "说明": "日历名称为空或工作时间小于等于0"})
            continue
        calendars[name] = WorkCalendar(name, daily_hours, min(weekly_workdays, 7.0))
    calendars.setdefault("默认日历", default["默认日历"])
    return calendars, issues


def load_workcenter_capacities(
    path: Path,
    calendars: dict[str, WorkCalendar] | None = None,
) -> tuple[dict[str, WorkCenterCapacity], list[dict[str, Any]]]:
    df = _read_workbook_first_sheet(path)
    required = ["工作中心", "资源组分类", "数量", "日历名称"]
    issues = _missing_columns_issue(df, required, path)
    if issues:
        missing = [column for column in required if column not in df.columns]
        raise ValueError(f"工作中心表缺少必填列：{', '.join(missing)}。请检查 {path.name}。")
    calendars = calendars or {"默认日历": WorkCalendar("默认日历", 24.0, 7.0)}
    capacities: dict[str, WorkCenterCapacity] = {}
    missing_calendar_rows: list[int] = []
    for idx, row in df.iterrows():
        wc = _clean_text(row.get("工作中心"))
        group = _clean_text(row.get("资源组分类"))
        qty = _to_number(row.get("数量"), default=0)
        calendar_name = _clean_text(row.get("日历名称"))
        work_calendar = calendars.get(calendar_name) if calendar_name else None
        capacity_calc_type = _normalize_capacity_calc_type(row.get("产能计算类型"))
        hot_surface_type = _hot_surface_type_from_text(row.get("热处表处类型"), group, wc)
        if not wc or qty <= 0:
            issues.append({"类型": "工作中心映射异常", "文件": str(path), "行号": idx + 2, "说明": "工作中心为空或数量小于等于0"})
            continue
        if work_calendar is None:
            missing_calendar_rows.append(idx + 2)
            issues.append({"类型": "工作中心日历异常", "文件": str(path), "行号": idx + 2, "说明": f"日历名称未在工作日历中定义：{calendar_name or '(空)'}"})
            continue
        capacities[wc] = WorkCenterCapacity(
            wc,
            group,
            max(int(math.floor(qty)), 1),
            work_calendar.name,
            work_calendar.average_daily_hours,
            capacity_calc_type,
            hot_surface_type,
            _to_number(row.get("单炉容量"), default=0.0),
            _clean_text(row.get("容量单位")) or "件",
            _positive_or_default(_to_number(row.get("单件容量占用默认值"), default=1.0), 1.0),
            _to_number(row.get("单炉周期小时"), default=0.0),
            _to_number(row.get("装卸/准备小时"), default=0.0),
            _to_bool(row.get("是否允许不足炉开炉")) if "是否允许不足炉开炉" in df.columns else True,
            _to_number(row.get("最低开炉率"), default=0.0),
            _to_number(row.get("流水线吞吐率"), default=0.0),
            _clean_text(row.get("吞吐率单位")) or "件/小时",
            _to_number(row.get("单件在炉时间小时"), default=0.0),
            _to_number(row.get("换型时间小时"), default=0.0),
            _clean_text(row.get("可处理工艺组")),
        )
    if missing_calendar_rows:
        preview = ", ".join(str(row) for row in missing_calendar_rows[:20])
        more = "..." if len(missing_calendar_rows) > 20 else ""
        raise ValueError(
            f"工作中心表有 {len(missing_calendar_rows)} 行日历名称没有匹配到工作日历。"
            f"缺失行号：{preview}{more}。请检查 {path.name} 与 {CALENDAR_TEMPLATE_NAME}。"
        )
    return capacities, issues


def _build_operation_unit_hours_lookup(path: Path) -> dict[tuple[str, str], float]:
    df = _read_workbook_first_sheet(path)
    required = ["物料", "活动", "标准值1", "标准值2", "标准值3"]
    if _missing_columns_issue(df, required, path):
        return {}
    lookup: dict[tuple[str, str], float] = {}
    for _idx, row in df.iterrows():
        material = _clean_text(row.get("物料"))
        activity = _activity_key(row.get("活动"))
        if not material or not activity:
            continue
        unit_hours = sum(_to_number(row.get(col), default=0.0) for col in ("标准值1", "标准值2", "标准值3"))
        if unit_hours <= 0:
            continue
        key = (material, activity)
        lookup[key] = max(float(unit_hours), lookup.get(key, 0.0))
    return lookup


def load_optional_operations(
    path: Path | None,
    capacities: dict[str, WorkCenterCapacity],
    *,
    operation_unit_hours: dict[tuple[str, str], float] | None = None,
) -> tuple[dict[tuple[str, str], list[OptionalOperation]], list[dict[str, Any]]]:
    if path is None or not Path(path).exists():
        return {}, []
    df = _read_workbook_first_sheet(Path(path))
    required = ["物料", "活动", "可选工作中心", "可选资源组分类", "可选单位工时(小时/pcs)"]
    issues = _missing_columns_issue(df, required, Path(path))
    if issues:
        return {}, issues
    options: dict[tuple[str, str], list[OptionalOperation]] = {}
    for idx, row in df.iterrows():
        material = _clean_text(row.get("物料"))
        activity = _activity_key(row.get("活动"))
        work_center = _clean_text(row.get("可选工作中心"))
        resource_group = _clean_text(row.get("可选资源组分类"))
        unit_hours_raw = row.get("可选单位工时(小时/pcs)")
        unit_hours = _to_number(unit_hours_raw, default=0.0)
        unit_hours_source = "可选工序表填写"
        priority_text = _clean_text(row.get("工序优先级"))
        priority_rank = _priority_rank(priority_text)
        is_outsource = resource_group == "外包"
        capacity_calc_type = _clean_text(row.get("可选产能计算类型"))
        process_group = _clean_text(row.get("可选工艺兼容组"))
        unit_capacity = _to_number(row.get("可选单件容量占用"), default=0.0)
        outsource_return_days = _positive_or_default(_to_number(row.get("外包返回天数"), default=7.0), 7.0)
        if is_outsource:
            work_center = "外包"
            unit_hours = 0.0
            unit_hours_source = "外包"
        if not material or not activity or (not is_outsource and not work_center):
            issues.append({"类型": "可选工序异常", "文件": str(path), "行号": idx + 2, "说明": "物料、活动或可选工作中心无效；外包行以可选资源组分类=外包判断，且可不填单位工时"})
            continue
        if not is_outsource and work_center not in capacities:
            issues.append({"类型": "可选工序异常", "文件": str(path), "行号": idx + 2, "说明": f"可选工作中心未在工作中心映射中定义：{work_center}"})
            continue
        if not resource_group and work_center in capacities:
            resource_group = capacities[work_center].resource_group
        if not is_outsource and _is_blank_cell(unit_hours_raw):
            inherited_hours = (operation_unit_hours or {}).get((material, activity), 0.0)
            if inherited_hours <= 0:
                issues.append({
                    "类型": "可选工序缺少可继承工时",
                    "文件": str(path),
                    "行号": idx + 2,
                    "物料": material,
                    "活动": activity,
                    "可选工作中心": work_center,
                    "说明": "可选单位工时为空，且生产订单工序中找不到同物料+活动的有效单位工时，无法自动继承。",
                    "处理建议": "请在可选工序表填写可选单位工时，或在生产订单工序表补充该物料+活动的标准值1/2/3。",
                })
                continue
            unit_hours = inherited_hours
            unit_hours_source = "从生产订单工序继承(同物料+活动最大单位工时)"
            issues.append({
                "类型": "可选工序工时自动继承",
                "文件": str(path),
                "行号": idx + 2,
                "物料": material,
                "活动": activity,
                "可选工作中心": work_center,
                "继承单位工时": round(unit_hours, 4),
                "工时来源": unit_hours_source,
                "说明": "可选单位工时为空，已从生产订单工序表按同物料+活动取最大单位工时自动继承。",
            })
        if not is_outsource and unit_hours <= 0:
            issues.append({
                "类型": "可选工序异常",
                "文件": str(path),
                "行号": idx + 2,
                "物料": material,
                "活动": activity,
                "可选工作中心": work_center,
                "说明": "可选单位工时必须大于0；如需自动继承原工时，请将该单元格留空，不要填写0。",
            })
            continue
        options.setdefault((material, activity), []).append(OptionalOperation(
            material=material,
            activity=activity,
            alternative_work_center=work_center,
            alternative_resource_group=resource_group or "外包",
            unit_hours=unit_hours,
            unit_hours_source=unit_hours_source,
            is_outsource=is_outsource,
            priority_rank=priority_rank,
            capacity_calc_type=capacity_calc_type,
            process_group=process_group,
            unit_capacity=unit_capacity,
            outsource_return_days=outsource_return_days,
        ))
    return options, issues


def load_order_demand(path: Path, config: FortuneBjConfig) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    df = _read_workbook_first_sheet(path)
    required = ["富创单据号", "华创需求数量", "供给日期", "紧急类型"]
    issues = _missing_columns_issue(df, required, path)
    if issues:
        return {}, issues, []
    scientific_order_rows = [
        idx + 2
        for idx, value in df["富创单据号"].items()
        if _looks_like_scientific_order_id(value)
    ]
    if scientific_order_rows:
        preview = ", ".join(str(row_number) for row_number in scientific_order_rows[:20])
        more = "..." if len(scientific_order_rows) > 20 else ""
        raise ValueError(
            f"{Path(path).name} 的 A列/富创单据号 已被保存为科学计数法，"
            f"例如行号 {preview}{more}。这种格式已经丢失原始订单号精度，不能继续用于产能分析。"
            "请从备份或原始系统重新恢复该 CSV，并避免直接用 Excel 打开后保存 CSV。"
        )

    demand: dict[str, dict[str, Any]] = {}
    placeholder_due_orders: list[dict[str, Any]] = []
    optimization_start = _optimization_start_period(config)
    for idx, row in df.iterrows():
        order_id = _normalize_order(row.get("富创单据号"))
        if not order_id:
            continue
        quantity = _to_number(row.get("华创需求数量"), default=0.0)
        due_date = _parse_datetime(row.get("供给日期"))
        if quantity <= 0 or due_date is None:
            issues.append({"类型": "订单需求异常", "文件": str(path), "行号": idx + 2, "说明": "数量小于等于0或供给日期无效"})
            continue
        source_urgent_type = _normalize_urgent_type(row.get("紧急类型"))
        if source_urgent_type and source_urgent_type not in URGENT_TYPE_PRIORITY:
            issues.append({
                "类型": "紧急类型异常",
                "文件": str(path),
                "行号": idx + 2,
                "订单": order_id,
                "当前填写值": source_urgent_type,
                "允许填写值": _allowed_urgent_types_text(),
                "说明": "紧急类型只允许填写标准值或留空；请修正后重新运行产能分析。",
            })
            continue
        urgent_type = source_urgent_type if config.enable_urgent else ""
        manual_urgent = bool(urgent_type)
        if _is_placeholder_due_date(due_date):
            placeholder_due_orders.append({
                "订单": order_id,
                "数量": quantity,
                "占位交期": due_date,
                "紧急类型": urgent_type,
                "是否紧急": bool(urgent_type),
                "源文件行号": idx + 2,
                "说明": f"{due_date.year} 年交期视为占位日期，不参与产能分析计算",
            })
            continue
        original_due_date = due_date
        adjusted_to_start_period = False
        overdue = False
        if optimization_start is not None and due_date < optimization_start:
            overdue = True
            adjusted_to_start_period = True
            issues.append({
                "类型": "过期订单转入优化开始日期",
                "文件": str(path),
                "行号": idx + 2,
                "订单": order_id,
                "原交期": original_due_date.strftime("%Y-%m-%d"),
                "优化粒度": _optimization_granularity(config),
                "优化开始日期": optimization_start.strftime("%Y-%m-%d"),
                "优化开始日期所在周期": _period_label_from_start(optimization_start, config),
                "优化开始日期所在周期跨度": _period_span_from_bounds(
                    optimization_start,
                    _optimization_period_end(optimization_start, config),
                ),
                "说明": "原交期早于优化开始日期；工具仍按原交期倒排，之后将整单工序链平移到优化开始日期起算，并按过期订单优先级处理",
            })
        priority_rank, priority_type, priority_reason = _priority_from_flags(urgent_type, overdue)
        existing = demand.get(order_id)
        if existing is None:
            demand[order_id] = {
                "订单": order_id,
                "数量": quantity,
                "交期": due_date,
                "紧急": bool(urgent_type),
                "手动紧急": manual_urgent,
                "紧急类型": urgent_type,
                "过期": overdue,
                "原始最早交期": original_due_date,
                "已转入优化开始周期": adjusted_to_start_period,
                "订单优先级": priority_rank,
                "优先级类型": priority_type,
                "优先级原因": priority_reason,
                "需求来源": "真实订单",
            }
            continue
        existing["数量"] += quantity
        existing["交期"] = min(existing["交期"], due_date)
        merged_urgent_type = _higher_priority_urgent_type(str(existing.get("紧急类型") or ""), urgent_type)
        existing["紧急类型"] = merged_urgent_type
        existing["手动紧急"] = bool(merged_urgent_type)
        existing["过期"] = bool(existing.get("过期") or overdue)
        existing["紧急"] = bool(merged_urgent_type)
        existing["原始最早交期"] = min(existing.get("原始最早交期", original_due_date), original_due_date)
        existing["已转入优化开始周期"] = bool(existing.get("已转入优化开始周期") or adjusted_to_start_period)
        rank, priority_type, priority_reason = _priority_from_flags(
            merged_urgent_type,
            bool(existing.get("过期")),
        )
        existing["订单优先级"] = rank
        existing["优先级类型"] = priority_type
        existing["优先级原因"] = priority_reason
    return demand, issues, placeholder_due_orders


def load_forecast_demand(config: FortuneBjConfig) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    if not config.enable_forecast:
        return {}, []
    path = config.forecast_path
    if path is None or not str(path).strip() or str(path).strip() == ".":
        return {}, [{
            "类型": "需求预测文件缺失",
            "文件": "",
            "说明": "已勾选需求预测导入，但未填写物料需求预测表路径。",
        }]
    path = Path(path)
    if not path.exists():
        return {}, [{
            "类型": "需求预测文件缺失",
            "文件": str(path),
            "说明": "已勾选需求预测导入，但物料需求预测表文件不存在。",
        }]

    try:
        df = _read_workbook_sheet(path, "清单")
    except ValueError as exc:
        return {}, [{
            "类型": "需求预测清单缺失",
            "文件": str(path),
            "说明": str(exc),
        }]

    required = ["物料编码"]
    issues = _missing_columns_issue(df, required, path)
    if issues:
        return {}, issues

    route_templates, route_issues = _load_forecast_route_templates()
    issues.extend(route_issues)
    material_col = "物料编码"
    drawing_col = "图号" if "图号" in df.columns else ""
    month_columns: list[tuple[Any, int, int]] = []
    for column in list(df.columns)[2:]:
        month = _parse_forecast_month_header(column)
        if month is None:
            if df[column].map(lambda value: bool(_clean_text(value))).any():
                issues.append({
                    "类型": "需求预测月份列异常",
                    "文件": str(path),
                    "列名": str(column),
                    "说明": "预测月份列必须使用 YYYYMM、YYYY-M 或可识别的月份日期格式。",
                })
            continue
        month_columns.append((column, month[0], month[1]))

    if not month_columns:
        issues.append({
            "类型": "需求预测月份列异常",
            "文件": str(path),
            "说明": "清单中没有找到可识别的月份预测列。",
        })

    demand: dict[str, dict[str, Any]] = {}
    missing_route_materials: set[str] = set()
    optimization_start = _optimization_start_period(config)
    for idx, row in df.iterrows():
        material = _clean_text(row.get(material_col))
        if not material:
            continue
        drawing = _clean_text(row.get(drawing_col)) if drawing_col else ""
        route_template = route_templates.get(material)
        route_source_order = str(route_template.get("source_order") or "") if route_template else ""
        forecast_route_source = str(route_template.get("route_source") or "") if route_template else ""
        has_positive_forecast = False
        for column, year, month in month_columns:
            raw_value = row.get(column)
            if _is_blank_cell(raw_value):
                continue
            quantity_value = _to_number(raw_value, default=0.0)
            if quantity_value < 0:
                issues.append({
                    "类型": "需求预测数量异常",
                    "文件": str(path),
                    "行号": idx + 2,
                    "物料": material,
                    "预测月份": f"{year:04d}-{month:02d}",
                    "当前值": raw_value,
                    "说明": "预测数量不能小于0。",
                })
                continue
            if quantity_value <= 0:
                continue
            has_positive_forecast = True
            quantity_int = int(round(quantity_value))
            if abs(quantity_value - quantity_int) > 1e-6:
                issues.append({
                    "类型": "需求预测数量异常",
                    "文件": str(path),
                    "行号": idx + 2,
                    "物料": material,
                    "预测月份": f"{year:04d}-{month:02d}",
                    "当前值": raw_value,
                    "说明": "预测数量必须是整数件数。",
                })
                continue
            if route_template is None:
                continue
            for due_date, weekly_quantity in _split_monthly_forecast_to_sundays(year, month, quantity_int):
                order_id = f"FCST-{due_date.strftime('%Y%m%d')}-{material}"
                original_due_date = due_date
                overdue = bool(optimization_start is not None and due_date < optimization_start)
                priority_rank, priority_type, priority_reason = _priority_from_flags("", overdue)
                existing = demand.get(order_id)
                if existing is None:
                    demand[order_id] = {
                        "订单": order_id,
                        "数量": float(weekly_quantity),
                        "交期": due_date,
                        "紧急": False,
                        "手动紧急": False,
                        "紧急类型": "",
                        "过期": overdue,
                        "原始最早交期": original_due_date,
                        "已转入优化开始周期": overdue,
                        "订单优先级": priority_rank,
                        "优先级类型": priority_type,
                        "优先级原因": priority_reason,
                        "需求来源": "预测需求",
                        "预测月份": f"{year:04d}-{month:02d}",
                        "预测周日": due_date.strftime("%Y-%m-%d"),
                        "预测物料": material,
                        "预测图号": drawing,
                        "预测工艺路线来源": forecast_route_source,
                        "路线来源订单": route_source_order,
                    }
                    continue
                existing["数量"] = float(existing.get("数量") or 0.0) + float(weekly_quantity)
        if route_template is None and has_positive_forecast and material not in missing_route_materials:
            missing_route_materials.add(material)
            issues.append({
                "类型": "预测物料缺少工艺路线，未参与计算",
                "文件": str(path),
                "行号": idx + 2,
                "物料": material,
                "图号": drawing,
                "说明": "该预测物料在北京、沈阳、南通工艺路线文件中均找不到匹配路线，已跳过，不参与本次产能计算。",
            })
    return demand, issues


def _blocking_forecast_issues(issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    blocking_types = {
        "缺少字段",
        "需求预测文件缺失",
        "需求预测清单缺失",
        "需求预测月份列异常",
        "需求预测数量异常",
    }
    return [row for row in issues if str(row.get("类型") or "") in blocking_types]


def _read_workbook_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return _read_workbook_first_sheet(path)
    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except ValueError as exc:
        raise ValueError(f"{Path(path).name} 缺少名为“{sheet_name}”的sheet。") from exc


def _parse_forecast_month_header(value: Any) -> tuple[int, int] | None:
    if isinstance(value, datetime):
        return value.year, value.month
    text = _clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d{6}", text):
        year = int(text[:4])
        month = int(text[4:6])
    else:
        match = re.search(r"(\d{4})\D?(\d{1,2})", text)
        if not match:
            return None
        year = int(match.group(1))
        month = int(match.group(2))
    if 1 <= month <= 12:
        return year, month
    return None


def _split_monthly_forecast_to_sundays(year: int, month: int, quantity: int) -> list[tuple[datetime, int]]:
    if quantity <= 0:
        return []
    _, days_in_month = calendar.monthrange(year, month)
    sundays = [
        datetime(year, month, day)
        for day in range(1, days_in_month + 1)
        if datetime(year, month, day).weekday() == 6
    ]
    if not sundays:
        return [(datetime(year, month, days_in_month), quantity)]
    base = quantity // len(sundays)
    remainder = quantity % len(sundays)
    rows: list[tuple[datetime, int]] = []
    for index, due_date in enumerate(sundays):
        weekly_quantity = base + (remainder if index == len(sundays) - 1 else 0)
        if weekly_quantity > 0:
            rows.append((due_date, weekly_quantity))
    return rows


def _load_forecast_route_templates() -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    route_templates: dict[str, dict[str, Any]] = {}
    issues: list[dict[str, Any]] = []
    for priority, (route_source, filename) in enumerate(FORECAST_ROUTE_TEMPLATE_SPECS, start=1):
        path = DATA_DIR / filename
        if not path.exists():
            issues.append({
                "类型": "预测工艺路线文件缺失",
                "文件": str(path),
                "预测工艺路线来源": route_source,
                "优先级": priority,
                "说明": "该工艺路线文件不存在；预测需求将继续尝试其他工艺路线文件。",
            })
            continue
        try:
            df = _read_workbook_sheet(path, FORECAST_ROUTE_SHEET_NAME)
        except ValueError as exc:
            issues.append({
                "类型": "预测工艺路线清单缺失",
                "文件": str(path),
                "预测工艺路线来源": route_source,
                "优先级": priority,
                "说明": str(exc),
            })
            continue
        source_templates, source_issues = _select_forecast_route_templates_by_material(
            df,
            route_source=route_source,
            route_file=path,
            priority=priority,
        )
        issues.extend(source_issues)
        for material, route in source_templates.items():
            if material not in route_templates:
                route_templates[material] = route
    return route_templates, issues


def _select_forecast_route_templates_by_material(
    route_df: pd.DataFrame,
    *,
    route_source: str,
    route_file: Path,
    priority: int,
) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    column_map = _forecast_route_column_map(route_df)
    if column_map is None:
        return {}, [{
            "类型": "预测工艺路线字段缺失",
            "文件": str(route_file),
            "预测工艺路线来源": route_source,
            "优先级": priority,
            "说明": "工艺路线Data sheet至少需要包含到I列：B列物料编码、E列工作中心描述、F列工序编码、G列准备/H、H列人工/H、I列设备/H。",
        }]

    identity_columns = [
        route_df.columns[index]
        for index in (0, 3, 11, 12, 15, 16)
        if index < len(route_df.columns)
    ]
    grouped: dict[tuple[str, tuple[str, ...]], dict[str, Any]] = {}
    for idx, row in route_df.iterrows():
        material = _clean_text(row.get(column_map["material"]))
        if not material:
            continue
        activity = _to_number(row.get(column_map["activity"]), default=0.0)
        process_text = _clean_text(row.get(column_map["process_text"]))
        unit_hour_values = {
            "标准值1": _to_number(row.get(column_map["standard_1"]), default=0.0),
            "标准值2": _to_number(row.get(column_map["standard_2"]), default=0.0),
            "标准值3": _to_number(row.get(column_map["standard_3"]), default=0.0),
        }
        route_identity = tuple(_clean_text(row.get(column)) for column in identity_columns)
        key = (material, route_identity)
        route_key_text = " | ".join(value for value in route_identity if value)
        item = grouped.setdefault(key, {
            "material": material,
            "source_order": "",
            "source_key": route_key_text,
            "route_source": route_source,
            "route_file": str(route_file),
            "route_priority": priority,
            "rows": [],
            "activities": set(),
            "total_unit_hours": 0.0,
            "first_row": idx + 2,
        })
        internal_row = {
            "物料": material,
            "活动": activity,
            "工序短文本": process_text,
            "标准值1": unit_hour_values["标准值1"],
            "标准值2": unit_hour_values["标准值2"],
            "标准值3": unit_hour_values["标准值3"],
            "预测工艺路线来源": route_source,
            "预测工艺路线文件": str(route_file),
            "预测工艺路线源文件行号": idx + 2,
            "预测工艺路线标识": route_key_text,
            "工作中心代码": _clean_text(row.get("工作中心代码")),
            "客户图号": _clean_text(row.get("客户图号")),
            "版本号": _clean_text(row.get("版本号")),
            "SAP版本号": _clean_text(row.get("SAP版本号")),
            "工艺类型": _clean_text(row.get("工艺类型")),
        }
        item["rows"].append((idx, internal_row))
        item["activities"].add(_activity_key(activity))
        item["total_unit_hours"] += sum(unit_hour_values.values())
        item["first_row"] = min(int(item["first_row"]), idx + 2)

    selected: dict[str, dict[str, Any]] = {}
    for route in grouped.values():
        material = str(route["material"])
        current = selected.get(material)
        route_key = (
            len(route["activities"]),
            len(route["rows"]),
            float(route["total_unit_hours"]),
            -int(route["first_row"]),
            str(route["source_key"]),
        )
        current_key = (
            len(current["activities"]),
            len(current["rows"]),
            float(current["total_unit_hours"]),
            -int(current["first_row"]),
            str(current["source_key"]),
        ) if current else None
        if current is None or route_key > current_key:
            selected[material] = route
    for route in selected.values():
        route["rows"] = sorted(route["rows"], key=lambda item: (_to_number(item[1].get("活动"), default=0.0), item[0]))
    return selected, []


def _forecast_route_column_map(route_df: pd.DataFrame) -> dict[str, Any] | None:
    if len(route_df.columns) <= 8:
        return None
    return {
        "material": route_df.columns[1],
        "process_text": route_df.columns[4],
        "activity": route_df.columns[5],
        "standard_1": route_df.columns[6],
        "standard_2": route_df.columns[7],
        "standard_3": route_df.columns[8],
    }


def _select_route_templates_by_material(operation_df: pd.DataFrame) -> dict[str, dict[str, Any]]:
    required = {"订单", "活动", "物料", "工序短文本", "标准值1", "标准值2", "标准值3"}
    if not required.issubset(set(operation_df.columns)):
        return {}
    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for idx, row in operation_df.iterrows():
        material = _clean_text(row.get("物料"))
        order_id = _normalize_order(row.get("订单"))
        if not material or not order_id:
            continue
        key = (material, order_id)
        item = grouped.setdefault(key, {
            "material": material,
            "source_order": order_id,
            "rows": [],
            "activities": set(),
            "total_unit_hours": 0.0,
            "first_row": idx + 2,
        })
        item["rows"].append((idx, row))
        item["activities"].add(_activity_key(row.get("活动")))
        item["total_unit_hours"] += sum(_to_number(row.get(col), default=0.0) for col in ("标准值1", "标准值2", "标准值3"))
        item["first_row"] = min(int(item["first_row"]), idx + 2)

    selected: dict[str, dict[str, Any]] = {}
    for route in grouped.values():
        material = str(route["material"])
        current = selected.get(material)
        route_key = (
            len(route["activities"]),
            len(route["rows"]),
            float(route["total_unit_hours"]),
            -int(route["first_row"]),
            str(route["source_order"]),
        )
        current_key = (
            len(current["activities"]),
            len(current["rows"]),
            float(current["total_unit_hours"]),
            -int(current["first_row"]),
            str(current["source_order"]),
        ) if current else None
        if current is None or route_key > current_key:
            selected[material] = route
    for route in selected.values():
        route["rows"] = sorted(route["rows"], key=lambda item: (_to_number(item[1].get("活动"), default=0.0), item[0]))
    return selected


def _optimization_start_month(config: FortuneBjConfig) -> datetime | None:
    return _optimization_start_period(config)


def _optimization_granularity(config: FortuneBjConfig | None) -> str:
    value = _clean_text(config.mode_b_optimization_granularity if config is not None else "") or "周"
    return "月" if "月" in value or value.lower() in {"month", "monthly"} else "周"


def _week_start(value: datetime) -> datetime:
    day = _day_start(value)
    return day - timedelta(days=day.weekday())


def _optimization_start_period(config: FortuneBjConfig | None) -> datetime | None:
    if config is None:
        return None
    value = config.mode_b_optimization_start_month
    if value is None:
        return None
    return _day_start(value)


def load_operation_tasks(
    path: Path,
    capacities: dict[str, WorkCenterCapacity],
    demand_by_order: dict[str, dict[str, Any]],
    config: FortuneBjConfig | None = None,
    progress: ProgressCallback | None = None,
) -> tuple[list[OperationTask], list[dict[str, Any]], list[dict[str, Any]]]:
    df = _read_workbook_first_sheet(path)
    required = ["订单", "活动", "物料", "工序短文本", "标准值1", "标准值2", "标准值3"]
    issues = _missing_columns_issue(df, required, path)
    if issues:
        return [], issues, []

    tasks: list[OperationTask] = []
    missing_mapping: list[dict[str, Any]] = []
    total_rows = len(df)
    started_at = time.perf_counter()
    _emit_progress(progress, "工序匹配", 0, total_rows, started_at, "开始匹配订单、工序和资源映射")
    for idx, row in df.iterrows():
        order_id = _normalize_order(row.get("订单"))
        if not order_id or order_id not in demand_by_order:
            continue
        activity = _to_number(row.get("活动"), default=0.0)
        process_text = _clean_text(row.get("工序短文本"))
        material = _clean_text(row.get("物料"))
        demand = demand_by_order[order_id]
        unit_hours = sum(_to_number(row.get(col), default=0.0) for col in ("标准值1", "标准值2", "标准值3"))
        is_outsource = "外协" in process_text
        capacity = capacities.get(process_text)
        is_hot_surface = _is_hot_surface(process_text)
        capacity_calc_type = _clean_text(row.get("工艺处理类型")) or (capacity.capacity_calc_type if capacity else "")
        hot_surface_type = _hot_surface_type_from_text(row.get("热处表处类型"), process_text, capacity.hot_surface_type if capacity else "")
        process_group = _clean_text(row.get("工艺兼容组")) or (capacity.process_groups if capacity else "") or process_text
        unit_capacity = _positive_or_default(
            _to_number(row.get("单件容量占用"), default=0.0),
            capacity.default_unit_capacity if capacity else 1.0,
        )
        duration = _operation_load_hours(
            quantity=float(demand["数量"]),
            unit_hours=unit_hours,
            capacity=capacity,
            is_hot_surface=is_hot_surface,
            capacity_calc_type=capacity_calc_type,
            unit_capacity=unit_capacity,
            config=config,
        )
        missing_work_center = bool(capacity is None and not is_outsource)
        if missing_work_center:
            missing_mapping.append({
                "订单": order_id,
                "活动": activity,
                "物料": material,
                "工序短文本": process_text,
                "源文件行号": idx + 2,
                "处理建议": "请在工作中心模板中补充该工序短文本对应的工作中心、资源组分类、设备数量和日历名称；当前运行将其作为未维护工作中心负荷单独报告。",
            })
        tasks.append(OperationTask(
            order_id=order_id,
            activity=activity,
            material=material,
            process_text=process_text,
            work_center=capacity.work_center if capacity else (UNMAINTAINED_WORKCENTER if missing_work_center else "外协"),
            resource_group=capacity.resource_group if capacity else (UNMAINTAINED_WORKCENTER if missing_work_center else "外协"),
            quantity=float(demand["数量"]),
            unit_hours=unit_hours,
            duration_hours=duration,
            due_date=demand["交期"],
            urgent=bool(demand["紧急"]),
            manual_urgent=bool(demand.get("手动紧急")),
            urgent_type=str(demand.get("紧急类型") or ""),
            overdue=bool(demand.get("过期")),
            adjusted_to_start_period=bool(demand.get("已转入优化开始周期")),
            original_due_date=demand.get("原始最早交期") if isinstance(demand.get("原始最早交期"), datetime) else demand["交期"],
            priority_rank=int(demand.get("订单优先级") or PRIORITY_NORMAL),
            priority_type=str(demand.get("优先级类型") or "普通订单"),
            priority_reason=str(demand.get("优先级原因") or "普通订单"),
            demand_source=str(demand.get("需求来源") or "真实订单"),
            forecast_month=str(demand.get("预测月份") or ""),
            forecast_week_end=str(demand.get("预测周日") or ""),
            forecast_drawing=str(demand.get("预测图号") or ""),
            forecast_route_source=str(demand.get("预测工艺路线来源") or ""),
            route_source_order=str(demand.get("路线来源订单") or ""),
            is_outsource=is_outsource,
            missing_work_center=missing_work_center,
            is_hot_surface=is_hot_surface,
            source_row=idx + 2,
            capacity_calc_type=capacity_calc_type,
            hot_surface_type=hot_surface_type,
            process_group=process_group,
            unit_capacity=unit_capacity,
            allow_batch_mix=_to_bool(row.get("是否允许合炉")) if "是否允许合炉" in df.columns else True,
            must_same_batch=_to_bool(row.get("是否必须整单同批")),
            treatment_program=_clean_text(row.get("热处/表处程序")),
        ))
        processed = idx + 1
        if processed % 5000 == 0 or processed == total_rows:
            _emit_progress(
                progress,
                "工序匹配",
                processed,
                total_rows,
                started_at,
                f"可分析 {len(tasks):,}，缺失映射 {len(missing_mapping):,}",
            )
    forecast_tasks, forecast_missing_mapping, forecast_task_issues = _build_forecast_operation_tasks(
        operation_df=df,
        capacities=capacities,
        demand_by_order=demand_by_order,
        config=config,
    )
    tasks.extend(forecast_tasks)
    missing_mapping.extend(forecast_missing_mapping)
    issues.extend(forecast_task_issues)
    tasks.sort(key=lambda task: (task.order_id, task.activity, task.source_row or 0))
    _emit_progress(
        progress,
        "工序匹配",
        total_rows,
        total_rows,
        started_at,
        f"完成：可分析 {len(tasks):,}，缺失映射 {len(missing_mapping):,}",
    )
    return tasks, issues, missing_mapping


def _build_forecast_operation_tasks(
    *,
    operation_df: pd.DataFrame,
    capacities: dict[str, WorkCenterCapacity],
    demand_by_order: dict[str, dict[str, Any]],
    config: FortuneBjConfig | None,
) -> tuple[list[OperationTask], list[dict[str, Any]], list[dict[str, Any]]]:
    forecast_demands = {
        order_id: demand
        for order_id, demand in demand_by_order.items()
        if str(demand.get("需求来源") or "") == "预测需求"
    }
    if not forecast_demands:
        return [], [], []

    route_templates, _route_issues = _load_forecast_route_templates()
    tasks: list[OperationTask] = []
    missing_mapping: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    for order_id, demand in sorted(forecast_demands.items()):
        material = _clean_text(demand.get("预测物料")) or _forecast_material_from_order_id(order_id)
        route = route_templates.get(material)
        if route is None:
            issues.append({
                "类型": "预测虚拟订单缺少工艺路线，未参与计算",
                "订单": order_id,
                "物料": material,
                "说明": "该预测虚拟订单未找到北京、沈阳、南通工艺路线，已跳过工序生成。",
            })
            continue
        for idx, row in route["rows"]:
            activity = _to_number(row.get("活动"), default=0.0)
            process_text = _clean_text(row.get("工序短文本"))
            route_material = _clean_text(row.get("物料")) or material
            unit_hours = sum(_to_number(row.get(col), default=0.0) for col in ("标准值1", "标准值2", "标准值3"))
            is_outsource = "外协" in process_text
            capacity = capacities.get(process_text)
            is_hot_surface = _is_hot_surface(process_text)
            capacity_calc_type = _clean_text(row.get("工艺处理类型")) or (capacity.capacity_calc_type if capacity else "")
            hot_surface_type = _hot_surface_type_from_text(row.get("热处表处类型"), process_text, capacity.hot_surface_type if capacity else "")
            process_group = _clean_text(row.get("工艺兼容组")) or (capacity.process_groups if capacity else "") or process_text
            unit_capacity = _positive_or_default(
                _to_number(row.get("单件容量占用"), default=0.0),
                capacity.default_unit_capacity if capacity else 1.0,
            )
            source_row = int(_to_number(row.get("预测工艺路线源文件行号"), default=idx + 2))
            quantity = float(demand.get("数量") or 0.0)
            duration = _operation_load_hours(
                quantity=quantity,
                unit_hours=unit_hours,
                capacity=capacity,
                is_hot_surface=is_hot_surface,
                capacity_calc_type=capacity_calc_type,
                unit_capacity=unit_capacity,
                config=config,
            )
            missing_work_center = bool(capacity is None and not is_outsource)
            if missing_work_center:
                missing_mapping.append({
                    "订单": order_id,
                    "活动": activity,
                    "物料": route_material,
                    "工序短文本": process_text,
                    "源文件行号": source_row,
                    "预测工艺路线来源": str(route.get("route_source") or row.get("预测工艺路线来源") or ""),
                    "预测工艺路线文件": str(route.get("route_file") or row.get("预测工艺路线文件") or ""),
                    "处理建议": "预测需求采用工艺路线文件；请在工作中心模板中补充该工序短文本对应的工作中心、资源组分类、设备数量和日历名称。",
                })
            tasks.append(OperationTask(
                order_id=order_id,
                activity=activity,
                material=route_material,
                process_text=process_text,
                work_center=capacity.work_center if capacity else (UNMAINTAINED_WORKCENTER if missing_work_center else "外协"),
                resource_group=capacity.resource_group if capacity else (UNMAINTAINED_WORKCENTER if missing_work_center else "外协"),
                quantity=quantity,
                unit_hours=unit_hours,
                duration_hours=duration,
                due_date=demand["交期"],
                urgent=bool(demand.get("紧急")),
                manual_urgent=bool(demand.get("手动紧急")),
                urgent_type=str(demand.get("紧急类型") or ""),
                overdue=bool(demand.get("过期")),
                adjusted_to_start_period=bool(demand.get("已转入优化开始周期")),
                original_due_date=demand.get("原始最早交期") if isinstance(demand.get("原始最早交期"), datetime) else demand["交期"],
                priority_rank=int(demand.get("订单优先级") or PRIORITY_NORMAL),
                priority_type=str(demand.get("优先级类型") or "普通订单"),
                priority_reason=str(demand.get("优先级原因") or "普通订单"),
                demand_source="预测需求",
                forecast_month=str(demand.get("预测月份") or ""),
                forecast_week_end=str(demand.get("预测周日") or ""),
                forecast_drawing=str(demand.get("预测图号") or ""),
                forecast_route_source=str(route.get("route_source") or row.get("预测工艺路线来源") or demand.get("预测工艺路线来源") or ""),
                route_source_order=str(route.get("source_order") or demand.get("路线来源订单") or ""),
                is_outsource=is_outsource,
                missing_work_center=missing_work_center,
                is_hot_surface=is_hot_surface,
                source_row=source_row,
                capacity_calc_type=capacity_calc_type,
                hot_surface_type=hot_surface_type,
                process_group=process_group,
                unit_capacity=unit_capacity,
                allow_batch_mix=_to_bool(row.get("是否允许合炉")) if "是否允许合炉" in row else True,
                must_same_batch=_to_bool(row.get("是否必须整单同批")),
                treatment_program=_clean_text(row.get("热处/表处程序")),
            ))
    return tasks, missing_mapping, issues


def _forecast_material_from_order_id(order_id: str) -> str:
    parts = str(order_id or "").split("-", 2)
    return parts[2] if len(parts) == 3 else ""


def _schedule_tasks(
    tasks: list[OperationTask],
    capacities: dict[str, WorkCenterCapacity],
    *,
    mode: str,
    optional_operations: dict[tuple[str, str], list[OptionalOperation]] | None = None,
    result: ScheduleResult | None = None,
    config: FortuneBjConfig | None = None,
    progress: ProgressCallback | None = None,
) -> list[ScheduledOperation]:
    if not tasks:
        return []
    if mode == "ModeB":
        return _schedule_tasks_mode_b_v2(
            tasks,
            capacities,
            optional_operations or {},
            result,
            config=config,
            max_window_tasks=config.mode_b_max_window_tasks if config else 2000,
            progress=progress,
        )
    start_period = _optimization_start_period(config)
    if progress is not None:
        progress(
            f"ModeA无限产能倒排: {len(tasks):,} 条工序；"
            f"优化粒度 {_optimization_granularity(config)}，"
            f"优化开始日期 {start_period.strftime('%Y-%m-%d') if start_period else '未指定'}；"
            f"逾期订单或倒排后早于优化开始日期的订单会整单平移到优化开始日期"
        )
    scheduled = _schedule_tasks_infinite_capacity(tasks, config=config)
    _mark_analysis_items(scheduled, status="ModeA倒排平移基线", source="ModeA", window_type="全量分析")
    if result is not None:
        result.bottleneck_report = _build_bottleneck_report(scheduled, capacities, config=config)
        result.monthly_capacity_report = _build_monthly_capacity_report(scheduled, scheduled, capacities, config=config)
    return scheduled


def _normalize_mode(value: str) -> str:
    text = str(value or "").strip().upper()
    if "B" in text:
        return "ModeB"
    return "ModeA"


def _shift_order_items_to_optimization_start(
    order_items: list[ScheduledOperation],
    *,
    config: FortuneBjConfig | None,
    mode_label: str,
) -> None:
    optimization_start = _optimization_start_period(config)
    if optimization_start is None or not order_items:
        return
    first_start = min(item.start for item in order_items)
    order_overdue = any(item.task.overdue or item.task.adjusted_to_start_period for item in order_items)
    starts_before_period = any(item.start < optimization_start for item in order_items)
    if not (order_overdue or starts_before_period) or first_start >= optimization_start:
        return

    delta = optimization_start - first_start
    for item in order_items:
        item.start += delta
        item.end += delta
        item.on_time = item.end <= item.task.due_date
        item.tardy_hours = max((item.end - item.task.due_date).total_seconds() / 3600.0, 0.0)
        if item.task.is_outsource:
            item.note = f"{mode_label}按交期倒排后整单平移至优化开始日期；外协固定7天，不占用本地资源"
        else:
            item.note = f"{mode_label}按交期倒排后整单平移至优化开始日期"


def _schedule_order_tasks_by_flow(
    order_tasks: list[OperationTask],
    *,
    config: FortuneBjConfig | None,
    mode_label: str,
) -> list[ScheduledOperation]:
    sorted_tasks = sorted(order_tasks, key=lambda task: task.activity)
    if not sorted_tasks:
        return []
    flow_mode = _operation_flow_mode(config)
    due_date = min(task.due_date for task in sorted_tasks)
    if flow_mode == "交期强制":
        return [
            ScheduledOperation(
                task=task,
                start=due_date,
                end=due_date,
                on_time=True,
                tardy_hours=0.0,
                note=f"{mode_label}交期强制口径；全部工序负荷归入交期当天",
            )
            for task in sorted_tasks
        ]

    quantity_units = max(max(_flow_quantity_units(task.quantity) for task in sorted_tasks), 1)
    lot_sizes = _operation_flow_lot_sizes(quantity_units, flow_mode)
    if not lot_sizes:
        lot_sizes = [quantity_units]
    lot_total = max(sum(lot_sizes), 1)

    previous_operation_finish: list[float] = [0.0 for _ in lot_sizes]
    spans: list[tuple[float, float]] = []
    makespan = 0.0
    for operation_index, task in enumerate(sorted_tasks):
        duration_hours = _task_flow_duration_hours(task)
        unit_duration = duration_hours / lot_total if lot_total > 0 else duration_hours
        current_operation_finish: list[float] = []
        previous_lot_finish = 0.0
        first_start: float | None = None
        last_finish = 0.0
        for lot_index, lot_size in enumerate(lot_sizes):
            upstream_finish = previous_operation_finish[lot_index] if operation_index > 0 else 0.0
            lot_start = max(previous_lot_finish, upstream_finish)
            lot_finish = lot_start + max(float(lot_size), 0.0) * unit_duration
            if first_start is None:
                first_start = lot_start
            previous_lot_finish = lot_finish
            current_operation_finish.append(lot_finish)
            last_finish = lot_finish
        spans.append((first_start or 0.0, last_finish))
        previous_operation_finish = current_operation_finish
        makespan = max(makespan, last_finish)

    order_start = due_date - timedelta(hours=makespan)
    items: list[ScheduledOperation] = []
    for task, (start_offset, end_offset) in zip(sorted_tasks, spans):
        start = order_start + timedelta(hours=start_offset)
        end = order_start + timedelta(hours=end_offset)
        if task.is_outsource:
            note = f"{mode_label}{flow_mode}倒排；外协固定7天，不占用本地资源"
        else:
            note = f"{mode_label}{flow_mode}倒排"
        items.append(ScheduledOperation(
            task=task,
            start=start,
            end=end,
            on_time=end <= task.due_date,
            tardy_hours=max((end - task.due_date).total_seconds() / 3600.0, 0.0),
            note=note,
        ))
    return items


def _operation_flow_mode(config: FortuneBjConfig | None) -> str:
    text = _clean_text(config.operation_flow_mode if config is not None else "") or "整批流转"
    if "交期" in text:
        return "交期强制"
    if "单件" in text:
        return "单件流转"
    if "半批" in text:
        return "半批流转"
    return "整批流转"


def _flow_quantity_units(quantity: Any) -> int:
    value = _to_number(quantity, default=0.0)
    if value <= 0:
        return 1
    rounded = int(round(value))
    if abs(value - rounded) <= 1e-6:
        return max(rounded, 1)
    return max(int(math.ceil(value)), 1)


def _operation_flow_lot_sizes(quantity_units: int, flow_mode: str) -> list[int]:
    quantity_units = max(int(quantity_units), 1)
    if flow_mode == "单件流转":
        return [1 for _ in range(quantity_units)]
    if flow_mode == "半批流转":
        first = int(math.ceil(quantity_units / 2))
        second = quantity_units - first
        return [first] + ([second] if second > 0 else [])
    return [quantity_units]


def _task_flow_duration_hours(task: OperationTask) -> float:
    if task.is_outsource:
        return float(OUTSOURCE_DURATION_HOURS)
    return max(float(task.duration_hours or 0.0), 0.0)


def _schedule_tasks_infinite_capacity(
    tasks: list[OperationTask],
    *,
    config: FortuneBjConfig | None = None,
) -> list[ScheduledOperation]:
    tasks_by_order: dict[str, list[OperationTask]] = {}
    for task in tasks:
        tasks_by_order.setdefault(task.order_id, []).append(task)
    scheduled: list[ScheduledOperation] = []
    def order_sort_key(key: str) -> tuple[datetime, int, datetime, datetime, str]:
        order_tasks = tasks_by_order[key]
        first_period_start = min(_optimization_period_start(task, config) for task in order_tasks)
        priority_rank = min(task.priority_rank for task in order_tasks)
        priority_date = min(_priority_sort_date(task) for task in order_tasks)
        due_date = min(task.due_date for task in order_tasks)
        return first_period_start, priority_rank, priority_date, due_date, key

    for order_id in sorted(
        tasks_by_order,
        key=order_sort_key,
    ):
        order_items = _schedule_order_tasks_by_flow(tasks_by_order[order_id], config=config, mode_label="ModeA")
        _shift_order_items_to_optimization_start(order_items, config=config, mode_label="ModeA")
        scheduled.extend(order_items)
    return scheduled


def _schedule_tasks_modeb_backward_with_start_shift(
    tasks: list[OperationTask],
    *,
    config: FortuneBjConfig | None = None,
) -> list[ScheduledOperation]:
    tasks_by_order: dict[str, list[OperationTask]] = {}
    for task in tasks:
        tasks_by_order.setdefault(task.order_id, []).append(task)
    scheduled: list[ScheduledOperation] = []

    def order_sort_key(key: str) -> tuple[datetime, int, datetime, datetime, str]:
        order_tasks = tasks_by_order[key]
        first_period_start = min(_optimization_period_start(task, config) for task in order_tasks)
        priority_rank = min(task.priority_rank for task in order_tasks)
        priority_date = min(_priority_sort_date(task) for task in order_tasks)
        due_date = min(task.due_date for task in order_tasks)
        return first_period_start, priority_rank, priority_date, due_date, key

    for order_id in sorted(tasks_by_order, key=order_sort_key):
        order_items = _schedule_order_tasks_by_flow(tasks_by_order[order_id], config=config, mode_label="ModeB")
        _shift_order_items_to_optimization_start(order_items, config=config, mode_label="ModeB")
        scheduled.extend(order_items)
    return scheduled


def _schedule_tasks_mode_b_v2(
    tasks: list[OperationTask],
    capacities: dict[str, WorkCenterCapacity],
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
    result: ScheduleResult | None = None,
    *,
    config: FortuneBjConfig | None = None,
    max_window_tasks: int = 2000,
    progress: ProgressCallback | None = None,
) -> list[ScheduledOperation]:
    max_window_tasks = max(int(max_window_tasks), 1)
    granularity = _optimization_granularity(config)
    start_period = _optimization_start_period(config)
    if progress is not None:
        progress(
            f"ModeB 100%产能优化建议: 先按订单交期倒排负荷，工序 {len(tasks):,} 条；"
            f"优化粒度 {granularity}，优化开始日期 {start_period.strftime('%Y-%m-%d') if start_period else '未指定'}；"
            f"逾期订单或倒排后早于优化开始日期的订单会整单平移到优化开始日期；"
            f"在每个{granularity}内按整数产品数量分配原工作中心、可选工作中心和外包；"
            f"记录参考规模 {max_window_tasks:,} 条工序/周期，求解时间上限 "
            f"{config.mode_b_solver_max_seconds if config else 60.0} 秒"
        )

    baseline = _schedule_tasks_modeb_backward_with_start_shift(tasks, config=config)
    _mark_analysis_items(baseline, status="ModeB倒排平移基线", source="ModeB", window_type="优化基线")
    bottleneck_report = _build_bottleneck_report(baseline, capacities, config=config)
    bottleneck_workcenters = {
        str(row["工作中心"])
        for row in bottleneck_report
        if row.get("是否瓶颈") == "是"
    }
    if result is not None:
        result.bottleneck_report = bottleneck_report
    if progress is not None:
        progress(f"ModeB瓶颈识别: 基于交期倒排和平移后的负荷识别瓶颈工作中心 {len(bottleneck_workcenters)} 个")

    route_started = time.perf_counter()
    final_items, allocations, option_rows, optimization_summary, optimization_stats = _optimize_modeb_integer_allocations(
        baseline,
        capacities,
        optional_operations,
        config=config,
        progress=progress,
    )
    route_seconds = round(time.perf_counter() - route_started, 3)
    if result is not None:
        result.optional_operation_report = option_rows
        result.order_operation_allocation_report = _build_modeb_order_allocation_rows(allocations, config=config)
        result.capacity_optimization_summary = optimization_summary
        result.capacity_recommendation_report = _build_capacity_recommendation_report(optimization_summary)
        result.capacity_optimization_stats = optimization_stats
        result.hot_surface_capacity_report = _build_hot_surface_capacity_report(allocations, capacities, config=config)
        result.window_report = _build_modeb_period_report_from_allocations(
            allocations,
            capacities,
            config=config,
            max_window_tasks=max_window_tasks,
            solve_seconds=route_seconds,
            status="已优化" if option_rows else "无可分流或无需分流",
        )
        result.monthly_capacity_report = _build_modeb_period_capacity_report_from_allocations(
            baseline,
            allocations,
            capacities,
            config=config,
        )
    return final_items


def _schedule_tasks_mode_b_legacy(
    tasks: list[OperationTask],
    capacities: dict[str, WorkCenterCapacity],
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
    result: ScheduleResult | None = None,
    *,
    config: FortuneBjConfig | None = None,
    max_window_tasks: int = 2000,
    progress: ProgressCallback | None = None,
) -> list[ScheduledOperation]:
    max_window_tasks = max(int(max_window_tasks), 1)
    granularity = _optimization_granularity(config)
    start_period = _optimization_start_period(config)
    if progress is not None:
        progress(
            f"ModeB 100%产能优化建议: 先执行 ModeA 无限产能分析，工序 {len(tasks):,} 条；"
            f"优化粒度 {granularity}，优化开始日期 {start_period.strftime('%Y-%m-%d') if start_period else '未指定'}，"
            f"按{granularity}覆盖全部数据；"
            f"记录参考规模 {max_window_tasks:,} 条工序/周期，求解时间上限 "
            f"{config.mode_b_solver_max_seconds if config else 60.0} 秒；当前版本不按规模阈值拦截"
        )
    baseline = _schedule_tasks_infinite_capacity(tasks, config=config)
    _mark_analysis_items(baseline, status="无限产能估算", source="ModeA", window_type="窗口外")
    bottleneck_report = _build_modeb_bottleneck_report(tasks, capacities, config=config)
    bottleneck_workcenters = {
        str(row["工作中心"])
        for row in bottleneck_report
        if row.get("是否瓶颈") == "是"
    }
    if result is not None:
        result.bottleneck_report = bottleneck_report
    if progress is not None:
        progress(f"ModeB瓶颈识别: 瓶颈工作中心 {len(bottleneck_workcenters)} 个")
    if not bottleneck_workcenters:
        if result is not None:
            result.window_report = _build_modeb_period_report(
                tasks,
                tasks,
                capacities,
                optional_operations,
                bottleneck_workcenters,
                config=config,
                max_window_tasks=max_window_tasks,
                solve_seconds=0,
                status="无瓶颈",
            )
            result.optional_operation_report = []
            result.capacity_optimization_summary = _build_capacity_optimization_summary(
                _task_load_by_optimization_period(tasks, config=config),
                _task_load_by_optimization_period(tasks, config=config),
                capacities,
                config=config,
            )
            result.capacity_recommendation_report = _build_capacity_recommendation_report(
                result.capacity_optimization_summary
            )
            result.capacity_optimization_stats = _modeb_capacity_stats_rows(
                stats={
                    "状态": "跳过",
                    "说明": "ModeA 无限产能分析未识别到瓶颈工作中心",
                    "输入工序数": len(tasks),
                    "候选决策工序数": 0,
                    "候选方案数": 0,
                    "布尔变量数": 0,
                    "短缺变量数": 0,
                    "周期数": 0,
                    "工作中心数": len(capacities),
                    "求解状态": "SKIPPED",
                    "建模耗时秒": 0,
                    "求解耗时秒": 0,
                    "总短缺小时": 0,
                    "超载工作中心周期数": 0,
                },
                config=config,
            )
            result.monthly_capacity_report = _build_modeb_period_capacity_report(tasks, tasks, capacities, config=config)
        return _build_modeb_capacity_items(tasks, config=config)

    route_started = time.perf_counter()
    adjusted_tasks, option_rows, optimization_summary, optimization_stats = _optimize_capacity_route_choices(
        tasks,
        baseline,
        bottleneck_report,
        capacities,
        optional_operations,
        config=config,
        progress=progress,
    )
    route_seconds = round(time.perf_counter() - route_started, 3)
    if result is not None:
        result.optional_operation_report = option_rows
        result.capacity_optimization_summary = optimization_summary
        result.capacity_recommendation_report = _build_capacity_recommendation_report(optimization_summary)
        result.capacity_optimization_stats = optimization_stats
    final_items = _build_modeb_capacity_items(adjusted_tasks, config=config)
    if result is not None:
        result.window_report = _build_modeb_period_report(
            tasks,
            adjusted_tasks,
            capacities,
            optional_operations,
            bottleneck_workcenters,
            config=config,
            max_window_tasks=max_window_tasks,
            solve_seconds=route_seconds,
            status="已优化",
        )
        result.monthly_capacity_report = _build_modeb_period_capacity_report(tasks, adjusted_tasks, capacities, config=config)
    return final_items


def _mark_analysis_items(
    items: Iterable[ScheduledOperation],
    *,
    status: str,
    source: str,
    window_type: str,
    window_number: int | None = None,
) -> None:
    for item in items:
        item.analysis_status = status
        item.analysis_source = source
        item.window_type = window_type
        item.window_number = window_number


def _is_fixed_within_window(item: ScheduledOperation, window_start: datetime, window_end: datetime) -> bool:
    if item.task.is_outsource:
        return False
    return window_start <= item.start and item.end <= window_end


def _optimization_period_start(task: OperationTask, config: FortuneBjConfig | None) -> datetime:
    due_day = _day_start(task.due_date)
    return _period_start_for_date(due_day, config)


def _optimization_period_label(task: OperationTask, config: FortuneBjConfig | None) -> str:
    return _period_label_from_start(_optimization_period_start(task, config), config)


def _optimization_period_end(period_start: datetime, config: FortuneBjConfig | None) -> datetime:
    if _optimization_granularity(config) == "月":
        return _month_add(_month_start(period_start), 1)
    return _week_start(period_start) + timedelta(days=7)


def _period_start_for_date(value: datetime, config: FortuneBjConfig | None) -> datetime:
    if _optimization_granularity(config) == "月":
        return datetime(value.year, value.month, 1)
    return _week_start(value)


def _period_label_from_start(period_start: datetime | None, config: FortuneBjConfig | None) -> str:
    if period_start is None:
        return ""
    if _optimization_granularity(config) == "月":
        return period_start.strftime("%Y-%m")
    iso = period_start.isocalendar()
    return f"{iso.year}-W{iso.week:02d}"


def _period_bounds_from_label(period: str, config: FortuneBjConfig | None) -> tuple[datetime, datetime]:
    granularity = _optimization_granularity(config)
    try:
        if granularity == "月" or "-W" not in period:
            year, month = (int(part) for part in period.split("-")[:2])
            start = datetime(year, month, 1)
            return _clip_period_bounds_to_optimization_start(start, _month_add(start, 1), config)
        year_text, week_text = period.split("-W", 1)
        start = datetime.fromisocalendar(int(year_text), int(week_text), 1)
        return _clip_period_bounds_to_optimization_start(start, start + timedelta(days=7), config)
    except (ValueError, IndexError):
        start = datetime(1900, 1, 1)
        return start, start + timedelta(days=30 if granularity == "月" else 7)


def _clip_period_bounds_to_optimization_start(
    period_start: datetime,
    period_end: datetime,
    config: FortuneBjConfig | None,
) -> tuple[datetime, datetime]:
    optimization_start = _optimization_start_period(config)
    if optimization_start is None:
        return period_start, period_end
    if period_start <= optimization_start < period_end:
        return optimization_start, period_end
    return period_start, period_end


def _period_display_end(period_end: datetime) -> datetime:
    return period_end - timedelta(days=1)


def _period_span_from_bounds(period_start: datetime, period_end: datetime) -> str:
    end = _period_display_end(period_end)
    return f"{period_start.strftime('%Y-%m-%d')} 至 {end.strftime('%Y-%m-%d')}"


def _period_date_span(period: Any, config: FortuneBjConfig | None) -> str:
    period_text = str(period or "").strip()
    if not period_text or period_text == "无":
        return ""
    start, end = _period_bounds_from_label(period_text, config)
    if start.year == 1900:
        return ""
    return _period_span_from_bounds(start, end)


def _period_display_label(period: Any, config: FortuneBjConfig | None) -> str:
    period_text = str(period or "").strip()
    if not period_text:
        return ""
    span = _period_date_span(period_text, config)
    return f"{period_text} ({span})" if span else period_text


def _period_capacity_hours(period: str, capacity: WorkCenterCapacity | None, *, config: FortuneBjConfig | None = None) -> float:
    if capacity is None:
        return 0.0
    start, end = _period_bounds_from_label(period, config)
    days = max((end - start).days, 1)
    return capacity.quantity * capacity.daily_hours * days


def _task_load_by_optimization_period(
    tasks: list[OperationTask],
    *,
    config: FortuneBjConfig | None,
) -> dict[tuple[str, str], float]:
    load: dict[tuple[str, str], float] = {}
    for task in tasks:
        if task.is_outsource or task.missing_work_center:
            continue
        period = _optimization_period_label(task, config)
        key = (period, task.work_center)
        load[key] = load.get(key, 0.0) + task.duration_hours
    return load


def _task_load_by_reporting_period(
    tasks: list[OperationTask],
    *,
    config: FortuneBjConfig | None,
) -> dict[tuple[str, str], dict[str, Any]]:
    buckets: dict[tuple[str, str], dict[str, Any]] = {}
    for task in tasks:
        if task.is_outsource or task.missing_work_center:
            continue
        period = _optimization_period_label(task, config)
        period_start = _optimization_period_start(task, config)
        period_end = _optimization_period_end(period_start, config)
        key = (period, task.work_center)
        bucket = buckets.setdefault(key, {
            "周期": period,
            "周期日期跨度": _period_span_from_bounds(period_start, period_end),
            "周期开始": period_start.strftime("%Y-%m-%d"),
            "周期结束": _period_display_end(period_end).strftime("%Y-%m-%d"),
            "工作中心": task.work_center,
            "资源组分类": task.resource_group,
            "工序数": 0,
            "负荷小时": 0.0,
        })
        bucket["工序数"] += 1
        bucket["负荷小时"] += task.duration_hours
    return buckets


def _build_modeb_capacity_items(tasks: list[OperationTask], *, config: FortuneBjConfig | None) -> list[ScheduledOperation]:
    items: list[ScheduledOperation] = []
    for task in tasks:
        period_start = _optimization_period_start(task, config)
        end = task.due_date
        if task.is_outsource:
            start = end - timedelta(hours=OUTSOURCE_DURATION_HOURS)
            note = "外包无限产能；固定按7天日历周期返回，不占用厂内工作中心产能"
        else:
            start = min(period_start, end)
            note = "ModeB周期内产能优化口径；不作为车间执行排程时间"
        items.append(ScheduledOperation(
            task=task,
            start=start,
            end=end,
            on_time=True,
            tardy_hours=0.0,
            note=note,
            analysis_status="周期内产能优化",
            analysis_source="ModeB_OR-Tools",
            window_number=None,
            window_type="优化周期",
        ))
    return sorted(items, key=lambda item: (item.task.due_date, item.task.order_id, item.task.activity))


def _build_modeb_bottleneck_report(
    tasks: list[OperationTask],
    capacities: dict[str, WorkCenterCapacity],
    *,
    config: FortuneBjConfig | None,
) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, str], dict[str, Any]] = {}
    for task in tasks:
        if task.is_outsource or task.missing_work_center:
            continue
        period = _optimization_period_label(task, config)
        key = (period, task.work_center)
        bucket = buckets.setdefault(key, {
            "期间": period,
            "期间日期跨度": _period_date_span(period, config),
            "工作中心": task.work_center,
            "资源组分类": task.resource_group,
            "工序数": 0,
            "负荷小时": 0.0,
        })
        bucket["工序数"] += 1
        bucket["负荷小时"] += task.duration_hours
    rows: list[dict[str, Any]] = []
    for (period, work_center), bucket in sorted(buckets.items(), key=lambda kv: (kv[0][0], kv[0][1])):
        capacity = capacities.get(work_center)
        capacity_hours = _period_capacity_hours(period, capacity, config=config)
        load_hours = float(bucket["负荷小时"])
        ratio = load_hours / capacity_hours if capacity_hours > 0 else 0.0
        rows.append({
            "期间": period,
            "期间日期跨度": bucket["期间日期跨度"],
            "工作中心": work_center,
            "资源组分类": bucket["资源组分类"],
            "日历名称": capacity.calendar_name if capacity else "",
            "设备数量": capacity.quantity if capacity else 0,
            "平均每日小时/台": round(capacity.daily_hours, 2) if capacity else 0,
            "工序数": bucket["工序数"],
            "负荷小时": round(load_hours, 2),
            "产能小时": round(capacity_hours, 2),
            "负荷率": round(ratio, 4),
            "是否瓶颈": "是" if ratio > 1 else "否",
        })
    rows.sort(key=lambda row: (row["是否瓶颈"] != "是", -float(row["负荷率"]), row["期间"], row["工作中心"]))
    return rows


def _modeb_quantity_units(task: OperationTask) -> int:
    quantity = float(task.quantity or 0)
    if quantity <= 0:
        return 0
    rounded = int(round(quantity))
    if abs(quantity - rounded) <= 1e-6:
        return max(rounded, 1)
    return max(int(math.floor(quantity)), 1)


def _modeb_unit_hours(task: OperationTask) -> float:
    quantity = float(task.quantity or 0)
    if task.is_hot_surface and quantity > 0 and task.duration_hours > 0:
        return max(float(task.duration_hours) / quantity, 0.0)
    if task.unit_hours > 0:
        return float(task.unit_hours)
    if quantity > 0:
        return max(float(task.duration_hours) / quantity, 0.0)
    return max(float(task.duration_hours), 0.0)


def _modeb_period_for_item(
    item: ScheduledOperation,
    *,
    config: FortuneBjConfig | None,
) -> tuple[str, datetime, datetime]:
    anchor = item.start
    modeb_start = _optimization_start_period(config)
    if modeb_start is not None and anchor < modeb_start:
        anchor = modeb_start
    period = _period_label_from_start(_period_start_for_date(anchor, config), config)
    start, end = _period_bounds_from_label(period, config)
    return period, start, end


def _modeb_route_options(
    task: OperationTask,
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
    capacities: dict[str, WorkCenterCapacity],
    config: FortuneBjConfig | None,
) -> list[dict[str, Any]]:
    original_unit_hours = _modeb_unit_hours(task)
    original_profile = _capacity_profile_for_route(
        task,
        option=None,
        work_center=task.work_center,
        unit_hours=original_unit_hours,
        capacities=capacities,
        config=config,
    )
    options: list[dict[str, Any]] = [{
        "destination_type": "原工作中心" if not task.is_outsource else "外包",
        "work_center": task.work_center if not task.is_outsource else "外包",
        "resource_group": task.resource_group if not task.is_outsource else "外包",
        "unit_hours": 0.0 if task.is_outsource else original_unit_hours,
        "unit_hours_source": "外包" if task.is_outsource else "原工序",
        "is_outsource": bool(task.is_outsource),
        "priority_rank": 0,
        "is_original": True,
        **({} if task.is_outsource else original_profile),
    }]
    for option in optional_operations.get((task.material, _activity_key(task.activity)), []):
        if option.is_outsource:
            options.append({
                "destination_type": "外包",
                "work_center": "外包",
                "resource_group": "外包",
                "unit_hours": 0.0,
                "unit_hours_source": "外包",
                "is_outsource": True,
                "priority_rank": option.priority_rank,
                "is_original": False,
                "capacity_calc_type": "外包",
                "hot_surface_type": task.hot_surface_type or _hot_surface_type_from_text(task.process_text),
                "process_group": option.process_group or task.process_group,
                "unit_capacity": 0.0,
                "batch_capacity": 0.0,
                "batch_cycle_hours": 0.0,
                "line_throughput_rate": 0.0,
                "residence_hours": 0.0,
            })
        else:
            profile = _capacity_profile_for_route(
                task,
                option=option,
                work_center=option.alternative_work_center,
                unit_hours=max(float(option.unit_hours), 0.0),
                capacities=capacities,
                config=config,
            )
            options.append({
                "destination_type": "可选工作中心",
                "work_center": option.alternative_work_center,
                "resource_group": option.alternative_resource_group,
                "unit_hours": max(float(profile["unit_hours"]), 0.0),
                "unit_hours_source": option.unit_hours_source,
                "is_outsource": False,
                "priority_rank": option.priority_rank,
                "is_original": False,
                **profile,
            })
    deduped: list[dict[str, Any]] = []
    seen: set[tuple[str, str, bool, float, str, str]] = set()
    for option in sorted(options, key=lambda item: (item["is_outsource"], item["priority_rank"], item["work_center"])):
        key = (
            str(option["destination_type"]),
            str(option["work_center"]),
            bool(option["is_outsource"]),
            round(float(option["unit_hours"]), 6),
            str(option.get("capacity_calc_type", "")),
            str(option.get("process_group", "")),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(option)
    return deduped


def _modeb_baseline_load_from_items(
    baseline_items: list[ScheduledOperation],
    *,
    config: FortuneBjConfig | None,
) -> dict[tuple[str, str], float]:
    load: dict[tuple[str, str], float] = {}
    for item in baseline_items:
        if item.task.is_outsource or item.task.missing_work_center:
            continue
        quantity = _modeb_quantity_units(item.task)
        unit_hours = _modeb_unit_hours(item.task)
        for period, _period_start, _period_end, hours in _modeb_item_period_load_segments(
            item,
            quantity * unit_hours,
            config=config,
        ):
            load[(period, item.task.work_center)] = load.get((period, item.task.work_center), 0.0) + hours
    return load


def _modeb_allocation_load(
    allocations: list[ModeBAllocation],
    *,
    config: FortuneBjConfig | None,
) -> dict[tuple[str, str], float]:
    load: dict[tuple[str, str], float] = {}
    for allocation in allocations:
        if allocation.is_outsource or allocation.is_unmaintained_work_center:
            continue
        for segment in _modeb_allocation_period_segments(allocation, config=config):
            key = (segment.period, allocation.destination_work_center)
            load[key] = load.get(key, 0.0) + segment.load_hours
    return load


def _allocation_to_scheduled_operation(allocation: ModeBAllocation) -> ScheduledOperation:
    source_task = allocation.source_item.task
    if allocation.is_unmaintained_work_center:
        start = allocation.period_start
        end = min(source_task.due_date, allocation.period_end)
        if end <= start:
            end = start
        duration_hours = 0.0
        note = "未维护工作中心，未参与ModeB产能优化；未维护负荷单独报告"
    elif allocation.is_outsource:
        end = source_task.due_date
        start = end - timedelta(hours=OUTSOURCE_DURATION_HOURS)
        duration_hours = OUTSOURCE_DURATION_HOURS
        note = (
            f"ModeB外包建议：{allocation.quantity}件，释放原工作中心"
            f"{allocation.original_released_hours:.2f}小时；按7天日历返回，不占用厂内产能"
        )
    else:
        start = allocation.period_start
        end = min(source_task.due_date, allocation.period_end)
        if end <= start:
            end = start + timedelta(hours=max(allocation.load_hours, 1.0))
        duration_hours = allocation.load_hours
        note = "ModeB周期内整数产品数量产能优化；不作为车间执行排程时间"
    task = replace(
        source_task,
        work_center=allocation.destination_work_center,
        resource_group=allocation.destination_resource_group,
        quantity=float(allocation.quantity),
        unit_hours=allocation.original_unit_hours if allocation.is_unmaintained_work_center else allocation.unit_hours,
        duration_hours=duration_hours,
        is_outsource=allocation.is_outsource,
        missing_work_center=allocation.is_unmaintained_work_center or source_task.missing_work_center,
        capacity_calc_type=allocation.capacity_calc_type,
        hot_surface_type=allocation.hot_surface_type,
        process_group=allocation.process_group,
        unit_capacity=allocation.unit_capacity,
    )
    return ScheduledOperation(
        task=task,
        start=start,
        end=end,
        on_time=True,
        tardy_hours=0.0,
        note=note,
        analysis_status="周期内产能优化",
        analysis_source="ModeB_OR-Tools",
        window_number=None,
        window_type="优化周期",
    )


def _original_allocation_for_item(
    item: ScheduledOperation,
    *,
    config: FortuneBjConfig | None,
) -> ModeBAllocation | None:
    quantity = _modeb_quantity_units(item.task)
    if quantity <= 0:
        return None
    period, period_start, period_end = _modeb_period_for_item(item, config=config)
    original_unit_hours = _modeb_unit_hours(item.task)
    is_outsource = bool(item.task.is_outsource)
    if item.task.missing_work_center:
        unmaintained_load = quantity * original_unit_hours
        return ModeBAllocation(
            period=period,
            period_start=period_start,
            period_end=period_end,
            source_item=item,
            quantity=quantity,
            destination_type=UNMAINTAINED_WORKCENTER,
            destination_work_center=UNMAINTAINED_WORKCENTER,
            destination_resource_group=UNMAINTAINED_WORKCENTER,
            unit_hours=0.0,
            unit_hours_source=UNMAINTAINED_WORKCENTER,
            load_hours=0.0,
            original_unit_hours=original_unit_hours,
            original_released_hours=0.0,
            extra_hours=0.0,
            unmaintained_load_hours=unmaintained_load,
            is_unmaintained_work_center=True,
            capacity_calc_type=UNMAINTAINED_WORKCENTER,
            hot_surface_type=item.task.hot_surface_type or _hot_surface_type_from_text(item.task.process_text),
            process_group=item.task.process_group or item.task.process_text,
            unit_capacity=item.task.unit_capacity,
        )
    load_hours = 0.0 if is_outsource else quantity * original_unit_hours
    return ModeBAllocation(
        period=period,
        period_start=period_start,
        period_end=period_end,
        source_item=item,
        quantity=quantity,
        destination_type="外包" if is_outsource else "原工作中心",
        destination_work_center="外包" if is_outsource else item.task.work_center,
        destination_resource_group="外包" if is_outsource else item.task.resource_group,
        unit_hours=0.0 if is_outsource else original_unit_hours,
        unit_hours_source="外包" if is_outsource else "原工序",
        load_hours=load_hours,
        original_unit_hours=original_unit_hours,
        original_released_hours=quantity * original_unit_hours if is_outsource else 0.0,
        extra_hours=0.0,
        is_outsource=is_outsource,
        capacity_calc_type="外包" if is_outsource else (item.task.capacity_calc_type or "普通工时"),
        hot_surface_type=item.task.hot_surface_type or _hot_surface_type_from_text(item.task.process_text),
        process_group=item.task.process_group or item.task.process_text,
        unit_capacity=item.task.unit_capacity,
        capacity_load_units=quantity * item.task.unit_capacity if not is_outsource else 0.0,
    )


def _optimize_modeb_integer_allocations(
    baseline_items: list[ScheduledOperation],
    capacities: dict[str, WorkCenterCapacity],
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
    *,
    config: FortuneBjConfig | None,
    progress: ProgressCallback | None = None,
) -> tuple[list[ScheduledOperation], list[ModeBAllocation], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Choose integer product quantities per route inside each ModeB analysis period."""
    from ortools.sat.python import cp_model

    started = time.perf_counter()
    scale = 100
    records: list[dict[str, Any]] = []
    baseline_load = _modeb_baseline_load_from_items(baseline_items, config=config)
    periods: set[str] = set()
    workcenters: set[str] = set(capacities)
    period_record_counts: dict[str, int] = {}
    candidate_decision_count = 0
    candidate_option_count = 0
    unmaintained_allocations: list[ModeBAllocation] = []

    for item in baseline_items:
        quantity = _modeb_quantity_units(item.task)
        if quantity <= 0:
            continue
        if item.task.missing_work_center:
            allocation = _original_allocation_for_item(item, config=config)
            if allocation is not None:
                unmaintained_allocations.append(allocation)
            continue
        period, period_start, period_end = _modeb_period_for_item(item, config=config)
        periods.add(period)
        period_record_counts[period] = period_record_counts.get(period, 0) + 1
        original_unit_hours = _modeb_unit_hours(item.task)
        options = _modeb_route_options(item.task, optional_operations, capacities, config)
        if len(options) > 1:
            candidate_decision_count += 1
            candidate_option_count += len(options)
        for option in options:
            if not option["is_outsource"]:
                workcenters.add(str(option["work_center"]))
        records.append({
            "item": item,
            "quantity": quantity,
            "period": period,
            "period_start": period_start,
            "period_end": period_end,
            "original_unit_hours": original_unit_hours,
            "options": options,
        })

    if progress is not None:
        progress(
            f"ModeB整数数量优化建模: 周期 {len(periods):,} 个，工作中心 {len(workcenters):,} 个，"
            f"输入工序 {len(records):,} 条，候选决策工序 {candidate_decision_count:,} 条，"
            f"候选方案 {candidate_option_count:,} 个"
        )
        for index, period in enumerate(sorted(period_record_counts), start=1):
            progress(f"ModeB优化周期{index}: {period} 工序 {period_record_counts[period]:,} 条")

    if not records:
        final_items = [_allocation_to_scheduled_operation(allocation) for allocation in unmaintained_allocations]
        stats = {
            "状态": "跳过",
            "说明": "没有可参与 ModeB 优化的工序；未维护工作中心工序已单独保留为未维护负荷" if unmaintained_allocations else "没有可参与 ModeB 优化的工序",
            "输入工序数": 0,
            "候选决策工序数": 0,
            "候选方案数": 0,
            "布尔变量数": 0,
            "炉次变量数": 0,
            "短缺变量数": 0,
            "周期数": 0,
            "工作中心数": len(capacities),
            "求解状态": "SKIPPED",
            "建模耗时秒": round(time.perf_counter() - started, 3),
            "求解耗时秒": 0,
            "总短缺小时": 0,
            "超载工作中心周期数": 0,
        }
        return final_items, unmaintained_allocations, [], [], _modeb_capacity_stats_rows(stats=stats, config=config)

    model = cp_model.CpModel()
    var_by_option: dict[tuple[int, int], Any] = {}
    load_terms: dict[tuple[str, str], list[Any]] = {}
    possible_load_units: dict[tuple[str, str], int] = {}
    batch_capacity_terms: dict[tuple[str, str, str], list[Any]] = {}
    possible_batch_capacity_units: dict[tuple[str, str, str], int] = {}
    batch_profiles: dict[tuple[str, str, str], dict[str, Any]] = {}
    batch_vars: dict[tuple[str, str, str], Any] = {}
    outsource_penalty_terms: list[Any] = []
    change_penalty_terms: list[Any] = []
    extra_hours_terms: list[Any] = []
    int_var_count = 0

    for record_index, record in enumerate(records):
        quantity = int(record["quantity"])
        original_unit_hours = float(record["original_unit_hours"])
        priority_factor = max(int(getattr(record["item"].task, "priority_rank", PRIORITY_NORMAL) or PRIORITY_NORMAL), 1)
        option_vars = []
        for option_index, option in enumerate(record["options"]):
            var = model.NewIntVar(0, quantity, f"q_{record_index}_{option_index}")
            var_by_option[(record_index, option_index)] = var
            option_vars.append(var)
            int_var_count += 1
            if option["is_outsource"]:
                outsource_penalty_terms.append(var * int(round(original_unit_hours * scale)) * priority_factor)
            else:
                work_center = str(option["work_center"])
                unit_hours = float(option["unit_hours"])
                calc_type = str(option.get("capacity_calc_type") or "普通工时")
                batch_capacity = float(option.get("batch_capacity") or 0.0)
                batch_cycle_hours = float(option.get("batch_cycle_hours") or 0.0)
                if calc_type == "批量处理" and batch_capacity > 0 and batch_cycle_hours > 0:
                    batch_key = (
                        str(record["period"]),
                        work_center,
                        str(option.get("process_group") or record["item"].task.process_text),
                    )
                    unit_capacity = _positive_or_default(float(option.get("unit_capacity") or 0.0), 1.0)
                    capacity_units_per_piece = max(int(round(unit_capacity * scale)), 1)
                    term = var * capacity_units_per_piece
                    batch_capacity_terms.setdefault(batch_key, []).append(term)
                    possible_batch_capacity_units[batch_key] = possible_batch_capacity_units.get(batch_key, 0) + quantity * capacity_units_per_piece
                    batch_profiles.setdefault(batch_key, option)
                else:
                    key = (str(record["period"]), work_center)
                    term = var * int(round(unit_hours * scale))
                    load_terms.setdefault(key, []).append(term)
                    possible_load_units[key] = possible_load_units.get(key, 0) + quantity * int(round(unit_hours * scale))
            if not option["is_original"]:
                change_penalty_terms.append(var * priority_factor)
                extra_unit_hours = max(float(option["unit_hours"]) - original_unit_hours, 0.0)
                if extra_unit_hours > 0:
                    extra_hours_terms.append(var * int(round(extra_unit_hours * scale)) * priority_factor)
        model.Add(sum(option_vars) == quantity)

    batch_var_count = 0
    for batch_key in sorted(batch_capacity_terms):
        period, work_center, _process_group = batch_key
        profile = batch_profiles[batch_key]
        batch_capacity_units = max(int(round(float(profile.get("batch_capacity") or 0.0) * scale)), 1)
        max_batches = max(math.ceil(possible_batch_capacity_units.get(batch_key, 0) / batch_capacity_units) + 1, 1)
        batch_var = model.NewIntVar(0, max_batches, f"batch_{batch_var_count}")
        batch_vars[batch_key] = batch_var
        batch_var_count += 1
        model.Add(sum(batch_capacity_terms[batch_key]) <= batch_var * batch_capacity_units)
        batch_cycle_units = max(int(round(float(profile.get("batch_cycle_hours") or 0.0) * scale)), 1)
        load_key = (period, work_center)
        load_terms.setdefault(load_key, []).append(batch_var * batch_cycle_units)
        possible_load_units[load_key] = possible_load_units.get(load_key, 0) + max_batches * batch_cycle_units

    shortage_vars: list[Any] = []
    overloaded_vars: list[Any] = []
    for key in sorted(load_terms):
        period, work_center = key
        capacity_hours = _period_capacity_hours(period, capacities.get(work_center), config=config)
        capacity_units = int(round(capacity_hours * scale))
        shortage_upper = max(possible_load_units.get(key, 0) + capacity_units + 1, 1)
        shortage = model.NewIntVar(0, shortage_upper, f"short_{len(shortage_vars)}")
        model.Add(sum(load_terms[key]) <= capacity_units + shortage)
        overloaded = model.NewBoolVar(f"over_{len(overloaded_vars)}")
        model.Add(shortage >= 1).OnlyEnforceIf(overloaded)
        model.Add(shortage == 0).OnlyEnforceIf(overloaded.Not())
        shortage_vars.append(shortage)
        overloaded_vars.append(overloaded)

    model.Minimize(
        sum(shortage_vars) * 1_000_000
        + sum(overloaded_vars) * 100_000
        + sum(outsource_penalty_terms) * 1_000
        + sum(change_penalty_terms) * 100
        + sum(extra_hours_terms) * 10
    )
    build_seconds = time.perf_counter() - started

    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = 8
    max_seconds = config.mode_b_solver_max_seconds if config else 60.0
    if max_seconds is not None and max_seconds > 0:
        solver.parameters.max_time_in_seconds = float(max_seconds)
    solve_started = time.perf_counter()
    status = solver.Solve(model)
    solve_seconds = time.perf_counter() - solve_started
    status_name = solver.StatusName(status)
    solved = status in (cp_model.OPTIMAL, cp_model.FEASIBLE)

    allocations: list[ModeBAllocation] = list(unmaintained_allocations)
    if solved:
        temp_allocations: list[dict[str, Any]] = []
        batch_group_units: dict[tuple[str, str, str], float] = {}
        batch_group_load_hours: dict[tuple[str, str, str], float] = {}
        for batch_key, batch_var in batch_vars.items():
            profile = batch_profiles[batch_key]
            batch_count = int(solver.Value(batch_var))
            batch_group_load_hours[batch_key] = batch_count * float(profile.get("batch_cycle_hours") or 0.0)
        for record_index, record in enumerate(records):
            item: ScheduledOperation = record["item"]
            original_unit_hours = float(record["original_unit_hours"])
            for option_index, option in enumerate(record["options"]):
                quantity = int(solver.Value(var_by_option[(record_index, option_index)]))
                if quantity <= 0:
                    continue
                is_outsource = bool(option["is_outsource"])
                calc_type = str(option.get("capacity_calc_type") or ("外包" if is_outsource else "普通工时"))
                unit_capacity = _positive_or_default(float(option.get("unit_capacity") or 0.0), 1.0)
                batch_key = None
                capacity_load_units = quantity * unit_capacity
                if (
                    not is_outsource
                    and calc_type == "批量处理"
                    and float(option.get("batch_capacity") or 0.0) > 0
                    and float(option.get("batch_cycle_hours") or 0.0) > 0
                ):
                    batch_key = (
                        str(record["period"]),
                        str(option["work_center"]),
                        str(option.get("process_group") or item.task.process_text),
                    )
                    batch_group_units[batch_key] = batch_group_units.get(batch_key, 0.0) + capacity_load_units
                unit_hours = 0.0 if is_outsource else float(option["unit_hours"])
                released_hours = quantity * original_unit_hours if is_outsource or not option["is_original"] else 0.0
                temp_allocations.append({
                    "record": record,
                    "item": item,
                    "option": option,
                    "quantity": quantity,
                    "is_outsource": is_outsource,
                    "unit_hours": unit_hours,
                    "released_hours": released_hours,
                    "capacity_load_units": capacity_load_units,
                    "batch_key": batch_key,
                    "original_unit_hours": original_unit_hours,
                })
        for pending in temp_allocations:
            record = pending["record"]
            item = pending["item"]
            option = pending["option"]
            quantity = int(pending["quantity"])
            is_outsource = bool(pending["is_outsource"])
            unit_hours = float(pending["unit_hours"])
            original_unit_hours = float(pending["original_unit_hours"])
            batch_key = pending["batch_key"]
            if is_outsource:
                load_hours = 0.0
            elif batch_key is not None:
                group_units = max(batch_group_units.get(batch_key, 0.0), 0.0)
                group_load = batch_group_load_hours.get(batch_key, 0.0)
                load_hours = group_load * float(pending["capacity_load_units"]) / group_units if group_units > 0 else 0.0
            else:
                load_hours = quantity * unit_hours
            allocations.append(ModeBAllocation(
                    period=str(record["period"]),
                    period_start=record["period_start"],
                    period_end=record["period_end"],
                    source_item=item,
                    quantity=quantity,
                    destination_type=str(option["destination_type"]),
                    destination_work_center="外包" if is_outsource else str(option["work_center"]),
                    destination_resource_group="外包" if is_outsource else str(option["resource_group"]),
                    unit_hours=unit_hours,
                    unit_hours_source=str(option.get("unit_hours_source") or ("外包" if is_outsource else "可选工序表填写")),
                    load_hours=load_hours,
                    original_unit_hours=original_unit_hours,
                    original_released_hours=float(pending["released_hours"]),
                    extra_hours=max(quantity * unit_hours - quantity * original_unit_hours, 0.0) if not is_outsource else 0.0,
                    is_outsource=is_outsource,
                    capacity_calc_type=str(option.get("capacity_calc_type") or ("外包" if is_outsource else "普通工时")),
                    hot_surface_type=str(option.get("hot_surface_type") or item.task.hot_surface_type or ""),
                    process_group=str(option.get("process_group") or item.task.process_group or ""),
                    unit_capacity=float(option.get("unit_capacity") or 0.0),
                    batch_capacity=float(option.get("batch_capacity") or 0.0),
                    batch_cycle_hours=float(option.get("batch_cycle_hours") or 0.0),
                    batch_count=(
                        int(solver.Value(batch_vars[batch_key])) if batch_key is not None and batch_key in batch_vars else 0
                    ),
                    line_throughput_rate=float(option.get("line_throughput_rate") or 0.0),
                    residence_hours=float(option.get("residence_hours") or 0.0),
                    capacity_load_units=float(pending["capacity_load_units"]),
                ))
    else:
        for item in baseline_items:
            if item.task.missing_work_center:
                continue
            allocation = _original_allocation_for_item(item, config=config)
            if allocation is not None:
                allocations.append(allocation)

    optimized_load = _modeb_allocation_load(allocations, config=config)
    summary = _build_capacity_optimization_summary(baseline_load, optimized_load, capacities, config=config)
    option_rows = _build_modeb_optional_allocation_rows(allocations, config=config)
    final_items = [_allocation_to_scheduled_operation(allocation) for allocation in allocations]
    final_items.sort(key=lambda item: (item.start, item.task.order_id, item.task.activity, item.task.work_center))

    total_shortage = sum(solver.Value(var) for var in shortage_vars) / scale if solved else None
    overloaded_count = sum(solver.Value(var) for var in overloaded_vars) if solved else None
    stats = {
        "状态": "成功" if solved else "失败",
        "说明": (
            "OR-Tools已按周期和整数产品数量分配原工作中心、可选工作中心和外包"
            if solved
            else "OR-Tools未在时间上限内返回可行分配，报告保留原工作中心路径"
        ),
        "输入工序数": len(records),
        "候选决策工序数": candidate_decision_count,
        "候选方案数": candidate_option_count,
        "布尔变量数": int_var_count + len(overloaded_vars),
        "炉次变量数": batch_var_count,
        "短缺变量数": len(shortage_vars),
        "周期数": len(periods),
        "工作中心数": len(workcenters),
        "求解状态": status_name,
        "建模耗时秒": round(build_seconds, 3),
        "求解耗时秒": round(solve_seconds, 3),
        "总短缺小时": round(total_shortage, 2) if total_shortage is not None else "",
        "超载工作中心周期数": overloaded_count if overloaded_count is not None else "",
    }
    if progress is not None:
        progress(
            f"ModeB整数数量优化完成: 状态 {status_name}，整数变量 {int_var_count:,}，"
            f"短缺变量 {len(shortage_vars):,}，求解 {round(solve_seconds, 3)} 秒，"
            f"总短缺 {stats['总短缺小时']} 小时"
        )
    return final_items, allocations, option_rows, summary, _modeb_capacity_stats_rows(stats=stats, config=config)


def _build_modeb_optional_allocation_rows(
    allocations: list[ModeBAllocation],
    *,
    config: FortuneBjConfig | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for allocation in allocations:
        if allocation.is_unmaintained_work_center:
            continue
        task = allocation.source_item.task
        if (allocation.destination_type == "原工作中心" and not allocation.is_outsource) or (
            task.is_outsource and allocation.is_outsource
        ):
            continue
        segments = _modeb_allocation_period_segments(allocation, config=config)
        for segment_index, segment in enumerate(segments, start=1):
            rows.append({
                "周期": segment.period,
                "周期日期跨度": _period_span_from_bounds(segment.period_start, segment.period_end),
                "订单": task.order_id,
                "需求来源": task.demand_source,
                "预测工艺路线来源": task.forecast_route_source,
                "物料": task.material,
                "活动": _activity_key(task.activity),
                "工序短文本": task.process_text,
                "紧急类型": task.urgent_type,
                "订单优先级": task.priority_rank,
                "优先级类型": task.priority_type,
                "优先级原因": task.priority_reason,
                "原始供给日期": (task.original_due_date or task.due_date).strftime("%Y-%m-%d"),
                "调整后供给日期": task.due_date.strftime("%Y-%m-%d"),
                "原工作中心": task.work_center,
                "建议工作中心": allocation.destination_work_center,
                "原资源组分类": task.resource_group,
                "建议资源组分类": allocation.destination_resource_group,
                "分配产品数量": allocation.quantity,
                "周期拆分行": f"{segment_index}/{len(segments)}",
                "原单位工时": round(allocation.original_unit_hours, 4),
                "建议单位工时": round(allocation.unit_hours, 4),
                "建议单位工时来源": allocation.unit_hours_source,
                "产能计算类型": allocation.capacity_calc_type,
                "热处/表处类型": allocation.hot_surface_type,
                "工艺兼容组": allocation.process_group,
                "单件容量占用": round(allocation.unit_capacity, 4),
                "容量占用": round(segment.capacity_load_units, 3),
                "容量占用总量": round(allocation.capacity_load_units, 3),
                "单炉容量": round(allocation.batch_capacity, 3) if allocation.batch_capacity else "",
                "单炉周期小时": round(allocation.batch_cycle_hours, 3) if allocation.batch_cycle_hours else "",
                "折算炉次": round(allocation.batch_count, 3) if allocation.batch_count else "",
                "流水线吞吐率": round(allocation.line_throughput_rate, 3) if allocation.line_throughput_rate else "",
                "单件在炉时间小时": round(allocation.residence_hours, 3) if allocation.residence_hours else "",
                "原工作中心减少负荷小时": round(segment.original_released_hours or segment.original_load_hours, 3),
                "建议工作中心增加负荷小时": 0 if allocation.is_outsource else round(segment.load_hours, 3),
                "外包释放本厂工时": round(segment.original_released_hours, 3) if allocation.is_outsource else 0,
                "额外工时": round(segment.extra_hours, 3),
                "是否外包": "是" if allocation.is_outsource else "否",
                "选择方式": "OR-Tools整数产品数量优化",
                "说明": "最小分配单位为1件产品；负荷小时按工序前推时间跨周期拆分；外包不占用厂内产能，按7天日历返回",
            })
    rows.sort(key=lambda row: (row["周期"], row["订单"], row["活动"], row["建议工作中心"]))
    return rows


def _build_modeb_order_allocation_rows(
    allocations: list[ModeBAllocation],
    *,
    config: FortuneBjConfig | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    split_counts: dict[tuple[str, float, str, int], int] = {}
    for allocation in allocations:
        task = allocation.source_item.task
        identity = _schedule_task_identity(task)
        split_counts[identity] = split_counts.get(identity, 0) + 1

    split_index: dict[tuple[str, float, str, int], int] = {}
    for allocation in sorted(
        allocations,
        key=lambda item: (
            item.source_item.task.order_id,
            item.source_item.task.material,
            item.source_item.task.activity,
            item.source_item.task.source_row or 0,
            item.period,
            item.destination_type,
            item.destination_work_center,
        ),
    ):
        task = allocation.source_item.task
        identity = _schedule_task_identity(task)
        split_index[identity] = split_index.get(identity, 0) + 1
        original_load_for_quantity = allocation.quantity * allocation.original_unit_hours
        changed = (
            allocation.is_outsource
            or allocation.destination_work_center != task.work_center
            or abs(allocation.unit_hours - allocation.original_unit_hours) > 1e-6
        )
        if allocation.is_unmaintained_work_center:
            action = "未维护工作中心，未参与产能优化"
        elif allocation.is_outsource:
            action = "转外包"
        elif changed:
            action = "转可选工作中心"
        else:
            action = "保留原工作中心"
        segments = _modeb_allocation_period_segments(allocation, config=config)
        for segment_index, segment in enumerate(segments, start=1):
            rows.append({
                "订单": task.order_id,
                "需求来源": task.demand_source,
                "预测月份": task.forecast_month,
                "预测周日": task.forecast_week_end,
                "预测图号": task.forecast_drawing,
                "预测工艺路线来源": task.forecast_route_source,
                "路线来源订单": task.route_source_order,
                "物料": task.material,
                "活动": _activity_key(task.activity),
                "工序短文本": task.process_text,
                "源文件行号": task.source_row or "",
                "原始供给日期": (task.original_due_date or task.due_date).strftime("%Y-%m-%d"),
                "调整后供给日期": task.due_date.strftime("%Y-%m-%d"),
                "需求日期": task.due_date.strftime("%Y-%m-%d"),
                "紧急类型": task.urgent_type,
                "是否紧急": "是" if task.urgent else "否",
                "订单优先级": task.priority_rank,
                "优先级类型": task.priority_type,
                "优先级原因": task.priority_reason,
                "是否手动紧急": "是" if task.manual_urgent else "否",
                "是否过期转入优化开始日期": "是" if task.adjusted_to_start_period else "否",
                "是否未维护工作中心": "是" if allocation.is_unmaintained_work_center else "否",
                "周期": segment.period,
                "周期粒度": _optimization_granularity(config),
                "周期日期跨度": _period_span_from_bounds(segment.period_start, segment.period_end),
                "原工序总产品数量": _modeb_quantity_units(task),
                "本行分配产品数量": allocation.quantity,
                "同工序分配行": f"{split_index[identity]}/{split_counts[identity]}",
                "同分配周期拆分行": f"{segment_index}/{len(segments)}",
                "原工作中心": task.work_center,
                "原资源组分类": task.resource_group,
                "原单位工时": round(allocation.original_unit_hours, 4),
                "原路径总负荷小时(按本行数量)": round(original_load_for_quantity, 3),
                "原路径本周期负荷小时": round(segment.original_load_hours, 3),
                "优化动作": action,
                "优化后工作中心": allocation.destination_work_center,
                "优化后资源组分类": allocation.destination_resource_group,
                "优化后单位工时": round(allocation.unit_hours, 4),
                "优化后单位工时来源": allocation.unit_hours_source,
                "产能计算类型": allocation.capacity_calc_type,
                "热处/表处类型": allocation.hot_surface_type,
                "工艺兼容组": allocation.process_group,
                "单件容量占用": round(allocation.unit_capacity, 4),
                "容量占用": round(segment.capacity_load_units, 3),
                "容量占用总量": round(allocation.capacity_load_units, 3),
                "单炉容量": round(allocation.batch_capacity, 3) if allocation.batch_capacity else "",
                "单炉周期小时": round(allocation.batch_cycle_hours, 3) if allocation.batch_cycle_hours else "",
                "折算炉次": round(allocation.batch_count, 3) if allocation.batch_count else "",
                "流水线吞吐率": round(allocation.line_throughput_rate, 3) if allocation.line_throughput_rate else "",
                "单件在炉时间小时": round(allocation.residence_hours, 3) if allocation.residence_hours else "",
                "本周期负荷小时": round(0.0 if allocation.is_unmaintained_work_center else segment.load_hours, 3),
                "优化后厂内负荷小时": round(0.0 if allocation.is_unmaintained_work_center else segment.load_hours, 3),
                "原工作中心释放小时": round(segment.original_load_hours if changed and not allocation.is_unmaintained_work_center else 0.0, 3),
                "未维护负荷小时": round(segment.unmaintained_load_hours, 3),
                "外包释放本厂工时": round(segment.original_released_hours, 3) if allocation.is_outsource else 0,
                "额外工时": round(segment.extra_hours, 3),
                "是否外包": "是" if allocation.is_outsource else "否",
                "外包返回日历天": 7 if allocation.is_outsource else "",
                "说明": (
                    "未维护工作中心，未参与ModeB可选工序/外包优化；未维护负荷按周期拆分后单独报告"
                    if allocation.is_unmaintained_work_center
                    else (
                        "最小分配单位为1件产品；负荷小时按工序前推时间跨周期拆分；同一工序可按整数件拆到多个路径"
                        if split_counts[identity] > 1
                        else "负荷小时按工序前推时间跨周期拆分；产品数量为涉及数量，不代表周期内完成数量"
                    )
                ),
            })
    return rows


def _optimize_capacity_route_choices(
    tasks: list[OperationTask],
    baseline: list[ScheduledOperation],
    bottleneck_report: list[dict[str, Any]],
    capacities: dict[str, WorkCenterCapacity],
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
    *,
    config: FortuneBjConfig | None,
    progress: ProgressCallback | None = None,
) -> tuple[list[OperationTask], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    """Use CP-SAT to choose original/alternative/outsource routes at period capacity level."""
    from ortools.sat.python import cp_model

    started = time.perf_counter()
    scale = 100
    bottleneck_periods = {
        (str(row.get("期间") or ""), str(row.get("工作中心") or ""))
        for row in bottleneck_report
        if row.get("是否瓶颈") == "是"
    }
    task_by_identity = {_schedule_task_identity(task): task for task in tasks}
    candidate_options: dict[tuple[str, float, str, int], list[dict[str, Any]]] = {}
    periods: set[str] = set()
    workcenters: set[str] = set(capacities)
    fixed_load: dict[tuple[str, str], float] = {}
    baseline_load: dict[tuple[str, str], float] = {}

    for task in tasks:
        identity = _schedule_task_identity(task)
        if task.is_outsource:
            continue
        period = _optimization_period_label(task, config)
        periods.add(period)
        workcenters.add(task.work_center)
        baseline_load[(period, task.work_center)] = baseline_load.get((period, task.work_center), 0.0) + task.duration_hours
        options = _capacity_route_options(task, optional_operations)
        if (period, task.work_center) in bottleneck_periods and len(options) > 1:
            candidate_options[identity] = options
            for option in options:
                if not option["is_outsource"]:
                    workcenters.add(str(option["work_center"]))
            continue
        fixed_load[(period, task.work_center)] = fixed_load.get((period, task.work_center), 0.0) + task.duration_hours

    if progress is not None:
        option_count = sum(len(options) for options in candidate_options.values())
        progress(
            f"ModeB路径选择优化: 候选工序 {len(candidate_options):,} 条，候选方案 {option_count:,} 个，"
            f"周期 {len(periods):,} 个，工作中心 {len(workcenters):,} 个"
        )

    if not candidate_options:
        stats = {
            "状态": "跳过",
            "说明": "瓶颈工序没有可选工序候选方案，保留原路径",
            "输入工序数": len(tasks),
            "候选决策工序数": 0,
            "候选方案数": 0,
            "布尔变量数": 0,
            "短缺变量数": 0,
            "周期数": len(periods),
            "工作中心数": len(workcenters),
            "求解状态": "SKIPPED",
            "建模耗时秒": round(time.perf_counter() - started, 3),
            "求解耗时秒": 0,
            "总短缺小时": 0,
            "超载工作中心周期数": 0,
        }
        summary = _build_capacity_optimization_summary(
            baseline_load,
            baseline_load,
            capacities,
            config=config,
        )
        return tasks, [], summary, _modeb_capacity_stats_rows(stats=stats, config=config)

    model = cp_model.CpModel()
    load_terms: dict[tuple[str, str], list[Any]] = {}
    selected_vars: dict[tuple[tuple[str, float, str, int], int], Any] = {}
    bool_var_count = 0
    outsource_penalty_terms: list[Any] = []
    change_penalty_terms: list[Any] = []
    extra_hours_terms: list[Any] = []

    for identity, options in candidate_options.items():
        original_task = task_by_identity[identity]
        period = _optimization_period_label(original_task, config)
        decision_vars = []
        for option_index, option in enumerate(options):
            var = model.NewBoolVar(f"x_{len(selected_vars)}")
            selected_vars[(identity, option_index)] = var
            decision_vars.append(var)
            bool_var_count += 1
            if option["is_outsource"]:
                outsource_penalty_terms.append(var * int(round(original_task.duration_hours * scale)))
                continue
            work_center = str(option["work_center"])
            load_terms.setdefault((period, work_center), []).append(var * int(round(float(option["load_hours"]) * scale)))
            if option["route_type"] != "原路径":
                change_penalty_terms.append(var)
                extra_hours = max(float(option["load_hours"]) - original_task.duration_hours, 0.0)
                if extra_hours > 0:
                    extra_hours_terms.append(var * int(round(extra_hours * scale)))
        model.Add(sum(decision_vars) == 1)

    shortage_vars: list[Any] = []
    overloaded_vars: list[Any] = []
    for period in periods:
        for work_center in workcenters:
            key = (period, work_center)
            capacity_hours = _period_capacity_hours(
                period,
                capacities.get(work_center),
                config=config,
            )
            fixed = int(round(fixed_load.get(key, 0.0) * scale))
            variable_terms = load_terms.get(key, [])
            if not variable_terms and fixed <= 0 and capacity_hours <= 0:
                continue
            capacity_units = int(round(capacity_hours * scale))
            shortage = model.NewIntVar(0, max(len(tasks) * 1000 * scale, capacity_units * 10 + fixed + 1), f"short_{period}_{abs(hash(work_center)) % 1_000_000}")
            model.Add(fixed + sum(variable_terms) <= capacity_units + shortage)
            overloaded = model.NewBoolVar(f"over_{period}_{abs(hash(work_center)) % 1_000_000}")
            model.Add(shortage >= 1).OnlyEnforceIf(overloaded)
            model.Add(shortage == 0).OnlyEnforceIf(overloaded.Not())
            shortage_vars.append(shortage)
            overloaded_vars.append(overloaded)

    model.Minimize(
        sum(shortage_vars) * 1_000_000
        + sum(overloaded_vars) * 100_000
        + sum(outsource_penalty_terms) * 100
        + sum(extra_hours_terms) * 50
        + sum(change_penalty_terms) * 10
    )
    build_seconds = time.perf_counter() - started
    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = 8
    max_seconds = config.mode_b_solver_max_seconds if config else 60.0
    if max_seconds is not None and max_seconds > 0:
        solver.parameters.max_time_in_seconds = float(max_seconds)
    solve_started = time.perf_counter()
    status = solver.Solve(model)
    solve_seconds = time.perf_counter() - solve_started
    status_name = solver.StatusName(status)

    selected_by_identity: dict[tuple[str, float, str, int], dict[str, Any]] = {}
    option_rows: list[dict[str, Any]] = []
    optimized_load = dict(fixed_load)
    if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        for identity, options in candidate_options.items():
            original_task = task_by_identity[identity]
            period = _optimization_period_label(original_task, config)
            selected_option = options[0]
            for option_index, option in enumerate(options):
                if solver.Value(selected_vars[(identity, option_index)]) == 1:
                    selected_option = option
                    break
            selected_by_identity[identity] = selected_option
            if not selected_option["is_outsource"]:
                key = (period, str(selected_option["work_center"]))
                optimized_load[key] = optimized_load.get(key, 0.0) + float(selected_option["load_hours"])
            if selected_option["route_type"] == "原路径":
                continue
            option_rows.append({
                "周期": period,
                "周期日期跨度": _period_date_span(period, config),
                "订单": original_task.order_id,
                "物料": original_task.material,
                "活动": _activity_key(original_task.activity),
                "工序短文本": original_task.process_text,
                "原工作中心": original_task.work_center,
                "建议工作中心": selected_option["work_center"],
                "原资源组分类": original_task.resource_group,
                "建议资源组分类": selected_option["resource_group"],
                "原工序工时": round(original_task.duration_hours, 3),
                "建议工序工时": round(float(selected_option.get("duration_hours", selected_option["load_hours"])), 3),
                "原工作中心减少负荷小时": round(original_task.duration_hours, 3),
                "建议工作中心增加负荷小时": 0 if selected_option["is_outsource"] else round(float(selected_option["load_hours"]), 3),
                "外包释放本厂工时": round(original_task.duration_hours, 3) if selected_option["is_outsource"] else 0,
                "是否外包": "是" if selected_option["is_outsource"] else "否",
                "选择方式": "OR-Tools全局产能平衡选择",
                "说明": "以最小化超100%产能缺口为首要目标，同时惩罚外包、额外工时和路径变更",
            })
    else:
        optimized_load = dict(baseline_load)

    adjusted_tasks: list[OperationTask] = []
    for task in tasks:
        selected_option = selected_by_identity.get(_schedule_task_identity(task))
        if selected_option is None or selected_option["route_type"] == "原路径":
            adjusted_tasks.append(task)
            continue
        new_duration = (
            OUTSOURCE_DURATION_HOURS
            if selected_option["is_outsource"]
            else max(task.quantity * float(selected_option["unit_hours"]), 0.01)
        )
        adjusted_tasks.append(replace(
            task,
            work_center="外包" if selected_option["is_outsource"] else str(selected_option["work_center"]),
            resource_group="外包" if selected_option["is_outsource"] else str(selected_option["resource_group"]),
            unit_hours=0.0 if selected_option["is_outsource"] else float(selected_option["unit_hours"]),
            duration_hours=new_duration,
            is_outsource=bool(selected_option["is_outsource"]),
        ))

    summary = _build_capacity_optimization_summary(
        baseline_load,
        optimized_load,
        capacities,
        config=config,
    )
    total_shortage = sum(solver.Value(var) for var in shortage_vars) / scale if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else None
    overloaded_count = sum(solver.Value(var) for var in overloaded_vars) if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else None
    stats = {
        "状态": "成功" if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else "失败",
        "说明": "OR-Tools已选择可选路径" if status in (cp_model.OPTIMAL, cp_model.FEASIBLE) else "OR-Tools未在时间上限内返回可行路径选择",
        "输入工序数": len(tasks),
        "候选决策工序数": len(candidate_options),
        "候选方案数": sum(len(options) for options in candidate_options.values()),
        "布尔变量数": bool_var_count + len(overloaded_vars),
        "短缺变量数": len(shortage_vars),
        "周期数": len(periods),
        "工作中心数": len(workcenters),
        "求解状态": status_name,
        "建模耗时秒": round(build_seconds, 3),
        "求解耗时秒": round(solve_seconds, 3),
        "总短缺小时": round(total_shortage, 2) if total_shortage is not None else "",
        "超载工作中心周期数": overloaded_count if overloaded_count is not None else "",
    }
    if progress is not None:
        progress(
            f"ModeB路径选择优化完成: 状态 {status_name}，候选工序 {len(candidate_options):,}，"
            f"候选方案 {sum(len(options) for options in candidate_options.values()):,}，"
            f"布尔变量 {stats['布尔变量数']:,}，短缺变量 {len(shortage_vars):,}，"
            f"求解 {round(solve_seconds, 3)} 秒，总短缺 {stats['总短缺小时']} 小时"
        )
    return adjusted_tasks, option_rows, summary, _modeb_capacity_stats_rows(stats=stats, config=config)


def _capacity_route_options(
    task: OperationTask,
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
) -> list[dict[str, Any]]:
    options = [{
        "route_type": "原路径",
        "work_center": task.work_center,
        "resource_group": task.resource_group,
        "unit_hours": task.unit_hours,
        "load_hours": task.duration_hours,
        "is_outsource": task.is_outsource,
        "priority_rank": 0,
    }]
    for option in optional_operations.get((task.material, _activity_key(task.activity)), []):
        new_duration = OUTSOURCE_DURATION_HOURS if option.is_outsource else max(task.quantity * option.unit_hours, 0.0)
        options.append({
            "route_type": "外包" if option.is_outsource else "可选路径",
            "work_center": "外包" if option.is_outsource else option.alternative_work_center,
            "resource_group": "外包" if option.is_outsource else option.alternative_resource_group,
            "unit_hours": 0.0 if option.is_outsource else option.unit_hours,
            "duration_hours": new_duration,
            "load_hours": 0.0 if option.is_outsource else new_duration,
            "is_outsource": option.is_outsource,
            "priority_rank": option.priority_rank,
        })
    return options


def _build_capacity_optimization_summary(
    baseline_load: dict[tuple[str, str], float],
    optimized_load: dict[tuple[str, str], float],
    capacities: dict[str, WorkCenterCapacity],
    *,
    config: FortuneBjConfig | None = None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    keys = sorted(set(baseline_load) | set(optimized_load), key=lambda item: (item[0], item[1]))
    for period, work_center in keys:
        capacity = capacities.get(work_center)
        capacity_hours = _period_capacity_hours(period, capacity, config=config)
        before = baseline_load.get((period, work_center), 0.0)
        after = optimized_load.get((period, work_center), 0.0)
        before_shortage = max(before - capacity_hours, 0.0)
        after_shortage = max(after - capacity_hours, 0.0)
        if before <= 0 and after <= 0:
            continue
        rows.append({
            "周期": period,
            "周期日期跨度": _period_date_span(period, config),
            "工作中心": work_center,
            "资源组分类": capacity.resource_group if capacity else "",
            "日历名称": capacity.calendar_name if capacity else "",
            "设备数量": capacity.quantity if capacity else 0,
            "周期产能小时": round(capacity_hours, 2),
            "优化前负荷小时": round(before, 2),
            "优化前负荷率": round(before / capacity_hours, 4) if capacity_hours > 0 else 0,
            "优化前缺口小时": round(before_shortage, 2),
            "优化后负荷小时": round(after, 2),
            "优化后负荷率": round(after / capacity_hours, 4) if capacity_hours > 0 else 0,
            "优化后缺口小时": round(after_shortage, 2),
            "缺口改善小时": round(before_shortage - after_shortage, 2),
            "状态": "仍超100%" if after_shortage > 0 else "已压回100%以内",
        })
    rows.sort(key=lambda row: (-float(row["优化前缺口小时"]), -float(row["优化后缺口小时"]), row["周期"], row["工作中心"]))
    return rows


def _modeb_capacity_stats_rows(*, stats: dict[str, Any], config: FortuneBjConfig | None) -> list[dict[str, Any]]:
    return [{
        "类型": "本次运行",
        "场景": "ModeB整数数量周期优化",
        "工序数": stats.get("输入工序数", ""),
        "工作中心数": stats.get("工作中心数", ""),
        "周期数": stats.get("周期数", ""),
        "每工序候选方案数": (
            round(float(stats.get("候选方案数", 0)) / float(stats.get("候选决策工序数", 1)), 2)
            if stats.get("候选决策工序数") else 0
        ),
        "候选决策工序数": stats.get("候选决策工序数", ""),
        "候选方案数": stats.get("候选方案数", ""),
        "整数/布尔变量约数": stats.get("布尔变量数", ""),
        "炉次变量数": stats.get("炉次变量数", 0),
        "短缺变量数": stats.get("短缺变量数", ""),
        "求解时间上限秒": config.mode_b_solver_max_seconds if config else 60.0,
        "求解状态": stats.get("求解状态", ""),
        "建模耗时秒": stats.get("建模耗时秒", ""),
        "求解耗时秒": stats.get("求解耗时秒", ""),
        "总短缺小时": stats.get("总短缺小时", ""),
        "超载工作中心周期数": stats.get("超载工作中心周期数", ""),
        "说明": stats.get("说明", ""),
    }]


def _build_monthly_capacity_report(
    infinite_items: list[ScheduledOperation],
    finite_items: list[ScheduledOperation],
    capacities: dict[str, WorkCenterCapacity],
    *,
    config: FortuneBjConfig | None = None,
) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, str], dict[str, Any]] = {}

    # ModeA is an infinite-capacity pressure analysis.  It should not carry the
    # old ModeB rolling-window "unsimulated baseline" columns into the report.
    _ = finite_items

    def add_item(item: ScheduledOperation, hours: float, period: str) -> None:
        if item.task.is_outsource or item.task.missing_work_center:
            return
        key = (period, item.task.work_center)
        bucket = buckets.setdefault(key, {
            "周期": period,
            "周期粒度": _optimization_granularity(config),
            "周期日期跨度": _period_date_span(period, config),
            "工作中心": item.task.work_center,
            "资源组分类": item.task.resource_group,
            "无限产能工序数": 0,
            "无限产能负荷小时": 0.0,
        })
        bucket["无限产能工序数"] += 1
        bucket["无限产能负荷小时"] += hours

    for item in infinite_items:
        for period, hours in _split_item_hours_by_reporting_period(item, config=config):
            add_item(item, hours, period)

    rows: list[dict[str, Any]] = []
    for (period, work_center), bucket in sorted(buckets.items()):
        capacity = capacities.get(work_center)
        capacity_hours = _period_capacity_hours(period, capacity, config=config)
        infinite_load = float(bucket["无限产能负荷小时"])
        rows.append({
            "周期": bucket["周期"],
            "周期粒度": bucket["周期粒度"],
            "周期日期跨度": bucket["周期日期跨度"],
            "工作中心": bucket["工作中心"],
            "资源组分类": bucket["资源组分类"],
            "日历名称": capacity.calendar_name if capacity else "",
            "设备数量": capacity.quantity if capacity else 0,
            "平均每日小时/台": round(capacity.daily_hours, 2) if capacity else 0,
            "周期产能小时": round(capacity_hours, 2),
            "无限产能工序数": bucket["无限产能工序数"],
            "无限产能负荷小时": round(infinite_load, 2),
            "无限产能负荷率": round(infinite_load / capacity_hours, 4) if capacity_hours > 0 else 0,
            "产能缺口小时(无限口径)": round(max(infinite_load - capacity_hours, 0.0), 2),
        })
    return rows


def _build_modeb_period_capacity_report_from_allocations(
    baseline_items: list[ScheduledOperation],
    allocations: list[ModeBAllocation],
    capacities: dict[str, WorkCenterCapacity],
    *,
    config: FortuneBjConfig | None,
) -> list[dict[str, Any]]:
    granularity = _optimization_granularity(config)
    before: dict[tuple[str, str], dict[str, Any]] = {}
    after: dict[tuple[str, str], dict[str, Any]] = {}
    outsource_release: dict[tuple[str, str], float] = {}

    for item in baseline_items:
        if item.task.is_outsource or item.task.missing_work_center:
            continue
        quantity = _modeb_quantity_units(item.task)
        load_hours = quantity * _modeb_unit_hours(item.task)
        for period, period_start, period_end, segment_hours in _modeb_item_period_load_segments(
            item,
            load_hours,
            config=config,
        ):
            key = (period, item.task.work_center)
            bucket = before.setdefault(key, {
                "period_start": period_start,
                "period_end": period_end,
                "resource_group": item.task.resource_group,
                "operation_count": 0,
                "quantity": 0,
                "load_hours": 0.0,
            })
            bucket["operation_count"] += 1
            bucket["quantity"] += quantity
            bucket["load_hours"] += segment_hours

    for allocation in allocations:
        if allocation.is_unmaintained_work_center:
            continue
        for segment in _modeb_allocation_period_segments(allocation, config=config):
            original_key = (segment.period, allocation.source_item.task.work_center)
            if allocation.is_outsource:
                outsource_release[original_key] = outsource_release.get(original_key, 0.0) + segment.original_released_hours
                continue
            key = (segment.period, allocation.destination_work_center)
            bucket = after.setdefault(key, {
                "period_start": segment.period_start,
                "period_end": segment.period_end,
                "resource_group": allocation.destination_resource_group,
                "operation_count": 0,
                "quantity": 0,
                "load_hours": 0.0,
            })
            bucket["operation_count"] += 1
            bucket["quantity"] += allocation.quantity
            bucket["load_hours"] += segment.load_hours

    keys = sorted(set(before) | set(after) | set(outsource_release), key=lambda item: (item[0], item[1]))
    rows: list[dict[str, Any]] = []
    for period, work_center in keys:
        capacity = capacities.get(work_center)
        before_bucket = before.get((period, work_center), {})
        after_bucket = after.get((period, work_center), {})
        start, end = _period_bounds_from_label(period, config)
        period_start = before_bucket.get("period_start") or after_bucket.get("period_start") or start
        period_end = before_bucket.get("period_end") or after_bucket.get("period_end") or end
        capacity_hours = _period_capacity_hours(period, capacity, config=config)
        before_load = float(before_bucket.get("load_hours", 0.0))
        after_load = float(after_bucket.get("load_hours", 0.0))
        before_shortage = max(before_load - capacity_hours, 0.0)
        after_shortage = max(after_load - capacity_hours, 0.0)
        days = max((period_end - period_start).days, 1)
        equipment_quantity = capacity.quantity if capacity else 0
        extra_daily_hours = after_shortage / max(days * equipment_quantity, 1) if after_shortage > 0 else 0.0
        added_equipment = max(math.ceil(after_load / capacity_hours) - 1, 0) if capacity_hours > 0 and after_shortage > 0 else 0
        resource_group = (
            after_bucket.get("resource_group")
            or before_bucket.get("resource_group")
            or (capacity.resource_group if capacity else "")
        )
        rows.append({
            "周期": period,
            "周期粒度": granularity,
            "周期日期跨度": _period_span_from_bounds(period_start, period_end),
            "周期开始": period_start.strftime("%Y-%m-%d"),
            "周期结束": _period_display_end(period_end).strftime("%Y-%m-%d"),
            "工作中心": work_center,
            "资源组分类": resource_group,
            "日历名称": capacity.calendar_name if capacity else "",
            "设备数量": equipment_quantity,
            "平均每日小时/台": round(capacity.daily_hours, 2) if capacity else 0,
            "周期产能小时": round(capacity_hours, 2),
            "原始工序数": int(before_bucket.get("operation_count", 0)),
            "原始产品数量": int(before_bucket.get("quantity", 0)),
            "原始负荷小时": round(before_load, 2),
            "原始负荷率": round(before_load / capacity_hours, 4) if capacity_hours > 0 else 0,
            "原始缺口小时": round(before_shortage, 2),
            "优化后工序分配行数": int(after_bucket.get("operation_count", 0)),
            "优化后产品数量": int(after_bucket.get("quantity", 0)),
            "优化后负荷小时": round(after_load, 2),
            "优化后负荷率": round(after_load / capacity_hours, 4) if capacity_hours > 0 else 0,
            "优化后缺口小时": round(after_shortage, 2),
            "缺口改善小时": round(before_shortage - after_shortage, 2),
            "OR-Tools已转外包释放小时": round(outsource_release.get((period, work_center), 0.0), 2),
            "建议外包小时": round(after_shortage, 2),
            "建议加班小时": round(after_shortage, 2),
            "建议每日增加小时/台": round(extra_daily_hours, 2),
            "建议新增设备数": added_equipment,
            "状态": "仍超100%，需补足产能" if after_shortage > 0 else "已压回100%以内",
        })
    return rows


def _build_modeb_period_report_from_allocations(
    allocations: list[ModeBAllocation],
    capacities: dict[str, WorkCenterCapacity],
    *,
    config: FortuneBjConfig | None,
    max_window_tasks: int,
    solve_seconds: float,
    status: str,
) -> list[dict[str, Any]]:
    by_period: dict[str, dict[str, Any]] = {}
    load_by_period_wc = _modeb_allocation_load(allocations, config=config)
    for allocation in allocations:
        if allocation.is_unmaintained_work_center:
            continue
        task = allocation.source_item.task
        for segment in _modeb_allocation_period_segments(allocation, config=config):
            bucket = by_period.setdefault(segment.period, {
                "period_start": segment.period_start,
                "period_end": segment.period_end,
                "orders": set(),
                "source_operations": set(),
                "allocation_rows": 0,
                "quantity": 0,
                "outsource_quantity": 0,
                "outsource_release": 0.0,
            })
            bucket["orders"].add(task.order_id)
            bucket["source_operations"].add(_schedule_task_identity(task))
            bucket["allocation_rows"] += 1
            bucket["quantity"] += allocation.quantity
            if allocation.is_outsource:
                bucket["outsource_quantity"] += allocation.quantity
                bucket["outsource_release"] += segment.original_released_hours

    rows: list[dict[str, Any]] = []
    for index, period in enumerate(sorted(by_period), start=1):
        bucket = by_period[period]
        period_load = {
            work_center: hours
            for (load_period, work_center), hours in load_by_period_wc.items()
            if load_period == period
        }
        shortage = 0.0
        for work_center, hours in period_load.items():
            shortage += max(hours - _period_capacity_hours(period, capacities.get(work_center), config=config), 0.0)
        period_start = bucket["period_start"]
        period_end = bucket["period_end"]
        rows.append({
            "优化周期": index,
            "周期标签": period,
            "周期日期跨度": _period_span_from_bounds(period_start, period_end),
            "周期粒度": _optimization_granularity(config),
            "优化开始": period_start.strftime("%Y-%m-%d"),
            "优化结束": _period_display_end(period_end).strftime("%Y-%m-%d"),
            "优化周期天数": max((period_end - period_start).days, 1),
            "参考工序数/周期": max_window_tasks,
            "求解时间上限秒": config.mode_b_solver_max_seconds if config else 60.0,
            "订单数": len(bucket["orders"]),
            "原始工序数": len(bucket["source_operations"]),
            "分配行数": bucket["allocation_rows"],
            "分配产品数量": int(bucket["quantity"]),
            "外包产品数量": int(bucket["outsource_quantity"]),
            "外包释放本厂工时": round(bucket["outsource_release"], 2),
            "优化后总负荷小时": round(sum(period_load.values()), 2),
            "优化后总缺口小时": round(shortage, 2),
            "状态": status,
            "耗时秒": solve_seconds,
            "说明": "OR-Tools按工序起始优化周期做整数产品数量路径分配；本页负荷小时按工序前推时间跨周期拆分，产品数量为涉及数量，不代表周期内完成数量",
        })
    return rows


def _build_hot_surface_capacity_report(
    allocations: list[ModeBAllocation],
    capacities: dict[str, WorkCenterCapacity],
    *,
    config: FortuneBjConfig | None,
) -> list[dict[str, Any]]:
    if not _use_hot_surface_special_logic(config):
        return []
    buckets: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}
    for allocation in allocations:
        if allocation.is_outsource or allocation.is_unmaintained_work_center or allocation.capacity_calc_type not in {"批量处理", "流水线处理"}:
            continue
        for segment in _modeb_allocation_period_segments(allocation, config=config):
            key = (
                segment.period,
                allocation.destination_work_center,
                allocation.capacity_calc_type,
                allocation.hot_surface_type,
                allocation.process_group,
            )
            bucket = buckets.setdefault(key, {
                "period_start": segment.period_start,
                "period_end": segment.period_end,
                "resource_group": allocation.destination_resource_group,
                "orders": set(),
                "source_operations": set(),
                "quantity": 0,
                "capacity_units": 0.0,
                "load_hours": 0.0,
                "batch_capacity": allocation.batch_capacity,
                "batch_cycle_hours": allocation.batch_cycle_hours,
                "line_throughput_rate": allocation.line_throughput_rate,
                "residence_hours": allocation.residence_hours,
            })
            task = allocation.source_item.task
            bucket["orders"].add(task.order_id)
            bucket["source_operations"].add(_schedule_task_identity(task))
            bucket["quantity"] += allocation.quantity
            bucket["capacity_units"] += segment.capacity_load_units
            bucket["load_hours"] += segment.load_hours

    rows: list[dict[str, Any]] = []
    for (period, work_center, calc_type, hot_type, process_group), bucket in sorted(buckets.items()):
        capacity = capacities.get(work_center)
        period_start = bucket["period_start"]
        period_end = bucket["period_end"]
        capacity_hours = _period_capacity_hours(period, capacity, config=config)
        load_hours = float(bucket["load_hours"])
        batch_count = (
            math.ceil(float(bucket["capacity_units"]) / float(bucket["batch_capacity"]))
            if calc_type == "批量处理" and float(bucket["batch_capacity"] or 0) > 0
            else 0
        )
        rows.append({
            "周期": period,
            "周期日期跨度": _period_span_from_bounds(period_start, period_end),
            "工作中心": work_center,
            "资源组分类": bucket["resource_group"],
            "热处/表处类型": hot_type,
            "产能计算类型": calc_type,
            "工艺兼容组": process_group,
            "订单数": len(bucket["orders"]),
            "工序数": len(bucket["source_operations"]),
            "产品数量": int(bucket["quantity"]),
            "容量占用": round(float(bucket["capacity_units"]), 3),
            "单炉容量": round(float(bucket["batch_capacity"]), 3) if bucket["batch_capacity"] else "",
            "单炉周期小时": round(float(bucket["batch_cycle_hours"]), 3) if bucket["batch_cycle_hours"] else "",
            "折算炉次合计": batch_count if batch_count else "",
            "流水线吞吐率": round(float(bucket["line_throughput_rate"]), 3) if bucket["line_throughput_rate"] else "",
            "单件在炉时间小时": round(float(bucket["residence_hours"]), 3) if bucket["residence_hours"] else "",
            "负荷小时": round(load_hours, 2),
            "周期产能小时": round(capacity_hours, 2),
            "负荷率": round(load_hours / capacity_hours, 4) if capacity_hours > 0 else 0,
            "说明": "批量处理按周期内容量占用折算炉次；流水线处理按吞吐率折算负荷小时。该页不是炉次执行顺序。"
        })
    return rows


def _build_unmaintained_workcenter_report(
    items: list[ScheduledOperation],
    *,
    config: FortuneBjConfig | None,
) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, str], dict[str, Any]] = {}
    for item in items:
        task = item.task
        if not task.missing_work_center:
            continue
        period = _reporting_period_label_for_item(item, config=config)
        load_hours = float(task.duration_hours or 0.0)
        if load_hours <= 0:
            load_hours = max(float(task.quantity or 0.0) * float(task.unit_hours or 0.0), 0.0)
        key = (period, task.process_text or UNMAINTAINED_WORKCENTER)
        bucket = buckets.setdefault(key, {
            "周期": period,
            "周期粒度": _optimization_granularity(config),
            "周期日期跨度": _period_date_span(period, config),
            "工序短文本": task.process_text,
            "涉及订单": set(),
            "工序数": 0,
            "产品数量": 0.0,
            "单位工时合计/参考": 0.0,
            "未维护负荷小时": 0.0,
            "示例订单": task.order_id,
            "示例物料": task.material,
            "示例活动": _activity_key(task.activity),
            "源文件行号": [],
        })
        bucket["涉及订单"].add(task.order_id)
        bucket["工序数"] += 1
        bucket["产品数量"] += float(task.quantity or 0.0)
        bucket["单位工时合计/参考"] += float(task.unit_hours or 0.0)
        bucket["未维护负荷小时"] += load_hours
        if task.source_row:
            bucket["源文件行号"].append(str(task.source_row))

    rows: list[dict[str, Any]] = []
    for (_period, _process_text), bucket in sorted(buckets.items(), key=lambda item: (item[0][0], item[0][1])):
        rows.append({
            "周期": bucket["周期"],
            "周期粒度": bucket["周期粒度"],
            "周期日期跨度": bucket["周期日期跨度"],
            "工序短文本": bucket["工序短文本"],
            "涉及订单数": len(bucket["涉及订单"]),
            "工序数": bucket["工序数"],
            "产品数量": round(bucket["产品数量"], 3),
            "单位工时合计/参考": round(bucket["单位工时合计/参考"], 4),
            "未维护负荷小时": round(bucket["未维护负荷小时"], 3),
            "示例订单": bucket["示例订单"],
            "示例物料": bucket["示例物料"],
            "示例活动": bucket["示例活动"],
            "源文件行号": ", ".join(bucket["源文件行号"][:20]),
            "处理建议": "请在工作中心表补充该工序短文本对应的工作中心、资源组分类、设备数量和日历；当前负荷未计入真实工作中心产能率。",
        })
    return rows


def _build_input_maintenance_report(config: FortuneBjConfig) -> list[dict[str, Any]]:
    mode_note = (
        "选择“热处/表处专用逻辑”时以下字段参与计算；选择“同机加逻辑”时仍按数量*单位工时计算。"
    )
    rows = [
        {
            "输入文件": WC_TEMPLATE_NAME,
            "字段": "产能计算类型",
            "是否必填": "热处/表处专用逻辑下建议维护",
            "适用逻辑": "普通工时 / 批量处理 / 流水线处理",
            "维护说明": "批量炉填“批量处理”，履带/连续炉填“流水线处理”；为空默认普通工时。",
        },
        {
            "输入文件": WC_TEMPLATE_NAME,
            "字段": "热处表处类型",
            "是否必填": "建议维护",
            "适用逻辑": "热处 / 表处 / 普通",
            "维护说明": "用于报告分组，不直接决定是否进专用逻辑；是否热处/表处仍优先从工序短文本识别。",
        },
        {
            "输入文件": WC_TEMPLATE_NAME,
            "字段": "单炉容量、容量单位、单件容量占用默认值",
            "是否必填": "批量处理必填",
            "适用逻辑": "批量处理",
            "维护说明": "周期内按产品数量*单件容量占用计算容量占用，再除以单炉容量向上取整为炉次。",
        },
        {
            "输入文件": WC_TEMPLATE_NAME,
            "字段": "单炉周期小时、装卸/准备小时",
            "是否必填": "批量处理必填",
            "适用逻辑": "批量处理",
            "维护说明": "每炉占用小时=单炉周期小时+装卸/准备小时；用于周期负荷，不生成炉次顺序。",
        },
        {
            "输入文件": WC_TEMPLATE_NAME,
            "字段": "流水线吞吐率、吞吐率单位、单件在炉时间小时、换型时间小时",
            "是否必填": "流水线处理必填",
            "适用逻辑": "流水线处理",
            "维护说明": "负荷小时=容量占用/吞吐率+换型时间；单件在炉时间用于报告提示，不作为独占等待产能。",
        },
        {
            "输入文件": OPS_TEMPLATE_NAME,
            "字段": "工艺处理类型、工艺兼容组、单件容量占用、是否允许合炉、是否必须整单同批、热处/表处程序",
            "是否必填": "可选维护",
            "适用逻辑": "热处/表处专用逻辑",
            "维护说明": "用于覆盖工作中心默认值和后续批次兼容分析；当前版本按周期容量汇总，不做炉次执行排程。",
        },
        {
            "输入文件": "工艺路线-北京.xlsx / 工艺路线-沈阳.xlsx / 工艺路线-南通.xlsx",
            "字段": "Data sheet：B列物料编码、E列工作中心描述、F列工序编码、G列准备/H、H列人工/H、I列设备/H",
            "是否必填": "启用需求预测时建议维护",
            "适用逻辑": "需求预测导入",
            "维护说明": "预测需求按北京、沈阳、南通优先级匹配工艺路线；匹配不到的预测物料不参与计算，并进入数据质量报告。",
        },
        {
            "输入文件": OPTIONAL_OPS_TEMPLATE_NAME,
            "字段": "可选产能计算类型、可选工艺兼容组、可选单件容量占用、外包返回天数",
            "是否必填": "可选维护",
            "适用逻辑": "ModeB可选工序",
            "维护说明": "可选工作中心为空时仍按其工作中心默认参数；外包仍视作无限产能，当前返回天数默认7天。",
        },
        {
            "输入文件": "本次运行",
            "字段": "热处/表处模式",
            "是否必填": "GUI选择",
            "适用逻辑": config.hot_surface_mode,
            "维护说明": mode_note,
        },
    ]
    return rows


def _build_modeb_period_capacity_report(
    original_tasks: list[OperationTask],
    optimized_tasks: list[OperationTask],
    capacities: dict[str, WorkCenterCapacity],
    *,
    config: FortuneBjConfig | None,
) -> list[dict[str, Any]]:
    granularity = _optimization_granularity(config)
    original = _task_load_by_reporting_period(original_tasks, config=config)
    optimized = _task_load_by_reporting_period(optimized_tasks, config=config)
    period_load = _task_load_by_optimization_period(optimized_tasks, config=config)
    peak_by_period: dict[tuple[str, str], dict[str, float]] = {}
    for (period, work_center), load_hours in period_load.items():
        capacity_hours = _period_capacity_hours(period, capacities.get(work_center), config=config)
        shortage = max(load_hours - capacity_hours, 0.0)
        ratio = load_hours / capacity_hours if capacity_hours > 0 else 0.0
        peak = peak_by_period.setdefault((period, work_center), {"ratio": 0.0, "shortage": 0.0})
        peak["ratio"] = max(peak["ratio"], ratio)
        peak["shortage"] = max(peak["shortage"], shortage)
    keys = sorted(set(original) | set(optimized), key=lambda item: (item[0], item[1]))
    rows: list[dict[str, Any]] = []
    for period, work_center in keys:
        capacity = capacities.get(work_center)
        capacity_hours = _period_capacity_hours(period, capacity, config=config)
        before_bucket = original.get((period, work_center), {})
        after_bucket = optimized.get((period, work_center), {})
        resource_group = (
            after_bucket.get("资源组分类")
            or before_bucket.get("资源组分类")
            or (capacity.resource_group if capacity else "")
        )
        before_load = float(before_bucket.get("负荷小时", 0.0))
        after_load = float(after_bucket.get("负荷小时", 0.0))
        before_shortage = max(before_load - capacity_hours, 0.0)
        after_shortage = max(after_load - capacity_hours, 0.0)
        peak = peak_by_period.get((period, work_center), {"ratio": 0.0, "shortage": 0.0})
        start, end = _period_bounds_from_label(period, config)
        days = max((end - start).days, 1)
        quantity = capacity.quantity if capacity else 0
        recommendation_shortage = max(after_shortage, float(peak["shortage"]))
        extra_daily_hours = recommendation_shortage / max(days * quantity, 1) if recommendation_shortage > 0 else 0.0
        added_equipment = max(math.ceil(after_load / capacity_hours) - 1, 0) if capacity_hours > 0 and after_shortage > 0 else 0
        rows.append({
            "周期": period,
            "周期粒度": granularity,
            "周期日期跨度": _period_span_from_bounds(start, end),
            "周期开始": before_bucket.get("周期开始") or after_bucket.get("周期开始") or start.strftime("%Y-%m-%d"),
            "周期结束": before_bucket.get("周期结束") or after_bucket.get("周期结束") or _period_display_end(end).strftime("%Y-%m-%d"),
            "工作中心": work_center,
            "资源组分类": resource_group,
            "日历名称": capacity.calendar_name if capacity else "",
            "设备数量": quantity,
            "平均每日小时/台": round(capacity.daily_hours, 2) if capacity else 0,
            "周期产能小时": round(capacity_hours, 2),
            "原始工序数": int(before_bucket.get("工序数", 0)),
            "原始负荷小时": round(before_load, 2),
            "原始负荷率": round(before_load / capacity_hours, 4) if capacity_hours > 0 else 0,
            "原始缺口小时": round(before_shortage, 2),
            "优化后工序数": int(after_bucket.get("工序数", 0)),
            "优化后负荷小时": round(after_load, 2),
            "优化后负荷率": round(after_load / capacity_hours, 4) if capacity_hours > 0 else 0,
            "优化后缺口小时": round(after_shortage, 2),
            "缺口改善小时": round(before_shortage - after_shortage, 2),
            "建议外包小时": round(recommendation_shortage, 2),
            "建议加班小时": round(recommendation_shortage, 2),
            "建议每日增加小时/台": round(extra_daily_hours, 2),
            "建议新增设备数": added_equipment,
            "状态": "仍超100%，需补足产能" if recommendation_shortage > 0 else "已压回100%以内",
        })
    return rows


def _build_capacity_recommendation_report(summary_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row in summary_rows:
        shortage = float(row.get("优化后缺口小时") or 0)
        if shortage <= 0:
            continue
        capacity_hours = float(row.get("周期产能小时") or 0)
        optimized_load = float(row.get("优化后负荷小时") or 0)
        added_equipment = max(math.ceil(optimized_load / capacity_hours) - 1, 0) if capacity_hours > 0 else ""
        rows.append({
            "周期": row.get("周期", ""),
            "周期日期跨度": row.get("周期日期跨度", ""),
            "工作中心": row.get("工作中心", ""),
            "资源组分类": row.get("资源组分类", ""),
            "周期产能小时": row.get("周期产能小时", ""),
            "优化后负荷小时": row.get("优化后负荷小时", ""),
            "优化后负荷率": row.get("优化后负荷率", ""),
            "剩余缺口小时": round(shortage, 2),
            "建议外包小时": round(shortage, 2),
            "建议加班小时": round(shortage, 2),
            "建议新增设备数": added_equipment,
            "建议排班调整": f"本周期至少补充 {shortage:.2f} 小时产能，可通过加班、临时班次或提高日历工作时间覆盖。",
            "建议设备投资": (
                f"如该缺口连续出现，按当前负荷约需新增 {added_equipment} 台同类设备。"
                if isinstance(added_equipment, int) and added_equipment > 0
                else "当前缺口优先用外包或排班调整处理。"
            ),
            "说明": "可选工序优化后仍超过100%，该缺口不跨周期隐藏，需管理动作补足。",
        })
    if rows:
        return rows
    return [{
        "周期": "",
        "周期日期跨度": "",
        "工作中心": "",
        "资源组分类": "",
        "周期产能小时": 0,
        "优化后负荷小时": 0,
        "优化后负荷率": 0,
        "剩余缺口小时": 0,
        "建议外包小时": 0,
        "建议加班小时": 0,
        "建议新增设备数": 0,
        "建议排班调整": "当前优化结果未发现超过100%的周期产能缺口。",
        "建议设备投资": "无需新增设备建议。",
        "说明": "无超产能缺口。",
    }]


def _build_modeb_period_report(
    original_tasks: list[OperationTask],
    optimized_tasks: list[OperationTask],
    capacities: dict[str, WorkCenterCapacity],
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
    bottleneck_workcenters: set[str],
    *,
    config: FortuneBjConfig | None,
    max_window_tasks: int,
    solve_seconds: float,
    status: str,
) -> list[dict[str, Any]]:
    by_period_orders: dict[str, set[str]] = {}
    by_period_tasks: dict[str, list[OperationTask]] = {}
    optimized_load = _task_load_by_optimization_period(optimized_tasks, config=config)
    for task in optimized_tasks:
        period = _optimization_period_label(task, config)
        by_period_orders.setdefault(period, set()).add(task.order_id)
        by_period_tasks.setdefault(period, []).append(task)
    rows: list[dict[str, Any]] = []
    for index, period in enumerate(sorted(by_period_tasks), start=1):
        period_start, period_end = _period_bounds_from_label(period, config)
        period_tasks = by_period_tasks[period]
        period_load = {
            work_center: hours
            for (load_period, work_center), hours in optimized_load.items()
            if load_period == period
        }
        shortage = 0.0
        for work_center, hours in period_load.items():
            shortage += max(hours - _period_capacity_hours(period, capacities.get(work_center), config=config), 0.0)
        row = {
            "优化周期": index,
            "周期标签": period,
            "周期日期跨度": _period_span_from_bounds(period_start, period_end),
            "周期粒度": _optimization_granularity(config),
            "优化开始": period_start.strftime("%Y-%m-%d"),
            "优化结束": _period_display_end(period_end).strftime("%Y-%m-%d"),
            "优化周期天数": max((period_end - period_start).days, 1),
            "参考工序数/周期": max_window_tasks,
            "求解时间上限秒": config.mode_b_solver_max_seconds if config else 60.0,
            "订单数": len(by_period_orders.get(period, set())),
            "工序数": len(period_tasks),
            "优化后总负荷小时": round(sum(period_load.values()), 2),
            "优化后总缺口小时": round(shortage, 2),
            "状态": status,
            "耗时秒": solve_seconds,
            "说明": (
                "已按周期内产能优化口径分析；超100%负荷保留在本周期并进入解决建议。"
                if len(period_tasks) <= max_window_tasks
                else f"周期工序数 {len(period_tasks):,} 超过参考规模 {max_window_tasks:,}；仅记录参考值，不按阈值跳过。"
            ),
        }
        row.update({
            f"当前{key}": value
            for key, value in _build_optional_pressure_summary(
                original_tasks,
                by_period_orders.get(period, set()),
                optional_operations,
                bottleneck_workcenters,
            ).items()
        })
        rows.append(row)
    return rows


def _month_days(month: str) -> int:
    try:
        year, month_number = (int(part) for part in month.split("-")[:2])
        return calendar.monthrange(year, month_number)[1]
    except (ValueError, IndexError):
        return 30


def _split_hours_by_reporting_period(
    start: datetime,
    end: datetime,
    total_hours: float,
    *,
    config: FortuneBjConfig | None = None,
) -> list[tuple[str, datetime, datetime, float]]:
    total_hours = max(float(total_hours or 0.0), 0.0)
    if total_hours <= 0:
        return []
    if end <= start:
        period = _period_label_from_start(_period_start_for_date(start, config), config)
        period_start, period_end = _period_bounds_from_label(period, config)
        return [(_period_label_from_start(period_start, config), period_start, period_end, total_hours)]

    elapsed_hours = (end - start).total_seconds() / 3600.0
    if elapsed_hours <= 0:
        period = _period_label_from_start(_period_start_for_date(start, config), config)
        period_start, period_end = _period_bounds_from_label(period, config)
        return [(_period_label_from_start(period_start, config), period_start, period_end, total_hours)]

    parts: list[tuple[str, datetime, datetime, float]] = []
    cursor = start
    while cursor < end:
        period = _period_label_from_start(_period_start_for_date(cursor, config), config)
        period_start, period_end = _period_bounds_from_label(period, config)
        segment_end = min(end, period_end)
        segment_elapsed_hours = (segment_end - cursor).total_seconds() / 3600.0
        if segment_elapsed_hours > 0:
            parts.append((
                period,
                period_start,
                period_end,
                total_hours * segment_elapsed_hours / elapsed_hours,
            ))
        cursor = segment_end
    return parts


def _modeb_item_period_load_segments(
    item: ScheduledOperation,
    total_hours: float,
    *,
    config: FortuneBjConfig | None,
) -> list[tuple[str, datetime, datetime, float]]:
    total_hours = max(float(total_hours or 0.0), 0.0)
    if total_hours <= 0:
        return []
    start = item.start
    modeb_start = _optimization_start_period(config)
    if modeb_start is not None and start < modeb_start:
        start = modeb_start
    end = item.end
    if end <= start:
        end = start + timedelta(hours=max(total_hours, 1.0))
    return _split_hours_by_reporting_period(start, end, total_hours, config=config)


def _modeb_allocation_period_segments(
    allocation: ModeBAllocation,
    *,
    config: FortuneBjConfig | None,
) -> list[ModeBAllocationPeriodSegment]:
    weights = _modeb_item_period_load_segments(allocation.source_item, 1.0, config=config)
    if not weights:
        weights = [(allocation.period, allocation.period_start, allocation.period_end, 1.0)]
    weight_total = sum(float(hours) for _period, _start, _end, hours in weights)
    if weight_total <= 0:
        weights = [(allocation.period, allocation.period_start, allocation.period_end, 1.0)]
        weight_total = 1.0

    original_load_hours = allocation.quantity * allocation.original_unit_hours
    segments: list[ModeBAllocationPeriodSegment] = []
    for period, period_start, period_end, weight_hours in weights:
        ratio = max(float(weight_hours) / weight_total, 0.0)
        segments.append(ModeBAllocationPeriodSegment(
            allocation=allocation,
            period=period,
            period_start=period_start,
            period_end=period_end,
            load_hours=allocation.load_hours * ratio,
            original_load_hours=original_load_hours * ratio,
            original_released_hours=allocation.original_released_hours * ratio,
            extra_hours=allocation.extra_hours * ratio,
            unmaintained_load_hours=allocation.unmaintained_load_hours * ratio,
            capacity_load_units=allocation.capacity_load_units * ratio,
        ))
    return segments


def _split_item_hours_by_reporting_period(
    item: ScheduledOperation,
    *,
    config: FortuneBjConfig | None = None,
) -> list[tuple[str, float]]:
    if _optimization_granularity(config) == "周":
        return _split_item_hours_by_week(item, config=config)
    return _split_item_hours_by_month(item)


def _operation_total_load_hours(task: OperationTask) -> float:
    if task.duration_hours > 0:
        return max(float(task.duration_hours), 0.0)
    return max(float(task.quantity or 0.0) * float(task.unit_hours or 0.0), 0.0)


def _split_item_hours_by_week(
    item: ScheduledOperation,
    *,
    config: FortuneBjConfig | None = None,
) -> list[tuple[str, float]]:
    if item.task.is_outsource or item.end <= item.start:
        return []
    parts: list[tuple[str, float]] = []
    elapsed_hours = (item.end - item.start).total_seconds() / 3600.0
    if elapsed_hours <= 0:
        return []
    load_hours = _operation_total_load_hours(item.task)
    cursor = item.start
    while cursor < item.end:
        next_week = _week_start(cursor) + timedelta(days=7)
        segment_end = min(item.end, next_week)
        segment_elapsed_hours = (segment_end - cursor).total_seconds() / 3600.0
        if segment_elapsed_hours > 0:
            parts.append((_period_label_from_start(_week_start(cursor), config), load_hours * segment_elapsed_hours / elapsed_hours))
        cursor = segment_end
    return parts


def _split_item_hours_by_month(item: ScheduledOperation) -> list[tuple[str, float]]:
    if item.task.is_outsource or item.end <= item.start:
        return []
    parts: list[tuple[str, float]] = []
    elapsed_hours = (item.end - item.start).total_seconds() / 3600.0
    if elapsed_hours <= 0:
        return []
    load_hours = _operation_total_load_hours(item.task)
    cursor = item.start
    while cursor < item.end:
        next_month = _month_add(_month_start(cursor), 1)
        segment_end = min(item.end, next_month)
        segment_elapsed_hours = (segment_end - cursor).total_seconds() / 3600.0
        if segment_elapsed_hours > 0:
            parts.append((cursor.strftime("%Y-%m"), load_hours * segment_elapsed_hours / elapsed_hours))
        cursor = segment_end
    return parts


def _scheduled_item_period_load_segments(
    item: ScheduledOperation,
    *,
    config: FortuneBjConfig | None = None,
) -> list[tuple[str, str, float]]:
    parts = _split_item_hours_by_reporting_period(item, config=config)
    if not parts:
        period = _reporting_period_label_for_item(item, config=config)
        return [(period, _period_date_span(period, config), round(_operation_total_load_hours(item.task), 3))]

    total_hours = round(_operation_total_load_hours(item.task), 3)
    rounded_parts = [(period, round(max(float(hours), 0.0), 3)) for period, hours in parts]
    diff = round(total_hours - sum(hours for _period, hours in rounded_parts), 3)
    if rounded_parts and abs(diff) >= 0.001:
        last_period, last_hours = rounded_parts[-1]
        rounded_parts[-1] = (last_period, round(max(last_hours + diff, 0.0), 3))
    return [
        (period, _period_date_span(period, config), hours)
        for period, hours in rounded_parts
    ]


def _schedule_detail_row_count(items: list[ScheduledOperation], *, config: FortuneBjConfig | None = None) -> int:
    return sum(max(len(_scheduled_item_period_load_segments(item, config=config)), 1) for item in items)


def _apply_optional_operations(
    tasks: list[OperationTask],
    bottleneck_workcenters: set[str],
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
) -> tuple[list[OperationTask], list[dict[str, Any]]]:
    adjusted: list[OperationTask] = []
    rows: list[dict[str, Any]] = []
    for task in tasks:
        if task.work_center not in bottleneck_workcenters:
            adjusted.append(task)
            continue
        selected = _select_optional_operation(task, optional_operations, bottleneck_workcenters)
        if selected is None:
            adjusted.append(task)
            continue
        new_duration = OUTSOURCE_DURATION_HOURS if selected.is_outsource else max(task.quantity * selected.unit_hours, 0.01)
        new_task = replace(
            task,
            work_center="外包" if selected.is_outsource else selected.alternative_work_center,
            resource_group="外包" if selected.is_outsource else selected.alternative_resource_group,
            unit_hours=0.0 if selected.is_outsource else selected.unit_hours,
            duration_hours=new_duration,
            is_outsource=selected.is_outsource,
        )
        adjusted.append(new_task)
        rows.append({
            "订单": task.order_id,
            "物料": task.material,
            "活动": _activity_key(task.activity),
            "工序短文本": task.process_text,
            "原工作中心": task.work_center,
            "可选工作中心": new_task.work_center,
            "原单位工时": task.unit_hours,
            "可选单位工时": 0 if selected.is_outsource else selected.unit_hours,
            "原工序工时": round(task.duration_hours, 3),
            "可选工序工时": round(new_duration, 3),
            "原工作中心减少负荷小时": round(task.duration_hours, 3),
            "可选工作中心增加负荷小时": round(0 if selected.is_outsource else new_duration, 3),
            "外包增加工时": OUTSOURCE_DURATION_HOURS if selected.is_outsource else 0,
            "是否外包": "是" if selected.is_outsource else "否",
            "说明": "原工作中心在基线负荷中识别为瓶颈，已按规则分流用于有限产能模拟",
        })
    return adjusted, rows


def _build_optional_pressure_summary(
    tasks: list[OperationTask],
    order_ids: set[str],
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
    bottleneck_workcenters: set[str],
) -> dict[str, Any]:
    if not order_ids:
        return {
            "原瓶颈负荷小时": 0,
            "可选工序数": 0,
            "可选路径负荷小时": 0,
            "外包工时": 0,
            "可选工作中心数": 0,
            "可选工作中心负荷TOP": "",
        }
    optional_count = 0
    original_bottleneck_hours = 0.0
    alternative_hours = 0.0
    outsource_hours = 0.0
    alternative_by_wc: dict[str, float] = {}
    for task in tasks:
        if task.order_id not in order_ids or task.work_center not in bottleneck_workcenters:
            continue
        original_bottleneck_hours += task.duration_hours
        selected = _select_optional_operation(task, optional_operations, bottleneck_workcenters)
        if selected is None:
            continue
        optional_count += 1
        if selected.is_outsource:
            outsource_hours += task.duration_hours
        else:
            new_duration = max(task.quantity * selected.unit_hours, 0.01)
            alternative_hours += new_duration
            alternative_by_wc[selected.alternative_work_center] = (
                alternative_by_wc.get(selected.alternative_work_center, 0.0) + new_duration
            )
    top = sorted(alternative_by_wc.items(), key=lambda item: (-item[1], item[0]))[:5]
    return {
        "原瓶颈负荷小时": round(original_bottleneck_hours, 2),
        "可选工序数": optional_count,
        "可选路径负荷小时": round(alternative_hours, 2),
        "外包工时": round(outsource_hours, 2),
        "可选工作中心数": len(alternative_by_wc),
        "可选工作中心负荷TOP": "；".join(f"{wc}:{hours:.1f}h" for wc, hours in top),
    }


def _select_optional_operation(
    task: OperationTask,
    optional_operations: dict[tuple[str, str], list[OptionalOperation]],
    bottleneck_workcenters: set[str],
) -> OptionalOperation | None:
    options = optional_operations.get((task.material, _activity_key(task.activity)), [])
    ranked_options = sorted(options, key=lambda item: (item.is_outsource, item.priority_rank, item.unit_hours))
    return next(
        (
            option
            for option in ranked_options
            if option.is_outsource
            or (
                option.alternative_work_center != task.work_center
                and option.alternative_work_center not in bottleneck_workcenters
            )
        ),
        None,
    )


def _build_bottleneck_report(
    scheduled: list[ScheduledOperation],
    capacities: dict[str, WorkCenterCapacity],
    *,
    config: FortuneBjConfig | None = None,
) -> list[dict[str, Any]]:
    load: dict[tuple[str, str], dict[str, Any]] = {}
    for item in scheduled:
        if item.task.is_outsource or item.task.missing_work_center:
            continue
        period = _reporting_period_label_for_item(item, config=config)
        key = (item.task.work_center, period)
        bucket = load.setdefault(key, {
            "工作中心": item.task.work_center,
            "期间": period,
            "期间日期跨度": _period_date_span(period, config),
            "资源组分类": item.task.resource_group,
            "工序数": 0,
            "负荷小时": 0.0,
        })
        bucket["工序数"] += 1
        bucket["负荷小时"] += item.task.duration_hours
    rows: list[dict[str, Any]] = []
    for (work_center, period), bucket in sorted(load.items(), key=lambda kv: (kv[0][1], kv[0][0])):
        capacity = capacities.get(work_center)
        capacity_hours = _period_capacity_hours(period, capacity, config=config)
        ratio = (bucket["负荷小时"] / capacity_hours) if capacity_hours > 0 else 0.0
        rows.append({
            "期间": period,
            "期间日期跨度": bucket["期间日期跨度"],
            "工作中心": work_center,
            "资源组分类": bucket["资源组分类"],
            "日历名称": capacity.calendar_name if capacity else "",
            "设备数量": capacity.quantity if capacity else 0,
            "平均每日小时/台": round(capacity.daily_hours, 2) if capacity else 0,
            "工序数": bucket["工序数"],
            "负荷小时": round(bucket["负荷小时"], 2),
            "产能小时": round(capacity_hours, 2),
            "负荷率": round(ratio, 4),
            "是否瓶颈": "是" if ratio > 1 else "否",
        })
    rows.sort(key=lambda row: (row["是否瓶颈"] != "是", -float(row["负荷率"]), row["期间"], row["工作中心"]))
    return rows


def _month_start(value: datetime) -> datetime:
    return datetime(value.year, value.month, 1)


def _day_start(value: datetime) -> datetime:
    return datetime(value.year, value.month, value.day)


def _month_add(value: datetime, months: int) -> datetime:
    month_index = value.year * 12 + value.month - 1 + months
    return datetime(month_index // 12, month_index % 12 + 1, 1)


def _month_in_range(value: datetime, start: datetime, end: datetime) -> bool:
    target = _month_start(value)
    return start <= target < end


def _reporting_period_label_for_item(item: ScheduledOperation, *, config: FortuneBjConfig | None = None) -> str:
    if _optimization_granularity(config) == "月":
        return item.start.strftime("%Y-%m")
    return _period_label_from_start(_week_start(item.start), config)


def _task_sort_key(task: OperationTask) -> tuple[str, float, int]:
    return (task.order_id, task.activity, task.source_row or 0)


def _schedule_task_identity(task: OperationTask) -> tuple[str, float, str, int]:
    return (task.order_id, task.activity, task.process_text, task.source_row or 0)


def _scheduled_item_by_identity(items: Iterable[ScheduledOperation]) -> dict[tuple[str, float, str, int], ScheduledOperation]:
    return {_schedule_task_identity(item.task): item for item in items}


def _candidate_window_orders_by_due_range_for_orders(
    tasks: list[OperationTask],
    candidate_order_ids: set[str],
    window_start: datetime,
    window_end: datetime,
) -> set[str]:
    by_order: dict[str, list[OperationTask]] = {}
    for task in tasks:
        by_order.setdefault(task.order_id, []).append(task)
    return {
        order_id
        for order_id, order_tasks in by_order.items()
        if order_id in candidate_order_ids
        and any(window_start <= task.due_date < window_end for task in order_tasks)
    }


def _filter_orders(tasks: list[OperationTask], order_ids: set[str]) -> list[OperationTask]:
    return sorted([task for task in tasks if task.order_id in order_ids], key=_task_sort_key)


def _status_text(status: int) -> str:
    try:
        from ortools.sat.python import cp_model
        mapping = {
            cp_model.OPTIMAL: "OPTIMAL",
            cp_model.FEASIBLE: "FEASIBLE",
            cp_model.INFEASIBLE: "INFEASIBLE",
            cp_model.MODEL_INVALID: "MODEL_INVALID",
            cp_model.UNKNOWN: "UNKNOWN",
        }
        return mapping.get(status, str(status))
    except Exception:
        return str(status)


def _schedule_tasks_with_ortools(
    tasks: list[OperationTask],
    capacities: dict[str, WorkCenterCapacity],
    fixed_items: list[ScheduledOperation] | None = None,
    max_seconds: float | None = None,
    release_time: datetime | None = None,
) -> list[ScheduledOperation]:
    """Schedule the production instance with OR-Tools CP-SAT."""
    from ortools.sat.python import cp_model

    fixed_items = fixed_items or []
    tasks_by_order: dict[str, list[OperationTask]] = {}
    for task in tasks:
        tasks_by_order.setdefault(task.order_id, []).append(task)
    for order_tasks in tasks_by_order.values():
        order_tasks.sort(key=lambda task: task.activity)
    fixed_by_order: dict[str, list[ScheduledOperation]] = {}
    for item in fixed_items:
        fixed_by_order.setdefault(item.task.order_id, []).append(item)
    for order_items in fixed_by_order.values():
        order_items.sort(key=lambda item: item.task.activity)

    durations = {
        id(task): _solver_duration_hours(task, capacities)
        for task in tasks
    }
    fixed_durations = {
        id(item): _scheduled_elapsed_hours(item)
        for item in fixed_items
    }
    total_duration = sum(durations.values())
    date_values = [task.due_date for task in tasks]
    date_values.extend(item.start for item in fixed_items)
    date_values.extend(item.end for item in fixed_items)
    if release_time is not None:
        date_values.append(release_time)
    earliest_due = min(date_values)
    latest_due = max(date_values)
    origin = earliest_due - timedelta(hours=total_duration + 24)
    horizon = max(int(math.ceil((latest_due - origin).total_seconds() / 3600.0)) + total_duration + 24, total_duration + 24)

    model = cp_model.CpModel()
    start_vars: dict[int, Any] = {}
    end_vars: dict[int, Any] = {}
    intervals_by_wc: dict[str, list[Any]] = {}
    demands_by_wc: dict[str, list[int]] = {}

    for index, task in enumerate(tasks):
        duration = durations[id(task)]
        start = model.NewIntVar(0, horizon, f"s_{index}")
        end = model.NewIntVar(0, horizon, f"e_{index}")
        model.Add(end == start + duration)
        if release_time is not None:
            release_offset = max(int(math.floor((release_time - origin).total_seconds() / 3600.0)), 0)
            model.Add(start >= release_offset)
        start_vars[index] = start
        end_vars[index] = end
        if not task.is_outsource:
            interval = model.NewIntervalVar(start, duration, end, f"i_{index}")
            intervals_by_wc.setdefault(task.work_center, []).append(interval)
            demands_by_wc.setdefault(task.work_center, []).append(1)

    task_index = {id(task): idx for idx, task in enumerate(tasks)}

    fixed_start_offsets: dict[int, int] = {}
    fixed_end_offsets: dict[int, int] = {}
    for item in fixed_items:
        fixed_start = max(int(math.floor((item.start - origin).total_seconds() / 3600.0)), 0)
        duration = fixed_durations[id(item)]
        fixed_start_offsets[id(item)] = fixed_start
        fixed_end_offsets[id(item)] = fixed_start + duration
        if item.task.is_outsource:
            continue
        interval = model.NewFixedSizeIntervalVar(fixed_start, duration, f"fixed_{id(item)}")
        intervals_by_wc.setdefault(item.task.work_center, []).append(interval)
        demands_by_wc.setdefault(item.task.work_center, []).append(1)

    for order_id in set(tasks_by_order) | set(fixed_by_order):
        steps: list[tuple[float, str, OperationTask | ScheduledOperation]] = []
        steps.extend((task.activity, "variable", task) for task in tasks_by_order.get(order_id, []))
        steps.extend((item.task.activity, "fixed", item) for item in fixed_by_order.get(order_id, []))
        steps.sort(key=lambda part: part[0])
        for previous, current in zip(steps, steps[1:]):
            prev_kind, prev_obj = previous[1], previous[2]
            curr_kind, curr_obj = current[1], current[2]
            if prev_kind == "variable" and curr_kind == "variable":
                prev_task = prev_obj
                curr_task = curr_obj
                model.Add(start_vars[task_index[id(curr_task)]] >= end_vars[task_index[id(prev_task)]])
            elif prev_kind == "fixed" and curr_kind == "variable":
                fixed_prev = prev_obj
                curr_task = curr_obj
                model.Add(start_vars[task_index[id(curr_task)]] >= fixed_end_offsets.get(id(fixed_prev), 0))
            elif prev_kind == "variable" and curr_kind == "fixed":
                prev_task = prev_obj
                fixed_curr = curr_obj
                model.Add(end_vars[task_index[id(prev_task)]] <= fixed_start_offsets.get(id(fixed_curr), horizon))

    for work_center, intervals in intervals_by_wc.items():
        capacity = capacities.get(work_center)
        model.AddCumulative(intervals, demands_by_wc[work_center], capacity.quantity if capacity else 1)

    objective_terms: list[Any] = []
    for order_id, order_tasks in tasks_by_order.items():
        final_task = order_tasks[-1]
        final_end = end_vars[task_index[id(final_task)]]
        due_offset = max(int(math.floor((final_task.due_date - origin).total_seconds() / 3600.0)), 0)
        tardiness = model.NewIntVar(0, horizon, f"tardy_{order_id}")
        late = model.NewBoolVar(f"late_{order_id}")
        model.Add(tardiness >= final_end - due_offset)
        model.Add(final_end <= due_offset).OnlyEnforceIf(late.Not())
        model.Add(final_end >= due_offset + 1).OnlyEnforceIf(late)
        urgent = any(task.urgent for task in order_tasks)
        late_penalty = 2_000_000 if urgent else 1_000_000
        objective_terms.append(late * late_penalty)
        objective_terms.append(tardiness * (20 if urgent else 10))
        objective_terms.append(final_end)
    model.Minimize(sum(objective_terms))

    solver = cp_model.CpSolver()
    solver.parameters.num_search_workers = 8
    if max_seconds is not None and max_seconds > 0:
        solver.parameters.max_time_in_seconds = float(max_seconds)
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError("OR-Tools 未找到可行有限产能模拟解")

    scheduled: list[ScheduledOperation] = []
    for index, task in enumerate(tasks):
        start = origin + timedelta(hours=int(solver.Value(start_vars[index])))
        end = origin + timedelta(hours=int(solver.Value(end_vars[index])))
        scheduled.append(ScheduledOperation(
            task=task,
            start=start,
            end=end,
            on_time=end <= task.due_date,
            tardy_hours=max((end - task.due_date).total_seconds() / 3600.0, 0.0),
            note="OR-Tools CP-SAT" if not task.is_outsource else "外协固定7天，不占用本地资源",
        ))
    return scheduled


def _solver_duration_hours(task: OperationTask, capacities: dict[str, WorkCenterCapacity]) -> int:
    if task.is_outsource:
        return OUTSOURCE_DURATION_HOURS
    capacity = capacities.get(task.work_center)
    daily_hours = capacity.daily_hours if capacity else 24.0
    if daily_hours <= 0:
        daily_hours = 24.0
    calendar_elapsed_hours = float(task.duration_hours) * 24.0 / min(daily_hours, 24.0)
    return max(int(math.ceil(calendar_elapsed_hours)), 1)


def _scheduled_elapsed_hours(item: ScheduledOperation) -> int:
    elapsed = (item.end - item.start).total_seconds() / 3600.0
    if elapsed <= 0:
        return OUTSOURCE_DURATION_HOURS if item.task.is_outsource else max(int(math.ceil(item.task.duration_hours)), 1)
    return max(int(math.ceil(elapsed)), 1)

def write_report(
    result: ScheduleResult,
    config: FortuneBjConfig,
    order_completion: dict[str, datetime],
    license_info: LicenseInfo | None = None,
    progress: ProgressCallback | None = None,
) -> Path:
    config.output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    mode = _normalize_mode(config.schedule_mode)
    is_mode_b = mode == "ModeB"
    path = config.output_dir / f"Fortune_BJ_产能分析报告_{mode}_{timestamp}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    report_started_at = time.perf_counter()
    schedule_detail_rows = 0 if is_mode_b else _schedule_detail_row_count(result.scheduled, config=config)
    total_report_rows = (
        5
        + schedule_detail_rows
        + max(len(result.bottleneck_report), 1)
        + max(len(result.monthly_capacity_report), 1)
        + (max(len(result.order_operation_allocation_report), 1) if is_mode_b else 0)
        + (max(len(result.window_report), 1) if is_mode_b else 0)
        + (len(result.optional_operation_report) if is_mode_b and result.optional_operation_report else 0)
        + (max(len(result.capacity_optimization_summary), 1) if is_mode_b else 0)
        + (max(len(result.capacity_recommendation_report), 1) if is_mode_b else 0)
        + (max(len(result.capacity_optimization_stats), 1) if is_mode_b else 0)
        + (max(len(result.hot_surface_capacity_report), 1) if _use_hot_surface_special_logic(config) else 0)
        + (max(len(result.unmaintained_workcenter_report), 1) if result.unmaintained_workcenter_report else 0)
        + max(len(result.input_maintenance_report), 1)
        + (max(len({str(row.get("订单") or "") for row in result.placeholder_due_orders}), 1) if result.placeholder_due_orders else 0)
        + (len({str(row.get("工序短文本") or "") for row in result.missing_mapping}) if result.missing_mapping else 0)
        + (max(len(result.data_issues), 1) if result.data_issues else 0)
    )
    written_rows = 0
    _emit_progress(progress, "写入报告", 0, total_report_rows, report_started_at, "初始化工作簿")
    _write_dashboard_sheet(wb, result, config, order_completion)
    written_rows += 1
    _emit_progress(progress, "写入报告", written_rows, total_report_rows, report_started_at, "仪表板")
    period_granularity = _optimization_granularity(config)
    capacity_label = f"{period_granularity}度产能分析"
    written_rows = _write_dict_sheet(
        wb,
        capacity_label,
        result.monthly_capacity_report,
        progress=progress,
        report_started_at=report_started_at,
        written_rows=written_rows,
        total_report_rows=total_report_rows,
    )
    _emit_progress(progress, "写入报告", written_rows, total_report_rows, report_started_at, capacity_label)
    written_rows = _write_bottleneck_analysis_sheet(wb, result.bottleneck_report, config, written_rows)
    _emit_progress(progress, "写入报告", written_rows, total_report_rows, report_started_at, "瓶颈分析")
    written_rows = _write_workcenter_heatmap_sheet(wb, result.monthly_capacity_report, config, written_rows)
    _emit_progress(progress, "写入报告", written_rows, total_report_rows, report_started_at, "工作组热力图")
    if is_mode_b:
        written_rows = _write_dict_sheet(
            wb,
            "订单工序分配明细",
            result.order_operation_allocation_report,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
    if not is_mode_b:
        written_rows = _write_schedule_sheet(
            wb,
            "订单工序分配明细",
            result.scheduled,
            include_outsource=False,
            config=config,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
    if result.placeholder_due_orders:
        written_rows = _write_placeholder_due_sheet(
            wb,
            result.placeholder_due_orders,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
    if is_mode_b:
        written_rows = _write_dict_sheet(
            wb,
            "ModeB优化周期明细",
            result.window_report,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
        written_rows = _write_dict_sheet(
            wb,
            "100%产能优化总览",
            result.capacity_optimization_summary,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
        if result.optional_operation_report:
            written_rows = _write_dict_sheet(
                wb,
                "可选工序分流分析",
                result.optional_operation_report,
                progress=progress,
                report_started_at=report_started_at,
                written_rows=written_rows,
                total_report_rows=total_report_rows,
            )
        written_rows = _write_dict_sheet(
            wb,
            "超产能解决建议",
            result.capacity_recommendation_report,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
        written_rows = _write_dict_sheet(
            wb,
            "OR-Tools本次求解规模",
            result.capacity_optimization_stats,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
        if _use_hot_surface_special_logic(config):
            written_rows = _write_dict_sheet(
                wb,
                "热处表处产能分析",
                result.hot_surface_capacity_report,
                progress=progress,
                report_started_at=report_started_at,
                written_rows=written_rows,
                total_report_rows=total_report_rows,
            )
    if result.unmaintained_workcenter_report:
        written_rows = _write_dict_sheet(
            wb,
            "未维护工作中心负荷汇总",
            result.unmaintained_workcenter_report,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
    written_rows = _write_dict_sheet(
        wb,
        "输入字段维护说明",
        result.input_maintenance_report,
        progress=progress,
        report_started_at=report_started_at,
        written_rows=written_rows,
        total_report_rows=total_report_rows,
    )
    if result.missing_mapping:
        written_rows = _write_missing_mapping_sheet(
            wb,
            result.missing_mapping,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
    if result.data_issues:
        written_rows = _write_dict_sheet(
            wb,
            "数据质量报告",
            result.data_issues,
            progress=progress,
            report_started_at=report_started_at,
            written_rows=written_rows,
            total_report_rows=total_report_rows,
        )
    _write_summary_sheet(wb, result, config, order_completion, license_info)
    written_rows += 1
    _emit_progress(progress, "写入报告", written_rows, total_report_rows, report_started_at, "调整列宽和保存文件")
    _autosize_workbook(wb)
    wb.save(path)
    _emit_progress(progress, "写入报告", total_report_rows, total_report_rows, report_started_at, f"保存完成：{path.name}")
    return path


def _write_dashboard_sheet(
    wb: Workbook,
    result: ScheduleResult,
    config: FortuneBjConfig,
    order_completion: dict[str, datetime],
) -> None:
    ws = wb.create_sheet("仪表板")
    mode = _normalize_mode(config.schedule_mode)
    all_items = result.scheduled + result.outsource
    order_ids = sorted({item.task.order_id for item in all_items})
    bottleneck_count = len({row.get("工作中心") for row in result.bottleneck_report if row.get("是否瓶颈") == "是"})
    window_success = sum(1 for row in result.window_report if row.get("状态") in {"成功", "已优化", "无可分流或无需分流"})
    window_failed = sum(1 for row in result.window_report if row.get("状态") == "失败")
    modeb_stats = next(
        (row for row in result.capacity_optimization_stats if row.get("类型") == "本次运行"),
        {},
    )
    shortage_improvement = sum(float(row.get("缺口改善小时") or 0) for row in result.capacity_optimization_summary)
    remaining_shortage = sum(float(row.get("优化后缺口小时") or 0) for row in result.capacity_optimization_summary)
    unmaintained_operations = sum(int(row.get("工序数") or 0) for row in result.unmaintained_workcenter_report)
    unmaintained_load = sum(float(row.get("未维护负荷小时") or 0) for row in result.unmaintained_workcenter_report)

    ws.append([f"执行摘要 - {mode} 产能分析"])
    modeb_start = _optimization_start_period(config)
    granularity = _optimization_granularity(config)
    ws.append([
        f"模式：{mode} | 运行时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
        f"优化粒度：{granularity} | "
        f"优化开始日期：{modeb_start.strftime('%Y-%m-%d') if modeb_start else '未指定'} | "
        f"工序流转：{_operation_flow_mode(config)} | "
        f"需求预测：{'启用' if config.enable_forecast else '未启用'} | "
        f"参考{config.mode_b_max_window_tasks}条工序/周期 | "
        f"工作日历：{Path(config.calendar_path).name if config.calendar_path else '默认日历'}"
    ])
    ws.append([])
    metrics = [
        ("分析订单数", len(order_ids), "订单"),
        ("厂内产能占用工序数", len(result.scheduled), "条"),
        ("外协模拟工序数", len(result.outsource), "条"),
        ("瓶颈工作中心数", bottleneck_count, "个"),
        ("可选工序分流数", len(result.optional_operation_report), "条"),
        ("ModeB整数分配候选工序", modeb_stats.get("候选决策工序数", 0), "条"),
        ("ModeB整数分配求解状态", modeb_stats.get("求解状态", ""), ""),
        ("ModeB产能缺口改善", round(shortage_improvement, 2), "小时"),
        ("ModeB剩余产能缺口", round(remaining_shortage, 2), "小时"),
        ("超产能解决建议", len(result.capacity_recommendation_report), "条"),
        ("ModeB优化周期数", len(result.window_report), "个"),
        ("ModeB异常周期数", window_failed, "个"),
        ("未维护工作中心工序数", unmaintained_operations, "条"),
        ("未维护工作中心负荷", round(unmaintained_load, 2), "小时"),
        (f"{_optimization_granularity(config)}度产能分析行数", len(result.monthly_capacity_report), "行"),
        ("占位交期订单数", len({str(row.get("订单") or "") for row in result.placeholder_due_orders}), "个"),
        ("数据质量/调整记录数", len(result.data_issues), "条"),
    ]
    ws.append(["核心指标", "数值", "单位"])
    for label, value, unit in metrics:
        ws.append([label, value, unit])
    ws.append([])
    ws.append(["重点瓶颈工作中心", "期间", "日期跨度", "负荷率", "负荷小时", "产能小时"])
    top_bottlenecks = [row for row in result.bottleneck_report if row.get("是否瓶颈") == "是"][:10]
    if not top_bottlenecks:
        ws.append(["无", "", "", "", "", ""])
    for row in top_bottlenecks:
        ws.append([
            row.get("工作中心", ""),
            row.get("期间", ""),
            row.get("期间日期跨度", ""),
            row.get("负荷率", ""),
            row.get("负荷小时", ""),
            row.get("产能小时", ""),
        ])
    ws.append([])
    ws.append(["运行判断", "说明"])
    ws.append(["ModeA", "无限产能倒排分析，用于识别需求负荷和潜在瓶颈；该口径允许负荷率超过100%。"])
    ws.append(["ModeB", "100%产能优化建议：先用 OR-Tools 在可选工序/外包路径中做周期级产能匹配；可选资源组分类为“外包”的路径视作无限产能且固定7天返回；优化后仍超100%的缺口保留在所选周/月周期内，并输出外包、加班和设备建议。"])
    _format_dashboard(ws)


def _write_monthly_trend_sheet(
    wb: Workbook,
    items: list[ScheduledOperation],
    order_completion: dict[str, datetime],
    config: FortuneBjConfig,
    written_rows: int,
) -> int:
    mode = _normalize_mode(config.schedule_mode)
    period_granularity = _optimization_granularity(config)
    period_header = "周期" if period_granularity == "周" else "月份"
    headers = [period_header, "日期跨度", "订单数", "厂内工序数", "外协工序数", "厂内负荷小时"]
    by_period: dict[str, dict[str, Any]] = {}
    for item in items:
        period = _reporting_period_label_for_item(item, config=config)
        bucket = by_period.setdefault(period, {
            "订单": set(),
            "厂内工序数": 0,
            "外协工序数": 0,
            "负荷小时": 0.0,
        })
        bucket["订单"].add(item.task.order_id)
        if item.task.is_outsource:
            bucket["外协工序数"] += 1
        else:
            bucket["厂内工序数"] += 1
            bucket["负荷小时"] += item.task.duration_hours
    rows = [
        [
            period,
            _period_date_span(period, config),
            len(bucket["订单"]),
            bucket["厂内工序数"],
            bucket["外协工序数"],
            round(bucket["负荷小时"], 2),
        ]
        for period, bucket in sorted(by_period.items())
    ]
    if not rows:
        rows = [["无", "", 0, 0, 0, 0]]
    _write_titled_table(
        wb,
        f"{period_granularity}度趋势",
        f"{period_granularity}度趋势 - {mode}",
        "按所选分析粒度汇总订单、工序和厂内负荷。",
        headers,
        rows,
    )
    return written_rows + len(rows)


def _write_bottleneck_analysis_sheet(
    wb: Workbook,
    rows: list[dict[str, Any]],
    config: FortuneBjConfig,
    written_rows: int,
) -> int:
    headers = ["期间", "期间日期跨度", "工作中心", "资源组分类", "日历名称", "设备数量", "平均每日小时/台", "工序数", "负荷小时", "产能小时", "负荷率", "是否瓶颈"]
    table_rows = [[row.get(header, "") for header in headers] for row in rows]
    if not table_rows:
        table_rows = [["无", "", "", "", "", 0, 0, 0, 0, 0, 0, "否"]]
    _write_titled_table(
        wb,
        "瓶颈分析",
        f"{_normalize_mode(config.schedule_mode)}瓶颈识别 - {_optimization_granularity(config)}度",
        "基于所选周/月分析粒度识别潜在瓶颈，负荷率允许超过100%。",
        headers,
        table_rows,
    )
    return written_rows + len(table_rows)


def _aggregate_heatmap_rows_by_resource_group(
    rows: list[dict[str, Any]],
    config: FortuneBjConfig,
) -> list[dict[str, Any]]:
    if not rows:
        return []
    is_period_capacity = bool("周期" in rows[0])
    is_monthly_capacity = bool("月份" in rows[0])
    is_modeb_capacity = bool((is_period_capacity or is_monthly_capacity) and "优化后负荷率" in rows[0])
    is_infinite_capacity = bool("无限产能负荷率" in rows[0])
    period_key = "周期" if is_period_capacity else ("月份" if is_monthly_capacity else "期间")
    buckets: dict[tuple[str, str], dict[str, Any]] = {}

    for row in rows:
        period = str(row.get(period_key) or "").strip()
        if not period:
            continue
        resource_group = str(row.get("资源组分类") or "").strip() or "未分组"
        bucket = buckets.setdefault((resource_group, period), {
            period_key: period,
            "工作组": resource_group,
            "周期产能小时": 0.0,
            "优化前负荷小时": 0.0,
            "优化后负荷小时": 0.0,
            "无限产能负荷小时": 0.0,
            "负荷小时": 0.0,
            "产能小时": 0.0,
        })
        capacity_hours = _to_number(row.get("周期产能小时"), default=_to_number(row.get("产能小时"), default=0.0))
        bucket["周期产能小时"] += capacity_hours
        bucket["产能小时"] += capacity_hours
        if is_modeb_capacity:
            bucket["优化前负荷小时"] += _to_number(
                row.get("优化前负荷小时"),
                default=_to_number(row.get("原始负荷小时"), default=0.0),
            )
            bucket["优化后负荷小时"] += _to_number(row.get("优化后负荷小时"), default=0.0)
        elif is_infinite_capacity:
            bucket["无限产能负荷小时"] += _to_number(row.get("无限产能负荷小时"), default=0.0)
        else:
            bucket["负荷小时"] += _to_number(row.get("负荷小时"), default=0.0)

    grouped_rows: list[dict[str, Any]] = []
    for (_resource_group, _period), bucket in sorted(buckets.items(), key=lambda item: (item[0][0], item[0][1])):
        capacity_hours = float(bucket["周期产能小时"])
        if is_modeb_capacity:
            before_load = float(bucket["优化前负荷小时"])
            after_load = float(bucket["优化后负荷小时"])
            grouped_rows.append({
                period_key: bucket[period_key],
                "工作组": bucket["工作组"],
                "周期产能小时": round(capacity_hours, 2),
                "优化前负荷小时": round(before_load, 2),
                "优化前负荷率": round(before_load / capacity_hours, 4) if capacity_hours > 0 else 0,
                "优化前缺口小时": round(max(before_load - capacity_hours, 0.0), 2),
                "优化后负荷小时": round(after_load, 2),
                "优化后负荷率": round(after_load / capacity_hours, 4) if capacity_hours > 0 else 0,
                "优化后缺口小时": round(max(after_load - capacity_hours, 0.0), 2),
            })
        elif is_infinite_capacity:
            infinite_load = float(bucket["无限产能负荷小时"])
            grouped_rows.append({
                period_key: bucket[period_key],
                "工作组": bucket["工作组"],
                "周期产能小时": round(capacity_hours, 2),
                "无限产能负荷小时": round(infinite_load, 2),
                "无限产能负荷率": round(infinite_load / capacity_hours, 4) if capacity_hours > 0 else 0,
                "产能缺口小时(无限口径)": round(max(infinite_load - capacity_hours, 0.0), 2),
            })
        else:
            load_hours = float(bucket["负荷小时"])
            grouped_rows.append({
                period_key: bucket[period_key],
                "工作组": bucket["工作组"],
                "产能小时": round(capacity_hours, 2),
                "负荷小时": round(load_hours, 2),
                "负荷率": round(load_hours / capacity_hours, 4) if capacity_hours > 0 else 0,
            })
    return grouped_rows


def _write_workcenter_heatmap_sheet(
    wb: Workbook,
    rows: list[dict[str, Any]],
    config: FortuneBjConfig,
    written_rows: int,
) -> int:
    rows = _aggregate_heatmap_rows_by_resource_group(rows, config)
    is_period_capacity = bool(rows and "周期" in rows[0])
    is_monthly_capacity = bool(rows and "月份" in rows[0])
    is_modeb_capacity = bool((is_period_capacity or is_monthly_capacity) and "优化后负荷率" in rows[0])
    is_infinite_capacity = bool(rows and "无限产能负荷率" in rows[0])
    period_granularity = _optimization_granularity(config)
    period_key = "周期" if is_period_capacity else ("月份" if is_monthly_capacity else "期间")
    periods = sorted({str(row.get(period_key) or "") for row in rows if row.get(period_key)})
    workgroups = sorted({str(row.get("工作组") or "") for row in rows if row.get("工作组")})
    lookup = {(str(row.get("工作组") or ""), str(row.get(period_key) or "")): row for row in rows}
    headers = ["工作组", "指标", *[_period_display_label(period, config) for period in periods]]
    table_rows: list[list[Any]] = []
    for group in workgroups:
        if is_modeb_capacity:
            table_rows.append([group, "优化后负荷率", *[lookup.get((group, period), {}).get("优化后负荷率", "") for period in periods]])
            table_rows.append([group, "优化后负荷小时", *[lookup.get((group, period), {}).get("优化后负荷小时", "") for period in periods]])
            table_rows.append([group, "优化后缺口小时", *[lookup.get((group, period), {}).get("优化后缺口小时", "") for period in periods]])
            table_rows.append([group, "周期产能小时", *[lookup.get((group, period), {}).get("周期产能小时", "") for period in periods]])
        elif is_infinite_capacity:
            table_rows.append([group, "无限产能负荷率", *[lookup.get((group, period), {}).get("无限产能负荷率", "") for period in periods]])
            table_rows.append([group, "无限产能负荷小时", *[lookup.get((group, period), {}).get("无限产能负荷小时", "") for period in periods]])
            table_rows.append([group, "周期产能小时", *[lookup.get((group, period), {}).get("周期产能小时", "") for period in periods]])
        else:
            table_rows.append([group, "负荷率", *[lookup.get((group, period), {}).get("负荷率", "") for period in periods]])
            table_rows.append([group, "负荷小时", *[lookup.get((group, period), {}).get("负荷小时", "") for period in periods]])
            table_rows.append([group, "产能小时", *[lookup.get((group, period), {}).get("产能小时", "") for period in periods]])
    if not table_rows:
        headers = ["工作组", "指标", "无数据"]
        table_rows = [["无", "优化后负荷率" if is_modeb_capacity else ("无限产能负荷率" if is_infinite_capacity else "负荷率"), ""]]
    ws = wb.create_sheet("工作组热力图")
    last_col = max(len(headers), 3)
    last_letter = get_column_letter(last_col)
    if is_modeb_capacity:
        ws.append([f"ModeB产能优化工作组热力图 - {_normalize_mode(config.schedule_mode)} | 按资源组分类汇总优化后负荷率、负荷小时、缺口小时"])
        ws.append(["周期汇总：热力图按工作组展示；周度/月度产能分析仍保留工作中心明细。"])
    elif is_infinite_capacity:
        ws.append([f"ModeA工作组热力图 - {_normalize_mode(config.schedule_mode)} | 按资源组分类汇总无限产能负荷率、负荷小时、产能小时"])
        ws.append([f"{period_granularity}度汇总：用于识别工作组压力；周度/月度产能分析仍保留工作中心明细。"])
    else:
        ws.append([f"工作组热力图 - {_normalize_mode(config.schedule_mode)} | 按资源组分类汇总负荷率、负荷小时、产能小时"])
        ws.append([f"{period_granularity}度汇总：用于识别工作组产能压力。"])
    ws.append(headers)
    for row in table_rows:
        ws.append(row)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    _format_workcenter_heatmap(ws, header_row=3)
    return written_rows + len(table_rows)


def _write_product_risk_sheet(
    wb: Workbook,
    items: list[ScheduledOperation],
    order_completion: dict[str, datetime],
    config: FortuneBjConfig,
    written_rows: int,
) -> int:
    grouped: dict[str, dict[str, Any]] = {}
    for item in items:
        key = item.task.material or "未填物料"
        bucket = grouped.setdefault(key, {
            "物料": key,
            "订单": set(),
            "工序数": 0,
            "厂内工序数": 0,
            "外协工序数": 0,
            "总工时": 0.0,
            "最早需求日期": item.task.due_date,
        })
        bucket["订单"].add(item.task.order_id)
        bucket["工序数"] += 1
        if item.task.is_outsource:
            bucket["外协工序数"] += 1
        else:
            bucket["厂内工序数"] += 1
            bucket["总工时"] += item.task.duration_hours
        if item.task.due_date < bucket["最早需求日期"]:
            bucket["最早需求日期"] = item.task.due_date
    rows = []
    for bucket in grouped.values():
        order_count = len(bucket["订单"])
        rows.append([
            bucket["物料"],
            order_count,
            bucket["工序数"],
            bucket["厂内工序数"],
            bucket["外协工序数"],
            round(bucket["总工时"], 2),
            bucket["最早需求日期"].strftime("%Y-%m-%d"),
        ])
    rows.sort(key=lambda row: (-float(row[5]), row[6], row[0]))
    if not rows:
        rows = [["无", 0, 0, 0, 0, 0, ""]]
    headers = ["物料", "订单数", "工序数", "厂内工序数", "外协工序数", "厂内总工时", "最早需求日期"]
    _write_titled_table(
        wb,
        "产品负荷分析",
        f"产品负荷分析 - {_normalize_mode(config.schedule_mode)}",
        "按物料汇总订单、工序、厂内负荷和外协工序，用于识别产品维度的产能压力。",
        headers,
        rows,
    )
    return written_rows + len(rows)


def _write_titled_table(
    wb: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append([title])
    ws.append([subtitle])
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _format_table(ws, header_row=3)


def _write_summary_sheet(
    wb: Workbook,
    result: ScheduleResult,
    config: FortuneBjConfig,
    order_completion: dict[str, datetime],
    license_info: LicenseInfo | None,
) -> None:
    ws = wb.create_sheet("运行信息")
    mode = _normalize_mode(config.schedule_mode)
    all_items = result.scheduled + result.outsource
    orders = {item.task.order_id for item in all_items}
    modeb_stats = next(
        (row for row in result.capacity_optimization_stats if row.get("类型") == "本次运行"),
        {},
    )
    modeb_start = _optimization_start_period(config)
    modeb_start_period = _period_label_from_start(modeb_start, config) if modeb_start else ""
    forecast_route_files = "；".join(filename for _source, filename in FORECAST_ROUTE_TEMPLATE_SPECS)
    unmatched_forecast_materials = {
        str(row.get("物料") or "")
        for row in result.data_issues
        if str(row.get("类型") or "") == "预测物料缺少工艺路线，未参与计算" and row.get("物料")
    }
    rows = [
        ("运行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("工具版本", APP_VERSION),
        ("生产订单工序表", str(config.operations_path)),
        ("订单交期数量表", str(config.demand_path)),
        ("工作中心表", str(config.workcenter_path)),
        ("工作日历表", str(config.calendar_path) if config.calendar_path else ""),
        ("热处/表处模式", config.hot_surface_mode),
        ("分析模式", mode),
        ("需求预测导入", "启用" if config.enable_forecast else "未启用"),
        ("需求预测表", str(config.forecast_path) if config.enable_forecast and config.forecast_path else ""),
        ("预测工艺路线优先级", "北京 > 沈阳 > 南通" if config.enable_forecast else ""),
        ("预测工艺路线文件", forecast_route_files if config.enable_forecast else ""),
        ("预测物料未匹配数", len(unmatched_forecast_materials) if config.enable_forecast else ""),
        ("工序流转逻辑", _operation_flow_mode(config)),
        ("优化粒度", _optimization_granularity(config)),
        ("优化开始日期", modeb_start.strftime("%Y-%m-%d") if modeb_start else ""),
        ("优化开始日期所在周期", modeb_start_period),
        ("优化开始日期所在周期跨度", _period_date_span(modeb_start_period, config)),
        ("ModeB参考工序数/周期", config.mode_b_max_window_tasks),
        ("ModeB求解时间上限(秒)", config.mode_b_solver_max_seconds if config.mode_b_solver_max_seconds else "不限制"),
        ("ModeB整数分配候选决策工序数", modeb_stats.get("候选决策工序数", "")),
        ("ModeB整数分配候选方案数", modeb_stats.get("候选方案数", "")),
        ("ModeB整数/布尔变量约数", modeb_stats.get("整数/布尔变量约数", modeb_stats.get("布尔变量约数", ""))),
        ("ModeB整数分配短缺变量数", modeb_stats.get("短缺变量数", "")),
        ("ModeB整数分配求解状态", modeb_stats.get("求解状态", "")),
        ("ModeB整数分配求解耗时秒", modeb_stats.get("求解耗时秒", "")),
        ("ModeB整数分配总短缺小时", modeb_stats.get("总短缺小时", "")),
        ("紧急类型优先级", "启用" if config.enable_urgent else "不启用"),
        ("分析订单数", len(orders)),
        ("厂内产能占用工序数", len(result.scheduled)),
        ("外协模拟工序数", len(result.outsource)),
        ("占位交期订单需求行数", len(result.placeholder_due_orders)),
        ("占位交期订单数", len({str(row.get("订单") or "") for row in result.placeholder_due_orders})),
        ("瓶颈工作中心报告行数", len(result.bottleneck_report)),
        ("ModeB优化周期数", len(result.window_report)),
        (f"{_optimization_granularity(config)}度产能分析行数", len(result.monthly_capacity_report)),
        ("100%产能优化总览行数", len(result.capacity_optimization_summary)),
        ("超产能解决建议行数", len(result.capacity_recommendation_report)),
        ("可选工序分流数", len(result.optional_operation_report)),
        ("未维护工作中心工序数", sum(int(row.get("工序数") or 0) for row in result.unmaintained_workcenter_report)),
        ("未维护工作中心负荷小时", round(sum(float(row.get("未维护负荷小时") or 0) for row in result.unmaintained_workcenter_report), 2)),
        ("缺失映射工序数", len(result.missing_mapping)),
        ("数据质量/调整记录数", len(result.data_issues)),
    ]
    if license_info is not None:
        rows.extend([
            ("授权状态", license_info.status),
            ("授权ID", license_info.license_id),
            ("授权类型", license_info.license_type),
            ("授权客户", license_info.customer_name),
            ("授权到期日", license_info.expiry_date),
            ("绑定方式", license_info.binding_mode),
            ("授权机器", license_info.machine_label),
            ("授权文件", license_info.license_path),
        ])
    ws.append(["模式", "参数", "值"])
    for row in rows:
        ws.append([mode, row[0], row[1]])
    _format_table(ws)


def _write_schedule_sheet(
    wb: Workbook,
    title: str,
    items: list[ScheduledOperation],
    include_outsource: bool,
    config: FortuneBjConfig | None = None,
    progress: ProgressCallback | None = None,
    report_started_at: float | None = None,
    written_rows: int = 0,
    total_report_rows: int = 0,
) -> int:
    ws = wb.create_sheet(title)
    headers = [
        "订单", "需求来源", "预测月份", "预测周日", "预测图号", "预测工艺路线来源", "路线来源订单", "活动", "物料", "工序短文本", "工作中心", "资源组分类", "订单数量",
        "单位工时(小时/pcs)", "工序生产时间(小时)", "本周期负荷小时", "开始时间", "完成时间", "需求日期",
        "原始供给日期", "调整后供给日期", "紧急类型", "是否紧急", "订单优先级", "优先级类型", "优先级原因",
        "是否手动紧急", "是否过期转入优化开始日期", "是否热处/表处", "是否外协", "是否未维护工作中心",
        "周期", "周期粒度", "周期日期跨度",
        "未维护负荷小时",
        "分析口径", "分析来源", "窗口编号", "窗口类型", "说明",
    ]
    ws.append(headers)
    sorted_items = sorted(items, key=lambda x: (x.start, x.task.order_id, x.task.activity))
    total_detail_rows = _schedule_detail_row_count(sorted_items, config=config)
    row_number = 0
    for item in sorted_items:
        total_load_hours = round(_operation_total_load_hours(item.task), 3)
        for period, period_span, period_load_hours in _scheduled_item_period_load_segments(item, config=config):
            row_number += 1
            ws.append([
                item.task.order_id,
                item.task.demand_source,
                item.task.forecast_month,
                item.task.forecast_week_end,
                item.task.forecast_drawing,
                item.task.forecast_route_source,
                item.task.route_source_order,
                item.task.activity,
                item.task.material,
                item.task.process_text,
                item.task.work_center,
                item.task.resource_group,
                item.task.quantity,
                item.task.unit_hours,
                total_load_hours,
                period_load_hours,
                item.start.strftime("%Y-%m-%d %H:%M"),
                item.end.strftime("%Y-%m-%d %H:%M"),
                item.task.due_date.strftime("%Y-%m-%d %H:%M"),
                (item.task.original_due_date or item.task.due_date).strftime("%Y-%m-%d %H:%M"),
                item.task.due_date.strftime("%Y-%m-%d %H:%M"),
                item.task.urgent_type,
                "是" if item.task.urgent else "否",
                item.task.priority_rank,
                item.task.priority_type,
                item.task.priority_reason,
                "是" if item.task.manual_urgent else "否",
                "是" if item.task.adjusted_to_start_period else "否",
                "是" if item.task.is_hot_surface else "否",
                "是" if include_outsource or item.task.is_outsource else "否",
                "是" if item.task.missing_work_center else "否",
                period,
                _optimization_granularity(config),
                period_span,
                round(period_load_hours if item.task.missing_work_center else 0.0, 3),
                item.analysis_status,
                item.analysis_source,
                item.window_number or "",
                item.window_type,
                item.note,
            ])
            current_written = written_rows + row_number
            if progress is not None and report_started_at is not None and (row_number % 3000 == 0 or row_number == total_detail_rows):
                _emit_progress(progress, "写入报告", current_written, total_report_rows, report_started_at, f"{title} {row_number:,}/{total_detail_rows:,}")
    _format_table(ws)
    return written_rows + total_detail_rows


def _write_order_summary_sheet(
    wb: Workbook,
    items: list[ScheduledOperation],
    order_completion: dict[str, datetime],
    progress: ProgressCallback | None = None,
    report_started_at: float | None = None,
    written_rows: int = 0,
    total_report_rows: int = 0,
) -> int:
    ws = wb.create_sheet("订单负荷汇总")
    ws.append(["订单", "需求日期", "工序数", "厂内工序数", "外协工序数", "厂内总工时", "是否紧急"])
    by_order: dict[str, list[ScheduledOperation]] = {}
    for item in items:
        by_order.setdefault(item.task.order_id, []).append(item)
    sorted_orders = sorted(by_order.items(), key=lambda kv: kv[0])
    for row_number, (order_id, order_items) in enumerate(sorted_orders, start=1):
        due = min(item.task.due_date for item in order_items)
        inhouse_items = [item for item in order_items if not item.task.is_outsource]
        outsource_items = [item for item in order_items if item.task.is_outsource]
        ws.append([
            order_id,
            due.strftime("%Y-%m-%d %H:%M"),
            len(order_items),
            len(inhouse_items),
            len(outsource_items),
            round(sum(item.task.duration_hours for item in inhouse_items), 2),
            "是" if any(item.task.urgent for item in order_items) else "否",
        ])
        current_written = written_rows + row_number
        if progress is not None and report_started_at is not None and (row_number % 1000 == 0 or row_number == len(sorted_orders)):
            _emit_progress(progress, "写入报告", current_written, total_report_rows, report_started_at, f"订单负荷汇总 {row_number:,}/{len(sorted_orders):,}")
    _format_table(ws)
    return written_rows + len(sorted_orders)


def _write_placeholder_due_sheet(
    wb: Workbook,
    rows: list[dict[str, Any]],
    progress: ProgressCallback | None = None,
    report_started_at: float | None = None,
    written_rows: int = 0,
    total_report_rows: int = 0,
) -> int:
    ws = wb.create_sheet("占位交期订单")
    headers = [
        "订单",
        "占位需求行数",
        "占位需求总数量",
        "最早占位交期",
        "最晚占位交期",
        "紧急类型",
        "是否紧急",
        "源文件行号",
        "处理说明",
    ]
    ws.append(headers)
    if not rows:
        ws.append(["无", 0, 0, "", "", "", "否", "", f"未发现 {PLACEHOLDER_DUE_YEAR} 年及以后占位交期订单"])
        _format_table(ws)
        return written_rows + 1

    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        order_id = str(row.get("订单") or "").strip()
        if not order_id:
            continue
        due_date = row.get("占位交期")
        item = grouped.setdefault(order_id, {
            "订单": order_id,
            "占位需求行数": 0,
            "占位需求总数量": 0.0,
            "最早占位交期": due_date,
            "最晚占位交期": due_date,
            "紧急类型": "",
            "是否紧急": False,
            "源文件行号": [],
            "处理说明": "占位交期订单未进入 OR-Tools 产能优化分析",
        })
        item["占位需求行数"] += 1
        item["占位需求总数量"] += _to_number(row.get("数量"), default=0.0)
        if isinstance(due_date, datetime):
            if not isinstance(item["最早占位交期"], datetime) or due_date < item["最早占位交期"]:
                item["最早占位交期"] = due_date
            if not isinstance(item["最晚占位交期"], datetime) or due_date > item["最晚占位交期"]:
                item["最晚占位交期"] = due_date
        item["紧急类型"] = _higher_priority_urgent_type(str(item.get("紧急类型") or ""), str(row.get("紧急类型") or ""))
        item["是否紧急"] = bool(item["是否紧急"] or row.get("是否紧急"))
        if row.get("源文件行号"):
            item["源文件行号"].append(str(row.get("源文件行号")))

    grouped_items = sorted(grouped.values(), key=lambda data: data["订单"])
    for row_number, item in enumerate(grouped_items, start=1):
        earliest = item["最早占位交期"]
        latest = item["最晚占位交期"]
        ws.append([
            item["订单"],
            item["占位需求行数"],
            round(item["占位需求总数量"], 2),
            earliest.strftime("%Y-%m-%d") if isinstance(earliest, datetime) else "",
            latest.strftime("%Y-%m-%d") if isinstance(latest, datetime) else "",
            item["紧急类型"],
            "是" if item["是否紧急"] else "否",
            ", ".join(item["源文件行号"]),
            item["处理说明"],
        ])
        if progress is not None and report_started_at is not None and (row_number % 1000 == 0 or row_number == len(grouped_items)):
            _emit_progress(progress, "写入报告", written_rows + row_number, total_report_rows, report_started_at, f"占位交期订单 {row_number:,}/{len(grouped_items):,}")
    _format_table(ws)
    return written_rows + len(grouped_items)


def _write_dict_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict[str, Any]],
    progress: ProgressCallback | None = None,
    report_started_at: float | None = None,
    written_rows: int = 0,
    total_report_rows: int = 0,
) -> int:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["状态"])
        ws.append(["无"])
        _format_table(ws)
        return written_rows + 1
    headers = list(dict.fromkeys(key for row in rows for key in row.keys()))
    ws.append(headers)
    for row_number, row in enumerate(rows, start=1):
        ws.append([row.get(header, "") for header in headers])
        if progress is not None and report_started_at is not None and (row_number % 1000 == 0 or row_number == len(rows)):
            _emit_progress(progress, "写入报告", written_rows + row_number, total_report_rows, report_started_at, f"{title} {row_number:,}/{len(rows):,}")
    _format_table(ws)
    return written_rows + len(rows)


def _write_missing_mapping_sheet(
    wb: Workbook,
    rows: list[dict[str, Any]],
    progress: ProgressCallback | None = None,
    report_started_at: float | None = None,
    written_rows: int = 0,
    total_report_rows: int = 0,
) -> int:
    ws = wb.create_sheet("缺失映射报告")
    headers = [
        "工序短文本",
        "缺失工序数",
        "涉及订单数",
        "示例订单",
        "示例活动",
        "示例物料",
        "示例源文件行号",
        "处理建议",
    ]
    ws.append(headers)
    if not rows:
        ws.append(["无", 0, 0, "", "", "", "", ""])
        _format_table(ws)
        return written_rows + 1

    grouped: dict[str, dict[str, Any]] = {}
    grouping_started_at = time.perf_counter()
    for row_number, row in enumerate(rows, start=1):
        key = _clean_text(row.get("工序短文本")) or "空工序短文本"
        item = grouped.setdefault(key, {
            "工序短文本": key,
            "缺失工序数": 0,
            "订单集合": set(),
            "示例订单": row.get("订单", ""),
            "示例活动": row.get("活动", ""),
            "示例物料": row.get("物料", ""),
            "示例源文件行号": row.get("源文件行号", ""),
            "处理建议": row.get("处理建议", ""),
        })
        item["缺失工序数"] += 1
        if row.get("订单"):
            item["订单集合"].add(str(row.get("订单")))
        if progress is not None and row_number % 20000 == 0:
            _emit_progress(progress, "汇总缺失映射", row_number, len(rows), grouping_started_at, f"已形成 {len(grouped):,} 个工序短文本")

    grouped_items = sorted(grouped.values(), key=lambda data: (-data["缺失工序数"], data["工序短文本"]))
    for row_number, item in enumerate(grouped_items, start=1):
        ws.append([
            item["工序短文本"],
            item["缺失工序数"],
            len(item["订单集合"]),
            item["示例订单"],
            item["示例活动"],
            item["示例物料"],
            item["示例源文件行号"],
            item["处理建议"],
        ])
        if progress is not None and report_started_at is not None and (row_number % 500 == 0 or row_number == len(grouped_items)):
            _emit_progress(progress, "写入报告", written_rows + row_number, total_report_rows, report_started_at, f"缺失映射报告 {row_number:,}/{len(grouped_items):,}")
    _format_table(ws)
    return written_rows + len(grouped_items)


def _read_workbook_first_sheet(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "gb18030", "utf-8"):
            try:
                return pd.read_csv(path, encoding=encoding, dtype=str, keep_default_na=False)
            except UnicodeDecodeError as exc:
                last_error = exc
        if last_error is not None:
            raise last_error
    return pd.read_excel(path, sheet_name=0)


def _format_table(ws, header_row: int = 1) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for row_idx in range(1, header_row):
        for cell in ws[row_idx]:
            if cell.value is None:
                continue
            cell.font = Font(bold=True, color="1F4E79", size=14 if row_idx == 1 else 11)
            cell.alignment = Alignment(vertical="center")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[header_row].height = max(ws.row_dimensions[header_row].height or 0, 24)
    if ws.max_row <= 5000:
        thin = Side(style="thin", color="D9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center")
        for cell in ws[header_row]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    if ws.max_row > header_row and ws.max_column >= 1:
        table_ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
        safe_title = re.sub(r"[^A-Za-z0-9_]", "", ws.title)
        table_name = f"T_{safe_title[:18]}" if safe_title else f"T_{abs(hash(ws.title)) % 1_000_000}"
        table = Table(displayName=table_name, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        try:
            ws.add_table(table)
        except ValueError:
            pass
    ws.freeze_panes = f"A{header_row + 1}"


def _format_workcenter_heatmap(ws, header_row: int = 3) -> None:
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = title_fill
        cell.font = Font(bold=True, color="1F4E79", size=14)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[2]:
        cell.fill = section_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    for cell in ws[header_row]:
        cell.fill = section_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(header_row + 1, ws.max_row + 1):
        metric = str(ws.cell(row_idx, 2).value or "")
        ws.cell(row_idx, 1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row_idx, 2).alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(3, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if isinstance(cell.value, (int, float)):
                if "负荷率" in metric:
                    cell.number_format = "0.0%"
                    cell.fill = _capacity_heatmap_fill(cell.value)
                    cell.font = Font(
                        color=_heatmap_font_color(cell.value),
                        bold=_heatmap_font_bold(cell.value),
                    )
                else:
                    cell.number_format = "#,##0.0"
            elif "负荷率" in metric and cell.value in (None, ""):
                cell.fill = PatternFill()
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[header_row].height = max(ws.row_dimensions[header_row].height or 0, 24)
    ws.freeze_panes = None


def _capacity_heatmap_fill(value: Any) -> PatternFill:
    try:
        ratio = float(value)
    except (TypeError, ValueError):
        ratio = 0.0
    ratio = max(0.0, ratio)
    if ratio <= 0.25:
        return PatternFill("solid", fgColor="006100")
    if ratio <= 0.75:
        return PatternFill("solid", fgColor="C6EFCE")
    if ratio <= 1.0:
        return PatternFill("solid", fgColor="FCE4D6")
    return PatternFill("solid", fgColor="C00000")


def _heatmap_font_color(value: Any) -> str:
    try:
        ratio = float(value)
    except (TypeError, ValueError):
        ratio = 0.0
    if ratio <= 0.25 or ratio > 1.0:
        return "FFFFFF"
    return "000000"


def _heatmap_font_bold(value: Any) -> bool:
    try:
        return float(value) >= 1.0
    except (TypeError, ValueError):
        return False


def _format_dashboard(ws) -> None:
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="1F4E79")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = title_fill
        cell.font = Font(bold=True, color="1F4E79", size=16)
    for cell in ws[2]:
        cell.font = Font(color="666666", italic=True)
    for row in ws.iter_rows():
        first_value = row[0].value
        if first_value in {"核心指标", "重点瓶颈工作中心", "运行判断"}:
            for cell in row:
                cell.fill = section_fill
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A4"


def _cell_display_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def _text_display_width(text: str) -> float:
    width = 0.0
    for char in text:
        if char in "\r\n\t":
            width += 1.0
            continue
        east_asian_width = unicodedata.east_asian_width(char)
        if east_asian_width in {"F", "W"}:
            width += 2.0
        elif east_asian_width == "A":
            width += 1.5
        else:
            width += 1.0
    return width


def _infer_autosize_header_row(ws) -> int:
    if ws.tables:
        try:
            first_table = next(iter(ws.tables.values()))
            _, min_row, _, _ = range_boundaries(first_table.ref)
            return min_row
        except Exception:
            pass
    if ws.max_row >= 3:
        row1_values = [cell.value for cell in ws[1] if cell.value not in (None, "")]
        row2_values = [cell.value for cell in ws[2] if cell.value not in (None, "")]
        row3_values = [cell.value for cell in ws[3] if cell.value not in (None, "")]
        if len(row3_values) >= 2 and len(row1_values) <= 1 and len(row2_values) <= 1:
            return 3
    return 1


def _sheet_has_table_filter(ws, header_row: int) -> bool:
    if ws.auto_filter and ws.auto_filter.ref:
        try:
            _, min_row, _, _ = range_boundaries(ws.auto_filter.ref)
            if min_row == header_row:
                return True
        except Exception:
            pass
    for table in ws.tables.values():
        try:
            _, min_row, _, _ = range_boundaries(table.ref)
            if min_row == header_row:
                return True
        except Exception:
            continue
    return False


def _column_width_cap(ws, column_index: int) -> float:
    if ws.max_column >= 20 and column_index >= 3:
        return 58.0
    if ws.max_column >= 12 and column_index >= 6:
        return 62.0
    return 78.0


def _recommended_column_width(ws, column_cells, header_row: int) -> float:
    sample_end = min(len(column_cells), 1000)
    sample_start = max(header_row - 1, 0)
    sample = list(column_cells[sample_start:sample_end])
    if not sample:
        sample = list(column_cells[:sample_end])

    max_width = 0.0
    for cell in sample:
        text = _cell_display_text(cell.value)
        if not text:
            continue
        max_width = max(max_width, _text_display_width(text))

    header_cell = ws.cell(header_row, column_cells[0].column)
    header_width = _text_display_width(_cell_display_text(header_cell.value))
    if header_width:
        filter_padding = 7.0 if _sheet_has_table_filter(ws, header_row) else 4.0
        max_width = max(max_width, header_width + filter_padding)

    padding = 3.0
    return min(max(max_width + padding, 12.0), _column_width_cap(ws, column_cells[0].column))


def _autosize_workbook(wb: Workbook) -> None:
    for ws in wb.worksheets:
        header_row = _infer_autosize_header_row(ws)
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[letter].width = _recommended_column_width(ws, column_cells, header_row)


def _missing_columns_issue(df: pd.DataFrame, required: Iterable[str], path: Path) -> list[dict[str, Any]]:
    missing = [column for column in required if column not in df.columns]
    if not missing:
        return []
    return [{"类型": "缺少字段", "文件": str(path), "说明": "缺少字段：" + "、".join(missing)}]


def _normalize_order(value: Any) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("'"):
        text = text[1:].strip()
    formula_match = re.fullmatch(r'="?([^"]+)"?', text)
    if formula_match and text.startswith("="):
        text = formula_match.group(1).strip()
    match = re.fullmatch(r"\d+(?:\.0+)?", text)
    if match:
        return str(int(float(text)))
    sci_match = re.fullmatch(r"[+-]?\d+(?:\.\d+)?[Ee][+-]?\d+", text)
    if sci_match:
        try:
            from decimal import Decimal
            number = Decimal(text)
            if number == number.to_integral_value():
                return str(number.quantize(Decimal(1)))
        except Exception:
            return None
    return None


def _is_numeric_order(value: Any) -> bool:
    return _normalize_order(value) is not None


def _looks_like_scientific_order_id(value: Any) -> bool:
    text = _clean_text(value).strip('"').strip()
    if not text:
        return False
    return bool(re.fullmatch(r"[+-]?\d+(?:\.\d+)?[eE][+-]?\d+", text))


def _clean_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def _is_blank_cell(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    return str(value).strip() == ""


def _activity_key(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    number = _to_number(value, default=math.nan)
    if not math.isnan(number):
        return str(int(number)) if float(number).is_integer() else f"{number:g}"
    return text


def _priority_rank(value: Any) -> int:
    text = _clean_text(value)
    if not text:
        return 999
    if text == "外包":
        return 900
    match = re.search(r"\d+", text)
    if match:
        return int(match.group())
    return 999


def _to_number(value: Any, default: float = 0.0) -> float:
    try:
        if pd.isna(value):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_datetime(value: Any) -> datetime | None:
    if pd.isna(value):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).replace(tzinfo=None)
        except OverflowError:
            return None
    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.\d+)?", text):
        try:
            serial = float(text)
            if 20000 <= serial <= 80000:
                return (datetime(1899, 12, 30) + timedelta(days=serial)).replace(tzinfo=None)
        except (TypeError, ValueError, OverflowError):
            return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime().replace(tzinfo=None)


def _is_placeholder_due_date(value: datetime) -> bool:
    return value.year >= PLACEHOLDER_DUE_YEAR


def _to_bool(value: Any) -> bool:
    text = str(value).strip().lower()
    return text in {"true", "1", "yes", "y", "是", "紧急", "允许"}


def _normalize_urgent_type(value: Any) -> str:
    return _clean_text(value)


def _allowed_urgent_types_text() -> str:
    return "、".join(ALLOWED_URGENT_TYPES)


def _higher_priority_urgent_type(current: str, incoming: str) -> str:
    current = _normalize_urgent_type(current)
    incoming = _normalize_urgent_type(incoming)
    if not current:
        return incoming
    if not incoming:
        return current
    return current if URGENT_TYPE_PRIORITY[current] <= URGENT_TYPE_PRIORITY[incoming] else incoming


def _priority_from_flags(urgent_type: str, overdue: bool) -> tuple[int, str, str]:
    urgent_type = _normalize_urgent_type(urgent_type)
    if urgent_type:
        return (
            URGENT_TYPE_PRIORITY[urgent_type],
            urgent_type,
            f"订单交期数量表手动填写紧急类型：{urgent_type}",
        )
    if overdue:
        return PRIORITY_OVERDUE, "过期订单", "原始供给日期早于优化开始日期"
    return PRIORITY_NORMAL, "普通订单", "未标记紧急且未过期"


def _priority_sort_date(task: OperationTask) -> datetime:
    return task.original_due_date or task.due_date


def _is_hot_surface(process_text: str) -> bool:
    return any(keyword in process_text for keyword in ("热处", "热处理", "表处", "表面处理", "表面"))


def _use_hot_surface_special_logic(config: FortuneBjConfig | None) -> bool:
    text = _clean_text(config.hot_surface_mode if config is not None else "")
    return "专用" in text and "同机加" not in text


def _normalize_capacity_calc_type(value: Any) -> str:
    text = _clean_text(value)
    if any(keyword in text for keyword in ("批", "炉", "箱")):
        return "批量处理"
    if any(keyword in text for keyword in ("流水", "履带", "连续", "线")):
        return "流水线处理"
    return "普通工时"


def _hot_surface_type_from_text(*values: Any) -> str:
    text = " ".join(_clean_text(value) for value in values if _clean_text(value))
    if any(keyword in text for keyword in ("表处", "表面")):
        return "表处"
    if any(keyword in text for keyword in ("热处", "热处理")):
        return "热处"
    return "普通"


def _positive_or_default(value: float, default: float) -> float:
    return float(value) if value and value > 0 else float(default)


def _task_uses_special_capacity(
    task: OperationTask,
    capacity: WorkCenterCapacity | None,
    config: FortuneBjConfig | None,
) -> bool:
    if not _use_hot_surface_special_logic(config) or task.is_outsource or not task.is_hot_surface:
        return False
    calc_type = _normalize_capacity_calc_type(task.capacity_calc_type or (capacity.capacity_calc_type if capacity else ""))
    return calc_type in {"批量处理", "流水线处理"}


def _task_unit_capacity(task: OperationTask, capacity: WorkCenterCapacity | None) -> float:
    return _positive_or_default(task.unit_capacity, capacity.default_unit_capacity if capacity else 1.0)


def _operation_load_hours(
    *,
    quantity: float,
    unit_hours: float,
    task: OperationTask | None = None,
    capacity: WorkCenterCapacity | None = None,
    is_hot_surface: bool = False,
    capacity_calc_type: str = "",
    unit_capacity: float = 1.0,
    config: FortuneBjConfig | None = None,
) -> float:
    if quantity <= 0:
        return 0.01
    if not _use_hot_surface_special_logic(config) or not is_hot_surface or capacity is None:
        return max(quantity * unit_hours, 0.01)
    calc_type = _normalize_capacity_calc_type(capacity_calc_type or capacity.capacity_calc_type)
    effective_unit_capacity = _positive_or_default(unit_capacity, capacity.default_unit_capacity)
    if calc_type == "批量处理":
        batch_capacity = _positive_or_default(capacity.batch_capacity, 0.0)
        cycle_hours = _positive_or_default(capacity.batch_cycle_hours, unit_hours)
        if batch_capacity > 0 and cycle_hours > 0:
            load_units = quantity * effective_unit_capacity
            batch_count = math.ceil(load_units / batch_capacity)
            return max(batch_count * (cycle_hours + max(capacity.setup_hours, 0.0)), 0.01)
    if calc_type == "流水线处理":
        throughput = _positive_or_default(capacity.line_throughput_rate, 0.0)
        if throughput > 0:
            load_units = quantity * effective_unit_capacity
            return max(load_units / throughput + max(capacity.changeover_hours, 0.0), 0.01)
    return max(quantity * unit_hours, 0.01)


def _capacity_profile_for_route(
    task: OperationTask,
    *,
    option: OptionalOperation | None,
    work_center: str,
    unit_hours: float,
    capacities: dict[str, WorkCenterCapacity],
    config: FortuneBjConfig | None,
) -> dict[str, Any]:
    capacity = capacities.get(work_center)
    is_hot_surface = bool(task.is_hot_surface and not task.is_outsource and capacity is not None)
    option_calc_type = option.capacity_calc_type if option is not None else ""
    calc_type = _normalize_capacity_calc_type(option_calc_type or task.capacity_calc_type or (capacity.capacity_calc_type if capacity else ""))
    if not _use_hot_surface_special_logic(config) or not is_hot_surface:
        calc_type = "普通工时"
    unit_capacity = _positive_or_default(
        option.unit_capacity if option is not None else 0.0,
        _task_unit_capacity(task, capacity),
    )
    process_group = (
        (option.process_group if option is not None else "")
        or task.process_group
        or (capacity.process_groups if capacity else "")
        or task.process_text
    )
    effective_unit_hours = max(unit_hours, 0.0)
    if calc_type == "流水线处理" and capacity is not None and capacity.line_throughput_rate > 0:
        effective_unit_hours = unit_capacity / capacity.line_throughput_rate
    return {
        "capacity_calc_type": calc_type,
        "hot_surface_type": task.hot_surface_type or (capacity.hot_surface_type if capacity else "") or _hot_surface_type_from_text(task.process_text),
        "process_group": process_group,
        "unit_capacity": unit_capacity,
        "batch_capacity": capacity.batch_capacity if capacity else 0.0,
        "batch_cycle_hours": (capacity.batch_cycle_hours + max(capacity.setup_hours, 0.0)) if capacity else 0.0,
        "line_throughput_rate": capacity.line_throughput_rate if capacity else 0.0,
        "residence_hours": capacity.residence_hours if capacity else 0.0,
        "unit_hours": effective_unit_hours,
    }
